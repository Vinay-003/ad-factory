#!/usr/bin/env python3
"""Standalone prompt assembler from an on-image-copy xlsx export.

Reads an xlsx file exported from the dashboard's "extract on-image copy"
function, regenerates prompt files with backgrounds, and creates a run
manifest so the web automation scripts (gemini_web_automation.py,
chatgpt_web_sutomation.py) can consume them.

Background strategy:
  - One background per FORMAT, shared across ALL personas.
  - Backgrounds are picked from background_variant.json catalog.

Usage:
    python scripts/assemble_from_xlsx.py --xlsx path/to/on-image-copy.xlsx [--aspect-ratio 4:5] [--batch v99]
"""

from __future__ import annotations

import argparse
import json
import random
import re
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import hashlib
import json
import random
import re
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    raise

ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = ROOT / "output"
BACKGROUNDS_PATH = ROOT / "background_variant.json"
REGISTRY_PATH = ROOT / "AD_GENERATION_REGISTRY.JSON"
STORAGE_ROOT = ROOT / "dashboard_storage"
RUNS_ROOT = STORAGE_ROOT / "runs"


def stable_signature_seed(*parts: Any) -> int:
    joined = "|".join(str(part or "") for part in parts)
    digest = hashlib.sha256(joined.encode("utf-8")).hexdigest()
    return int(digest[:8], 16) or 1


FORMAT_VISUAL_ARCHETYPES: dict[str, list[dict[str, Any]]] = {
    "HERO": [
        {"id": "hero_center_stage", "label": "Centered premium packshot",
         "layout_lines": ["- Archetype: centered premium packshot with headline centered above the product block.", "- Keep the product cluster compact in the middle third with a balanced pedestal feel and strong whitespace around it.", "- Use symmetrical spacing so the ad reads as premium and stable rather than busy or collage-like.", "- Keep both text and product centered; do not introduce obvious left/right asymmetry in this variant."],
         "direction_lines": ["- Archetype direction: centered composition with soft pedestal energy and minimal side distractions.", "- Text placement: headline/support centered in upper safe field; CTA centered below support with generous breathing room.", "- This is the only HERO variant that should read as fully centered and symmetrical."]},
        {"id": "hero_left_copy_right_product", "label": "Left-copy right-product hero",
         "layout_lines": ["- Archetype: left-aligned copy block with product cluster weighted to the right half.", "- Build a clear asymmetric composition where text owns the left safe field and product owns the right safe field.", "- Preserve clean negative space between copy and product so the frame feels intentional, not cramped.", "- Headline, support, and CTA must align left in a clean copy column; do not center the typography in this variant."],
         "direction_lines": ["- Archetype direction: editorial left-copy/right-product split with crisp column alignment.", "- Product emphasis: kit box and product stack dominate the right-center while copy stays in a protected left column.", "- Force visible asymmetry: if the result reads centered at a glance, reject and regenerate."]},
        {"id": "hero_close_crop_pedestal", "label": "Close crop pedestal hero",
         "layout_lines": ["- Archetype: larger, tighter product crop with the pack cluster feeling closer to camera.", "- Let product presence feel slightly oversized while keeping every label readable and fully inside safe field.", "- Copy should sit in a compact high-contrast block occupying one upper corner rather than a wide centered text stack.", "- Crop the product group noticeably larger than other HERO variants so the pack mass dominates the lower half."],
         "direction_lines": ["- Archetype direction: premium close crop with a luxury pedestal feel and shallow scene complexity.", "- Visual balance: oversized product hero, minimal props, restrained typography footprint.", "- Composition should feel tighter and more zoomed-in than the centered or split variants; reject if it reads like a standard balanced packshot."]},
        {"id": "hero_soft_lifestyle_frame", "label": "Lifestyle-assisted premium hero",
         "layout_lines": ["- Archetype: premium hero layout softened by subtle lifestyle context, not pure studio isolation.", "- Keep product dominant while allowing background context cues to add warmth and realism.", "- Headline and support should sit in a calm editorial block with more atmospheric breathing room than strict commercial grid.", "- Product cluster should sit lower-left or lower-right with the opposite upper zone reserved for copy; do not center both elements together."],
         "direction_lines": ["- Archetype direction: premium lifestyle frame with product dominance preserved over the environment.", "- Background behavior: warm believable context adds credibility but never competes with the packshot.", "- This variant must feel more environmental and editorial than the studio-centered HERO options."]},
    ],
    "BA": [
        {"id": "ba_classic_split", "label": "Classic vertical split contrast",
         "layout_lines": ["- Archetype: strict left/right split with a clear vertical divider and fast-scroll readability.", "- Left side must instantly read as friction; right side must instantly read as first practical win.", "- Keep products bridging the lower center so both panels still feel like one branded story."],
         "direction_lines": ["- Archetype direction: disciplined vertical split with highly legible contrast in mood, props, and posture.", "- Contrast behavior: same overall palette family, different clarity and routine signals on each side."]},
        {"id": "ba_soft_diagonal_transition", "label": "Soft diagonal transition",
         "layout_lines": ["- Archetype: transition from struggle to control happens along a soft diagonal flow instead of a hard binary panel wall.", "- Use composition to move the eye from upper-left friction toward lower-right control without losing split-story clarity.", "- Keep the product cluster near lower center as the bridge that unifies the two states."],
         "direction_lines": ["- Archetype direction: diagonal story flow with subtle transition edge and premium editorial polish.", "- Panel behavior: struggle cues and control cues remain distinct even without a thick central divider."]},
        {"id": "ba_desk_to_discipline", "label": "Desk-scene routine contrast",
         "layout_lines": ["- Archetype: same broad routine surface evolves from cluttered impulsive setup to clean repeatable setup.", "- Let props and orderliness carry much of the contrast, not only facial expression.", "- Keep the headline spanning both states while the lower product bridge anchors the solution."],
         "direction_lines": ["- Archetype direction: one routine zone shown in two states, emphasizing practical environmental contrast.", "- Visual proof: clutter-to-clarity progression should feel plausible and grounded, never dramatic or fake."]},
        {"id": "ba_emotion_to_control", "label": "Emotion-to-control contrast",
         "layout_lines": ["- Archetype: subject emotion shift is the primary contrast driver, with supporting routine cues around it.", "- The left panel should feel mentally stuck; the right panel should feel calmer, steadier, and more in control.", "- Product cluster remains stable and premium so the frame does not become too portrait-heavy."],
         "direction_lines": ["- Archetype direction: human emotional contrast leads, product proof anchors, environment stays secondary.", "- Contrast behavior: subtle mood correction rather than exaggerated transformation theatrics."]},
    ],
    "TEST": [
        {"id": "test_editorial_quote_card", "label": "Editorial quote card",
         "layout_lines": ["- Archetype: large editorial quote card is the primary visual element, with strong quote-first hierarchy.", "- Quote card should occupy the upper safe field prominently while product remains a secondary but visible proof anchor below.", "- Use clean premium spacing so the testimonial feels designed, not templated."],
         "direction_lines": ["- Archetype direction: oversized quote card with calm editorial rhythm and a restrained premium card treatment.", "- Product placement: compact bottom-right or bottom-center proof anchor beneath the quote hierarchy."]},
        {"id": "test_portrait_overlay_card", "label": "Portrait with floating review card",
         "layout_lines": ["- Archetype: believable human portrait is more visible, while the review card floats over part of the scene.", "- The quote must still win attention first, but the portrait should noticeably increase authenticity and relatability.", "- Product cluster should remain smaller than in HERO and act as proof, not main subject."],
         "direction_lines": ["- Archetype direction: portrait-led testimonial with floating review panel overlap and controlled depth.", "- Layering: quote card overlaps portrait space without obscuring the face or reducing readability."]},
        {"id": "test_minimal_review_poster", "label": "Minimal review poster",
         "layout_lines": ["- Archetype: very large typography with minimal card chrome, almost like an editorial poster.", "- Reduce decorative panel treatment and let type scale, spacing, and contrast carry the testimonial weight.", "- Product cluster should stay compact and lower in frame so the quote remains the unmistakable hero."],
         "direction_lines": ["- Archetype direction: minimal poster-like layout, premium typography-first attitude, almost no ornamental card styling.", "- Hierarchy: quote dominates first, proof strip and product sit as controlled secondary anchors."]},
        {"id": "test_proof_strip_layout", "label": "Proof-strip testimonial",
         "layout_lines": ["- Archetype: compact quote block with a clearer verification/proof strip and slightly stronger product presence.", "- Trust line should feel like a deliberate proof rail, not just a small subheading.", "- Product cluster can be larger here than other TEST variants, but testimonial hierarchy must still stay primary."],
         "direction_lines": ["- Archetype direction: compact testimonial card plus strong proof strip and polished product confirmation below.", "- Structure: quote block -> attribution -> proof strip -> product cluster -> CTA with clean separation between each layer."]},
    ],
    "FEAT": [
        {"id": "feat_bullet_panel", "label": "Classic bullet panel",
         "layout_lines": ["- Archetype: structured bullet panel on one side and product proof on the opposite side.", "- Bullets should read as a calm functional stack, never as dense paragraph copy.", "- Keep the information panel crisp and balanced against the product cluster."],
         "direction_lines": ["- Archetype direction: classic info-column layout with bullets grouped neatly in one protected text panel.", "- Product placement: kit stack occupies the opposite side as a clear proof anchor."]},
        {"id": "feat_modular_cards", "label": "Modular feature cards",
         "layout_lines": ["- Archetype: features appear as small modular cards or tiles rather than one continuous bullet column.", "- Maintain generous spacing between feature modules so each benefit scans instantly on mobile.", "- Product cluster remains visually stable while feature modules create structured variation around it."],
         "direction_lines": ["- Archetype direction: modular information architecture with 3-4 clearly separated feature tiles.", "- Layout behavior: feature cards can wrap around the product zone without feeling cluttered."]},
        {"id": "feat_mechanism_steps", "label": "Mechanism step stack",
         "layout_lines": ["- Archetype: present features as a clear step-by-step or mechanism ladder instead of isolated bullets.", "- Create a directional reading flow from headline into numbered or sequenced feature logic.", "- Product cluster should support the mechanism story rather than overpower it."],
         "direction_lines": ["- Archetype direction: vertical step stack with ordered reading rhythm and low-noise spacing.", "- Information behavior: each feature line should feel like one stage of a practical routine or benefit chain."]},
        {"id": "feat_callout_annotations", "label": "Annotated product callouts",
         "layout_lines": ["- Archetype: feature lines behave like concise callout annotations pointing toward the product group.", "- Keep annotation lines short and clean so the frame still reads premium, not technical or noisy.", "- Product cluster stays central enough that the callouts feel attached to a real focal object."],
         "direction_lines": ["- Archetype direction: annotated product explainer with short callout labels and disciplined connector logic.", "- Visual balance: callouts orbit the product safely without crowding the safe zones."]},
    ],
    "UGC": [
        {"id": "ugc_selfie_hold", "label": "Selfie beside-product frame",
         "layout_lines": ["- Archetype: creator appears beside the kit in a selfie-like composition while products rest on a surface at correct scale.", "- The frame should feel personal and immediate without losing product label readability.", "- Text should stay compact and social-native, with clear protected space around the face and product."],
         "direction_lines": ["- Archetype direction: near-product selfie shot with intimate creator energy and stable phone realism, but products remain placed on a table/counter.", "- Subject behavior: easy facial expression and optional open-palm gesture toward the product, never holding or gripping any package."]},
        {"id": "ugc_desk_review", "label": "Desk review setup",
         "layout_lines": ["- Archetype: creator is seated or standing by a desk/surface, presenting the kit like a practical recommendation.", "- Product cluster sits on the routine surface while one hand or gesture supports explanation.", "- Text can occupy a cleaner side of the frame as if added to a real review moment."],
         "direction_lines": ["- Archetype direction: review-at-desk composition with believable routine props and calm speaking posture.", "- Product behavior: some products rest on the desk while the kit box stays clearly visible in the creator zone."]},
        {"id": "ugc_morning_routine", "label": "Morning routine scene",
         "layout_lines": ["- Archetype: morning-routine context with subtle vanity, counter, or getting-ready cues.", "- The scene should feel like a real use moment, not a polished set or stock ad template.", "- Keep text lighter and more integrated so the environment can carry authenticity."],
         "direction_lines": ["- Archetype direction: morning routine realism with warm, soft, believable domestic context.", "- Environment behavior: lightly styled vanity or counter cues, never cluttered or glamorous."]},
        {"id": "ugc_unboxing_reaction", "label": "Soft unboxing / discovery moment",
         "layout_lines": ["- Archetype: creator appears near an already-opened kit in a first-impression style frame, with all products placed on the surface.", "- No product is held closer to camera; the full set remains grouped visibly at real-world package scale.", "- Preserve premium cleanliness so the ad feels authentic but not sloppy."],
         "direction_lines": ["- Archetype direction: discovery moment with a pointing/open-palm reaction beside the arranged pack, not a hand-held reveal.", "- Subject behavior: natural curiosity and calm approval, no exaggerated reaction faces."]},
    ],
}


def now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def next_batch_name() -> str:
    if not OUTPUT_DIR.exists():
        return "v1"
    batches = []
    for child in OUTPUT_DIR.iterdir():
        if child.is_dir():
            m = re.match(r"^v(\d+)$", child.name)
            if m:
                batches.append(int(m.group(1)))
    return "v1" if not batches else f"v{max(batches) + 1}"


def make_run_id() -> str:
    return f"run_{int(time.time())}_{random.randint(1000, 9999)}"


def load_backgrounds() -> dict[str, Any]:
    return json.loads(BACKGROUNDS_PATH.read_text(encoding="utf-8"))


def pick_one_background_per_format(backgrounds: dict[str, Any], formats: set[str]) -> dict[str, dict[str, Any]]:
    """Pick one background per format from the catalog."""
    variants = backgrounds.get("variants", [])
    result = {}
    for fmt in formats:
        pool = [v for v in variants if fmt in (v.get("formats") or [])]
        if not pool:
            raise RuntimeError(f"No background variants found for format {fmt}")
        result[fmt] = pool[0]
    return result


def build_seeded_background_sentence(bg: dict[str, Any], seed: int, aspect_ratio: str) -> str:
    rng = random.Random(seed)
    base = bg["base"]
    lighting = rng.choice(bg["lighting"])
    surface = rng.choice(bg["surface"])
    environment = rng.choice(bg["environment"])
    mood = rng.choice(bg["mood"])
    camera = rng.choice(bg["camera"])
    color_tone = rng.choice(bg["color_tone"])
    composition = rng.choice(bg.get("composition") or ["balanced feed composition inside the central safe field"])
    layout_intent = rng.choice(bg.get("layout_intent") or ["preserve a stable center-of-interest corridor"])
    cta_safe_space = rng.choice(bg.get("cta_safe_space") or ["maintain subtle low-contrast space near the lower edge"])
    crop_safety = rng.choice(bg.get("crop_safety") or ["maintain protected margin buffers"])
    text_overlay_treatment = rng.choice(
        bg.get("text_overlay_treatment")
        or ["if a text readability panel is used, keep it in the upper text zone only"]
    )
    edge_tone_control = rng.choice(
        bg.get("edge_tone_control")
        or ["keep all frame edges tonally neutral with no orange, amber, or sepia cast"]
    )

    if aspect_ratio == "9:16":
        format_clause = (
            "designed for 9:16 vertical placement with key subject content constrained to the 14-65 percent safe band"
        )
    else:
        format_clause = (
            "designed for 4:5 feed framing with key content held inside the central safe field"
        )

    return (
        f"{base} on a {surface}, with {environment}, lit by {lighting}, conveying {mood}; "
        f"{camera}, {composition}, {layout_intent}, {cta_safe_space}, {crop_safety}, {text_overlay_treatment}, {edge_tone_control}, {color_tone}, "
        f"{format_clause}, clean premium studio ad photography, ultra-detailed, flawless commercial finish."
    )


def build_ugc_subject_line(seed: int) -> str:
    rng = random.Random(seed)
    age_band = rng.choice(["24-29", "27-33", "30-36", "32-38"])
    tone = rng.choice(["calm confident", "warm assured", "focused optimistic", "grounded and practical"])
    attire = rng.choice(["solid kurta", "casual cotton shirt", "minimal blouse", "everyday premium casual wear"])
    hair = rng.choice(["neat tied-back hair", "simple ponytail", "natural shoulder-length hair", "clean center-part tied style"])
    return f"Indian woman {age_band}, {tone} expression, {attire}, {hair}; natural and unposed."


def safezone_enforcement_block(aspect_ratio: str) -> str:
    if aspect_ratio == "9:16":
        return (
            "SAFE-ZONE ENFORCEMENT (NON-NEGOTIABLE)\n"
            "- Keep all critical content inside the 14%-65% vertical safe band.\n"
            "- Reserve the lower 35% as quiet space; no product labels or headline copy in this band.\n"
            "- Keep side margins clean; no key content touching left or right frame edges.\n"
            "- Product cluster may be centered or left/right weighted according to the selected archetype, but must remain fully inside safe zones.\n"
            "- Reject and regenerate if any headline, CTA, or product detail crosses restricted zones."
        )
    return (
        "SAFE-ZONE ENFORCEMENT (NON-NEGOTIABLE)\n"
        "- Frame: 1080x1350 (4:5).\n"
        "- Restricted bands: top 10% (0-135px), bottom 15% (1148-1350px), side edges outer 8% (0-86px and 994-1080px).\n"
        "- Keep all products and all on-image text fully inside the central safe field: x=86-994 and y=135-1148.\n"
        "- Product cluster may be centered or left/right weighted according to the selected archetype, with mild upward bias allowed, but must not touch any restricted band.\n"
        "- Reject and regenerate if any headline, CTA, or product detail crosses restricted zones."
    )


def outpaint_lock_block(aspect_ratio: str) -> str:
    if aspect_ratio != "9:16":
        return ""
    return (
        "9:16 CONVERSION LOCK (OUTPAINT ONLY)\n"
        "- Treat base 4:5 composition as fixed ground truth; preserve it exactly.\n"
        "- Extend canvas vertically only; top/bottom extension zones are background only.\n"
        "- Do not stretch, warp, zoom, crop, or recomposite existing content.\n"
        "- Keep product cluster size, spacing, and relative proportions identical to base 4:5.\n"
        "- Keep product cluster anchored at roughly same visual vertical position (~45%).\n"
        "- Keep headline/support/CTA hierarchy and spacing tight; do not add vertical gaps.\n"
        "- Keep all critical text+product content inside the central active band; do not push into extension zones.\n"
        "- If any distortion, spacing drift, or repositioning appears, reject and regenerate."
    )


def base_layout_lines_for_format(fmt: str) -> list[str]:
    if fmt == "HERO":
        return [
            "- HERO format: strong headline, one support line, and one CTA.",
            "- Focal hierarchy: product dominant, text secondary, background tertiary.",
            "- Product zone: all key pack details stay away from edge-risk zones, but left/right weighting must follow the selected archetype rather than defaulting to a centered stack.",
            "- CTA must be rendered as a filled rounded button chip (high-contrast), never as plain text.",
            "- Camera framing should feel stable, premium, and label-safe.",
            "- Do not collapse every HERO into centered text over centered products unless the selected archetype explicitly requires a centered composition.",
        ]
    if fmt == "BA":
        return [
            "- Format: BA.",
            "- Build a clear struggle-to-progress contrast story without literal BEFORE/AFTER labels on-image.",
            "- Contrast must be visible in scene, props, posture, and lighting, not text alone.",
            "- Keep products grouped near lower center as the bridge between both states.",
            "- CTA must be rendered as a filled rounded button chip centered in lower safe band, never plain text.",
        ]
    if fmt == "TEST":
        return [
            "- Format: TEST.",
            "- This must read testimonial-first, not HERO with a person added.",
            "- Quote, attribution, and trust proof hierarchy must be obvious on first scroll.",
            "- Human presence can support credibility, but the testimonial message remains primary.",
            "- CTA must be rendered as a filled rounded button chip in lower safe band, never plain text.",
        ]
    if fmt == "FEAT":
        return [
            "- Format: FEAT.",
            "- Build a clean information hierarchy with headline, 3-4 concise feature points, and one CTA.",
            "- Product cluster must stay fully visible as proof while information remains fast to scan.",
            "- Bullets or callouts must be functional benefits only; short, concrete, and readable.",
            "- CTA must be rendered as a filled rounded button chip, never plain text.",
        ]
    if fmt == "UGC":
        return [
            "- Format: UGC.",
            "- Creator-style authenticity with premium cleanliness; avoid stock-template look.",
            "- Subject and product should feel naturally integrated into a believable routine moment.",
            "- Headline/support/context/CTA stack must stay compact and mobile-readable.",
            "- Hands must look anatomically correct; no extra fingers or warped nails.",
        ]
    raise RuntimeError(f"Unsupported format: {fmt}")



def pick_visual_archetype(fmt: str, persona_number: int, headline: str, seed: int) -> dict[str, Any]:
    variants = FORMAT_VISUAL_ARCHETYPES.get(fmt, [])
    if not variants:
        raise RuntimeError(f"No visual archetypes configured for format {fmt}")
    selector_seed = stable_signature_seed(fmt, persona_number, seed, headline)
    rng = random.Random(selector_seed)
    return variants[rng.randrange(len(variants))]


def render_prompt(
    fmt: str,
    lang: str,
    aspect_ratio: str,
    persona_name: str,
    persona_number: int,
    persona_pain: str,
    persona_desire: str,
    persona_friction: str,
    persona_proof: str,
    persona_tone: str,
    awareness_stage: str,
    concept_angle: str,
    concept_structure: str,
    exact_on_image_copy_block: str,
    bg: dict[str, Any],
    bg_seed: int,
    seeded_sentence: str,
    visual_archetype: dict[str, Any],
) -> str:
    if fmt == "HERO":
        style = "HERO, polished enough for paid ad deployment."
    elif fmt == "BA":
        style = "BA (before/after journey without body-shaming visuals)."
    elif fmt == "TEST":
        style = "TEST (trust-first testimonial/review framing)."
    elif fmt == "FEAT":
        style = "FEAT (features and mechanism clarity)."
    elif fmt == "UGC":
        style = "UGC (creator-style authenticity, premium and clean)."
    else:
        raise RuntimeError(f"Unsupported format: {fmt}")

    pain = persona_pain or f"Pain point for {persona_name}"
    desire = persona_desire or f"Desired outcome for {persona_name}"
    friction = persona_friction or f"Friction for {persona_name}"
    proof = persona_proof or f"Proof needed for {persona_name}"
    tone = persona_tone or "calm and confident"

    layout_lines = base_layout_lines_for_format(fmt) + list(visual_archetype.get("layout_lines") or [])
    archetype_direction_lines = [str(line) for line in (visual_archetype.get("direction_lines") or []) if isinstance(line, str) and line.strip()]

    # Parse the exact_on_image_copy_block into copy_lines
    copy_lines: list[str] = []
    for raw_line in exact_on_image_copy_block.splitlines():
        line = raw_line.strip()
        if line:
            copy_lines.append(line)

    if fmt == "UGC":
        subject_line = build_ugc_subject_line(bg_seed)
        action_line = "Products arranged on a clean table/counter at correct scale; creator may point or gesture near them but must not touch, hold, grab, lift, or open any product or box."
        camera_line = "Handheld close-to-medium framing, phone-like realism, stable focus on labels."
        realism_line = "True-to-life proportions; no stock-template look; natural skin and correct hand anatomy."
    elif fmt == "BA":
        subject_line = "Same adult subject identity appears in both panels: left shows routine struggle, right shows practical control; no body-shaming or body-morph visuals."
        action_line = "Split action: BEFORE panel shows rushed/impulsive routine cue; AFTER panel shows one clear repeatable step with cleaner setup and calmer behavior."
        camera_line = "Eye-level medium framing with clean edge discipline."
        realism_line = "True-to-life proportions; no stock-template look."
    else:
        subject_line = "No human subject, products only."
        action_line = "Arrange all 5 products as a cohesive premium cluster; kit box acts as anchor."
        camera_line = "Eye-level medium framing with clean edge discipline."
        realism_line = "True-to-life proportions; no stock-template look."

    lines: list[str] = []
    canvas_spec = "1080 x 1920" if aspect_ratio == "9:16" else "1080 x 1350"
    lines.append("PRODUCT LOCK BLOCK")
    lines.extend([
        "- Use the uploaded Obesity Killer product packshot images as absolute visual truth.",
        "- Use provided product references as exact appearance truth for pack shape, label, and color.",
        "- Do not redesign, redraw, relabel, or alter any product or packaging in any way.",
        "- Do not change brand name, label colors, illustrations, proportions, or any text (Hindi or English).",
        "- If any label text is unclear, preserve the original image as-is. Do not reinterpret it.",
        "- Only permitted: placement, scaling, subtle drop shadows, mild warm lighting correction.",
    ])
    lines.append("")
    lines.append("OUTPUT SPEC")
    lines.extend([
        f"- Canvas: {canvas_spec} pixels. Portrait. {aspect_ratio} ratio.",
        f"- Style: {style}",
        f"- Visual archetype: {visual_archetype['id']} ({visual_archetype['label']})",
        "- Full-bleed requirement: scene must reach all canvas edges; no inset poster card, no inner frame, no side matte bands, no faux border margins.",
        "- Text policy: low text by default; all copy minimal and mobile-readable at 375px width.",
        "- Rendering: no compression artifact; no soft edges on text or product labels.",
        "- All 5 products present, proportionally sized per reference dimensions, unmodified.",
        "- Readability: maintain high-contrast foreground/background treatment for ad-platform legibility.",
    ])
    lines.append("")
    lines.append("FORMAT LAYOUT INSTRUCTIONS")
    lines.extend(layout_lines)
    lines.append("")
    lines.append("PERSONA INPUT BLOCK")
    lines.extend([
        f"- Persona: {persona_name} (Persona {persona_number})",
        f"- Pain: {pain}",
        f"- Desire: {desire}",
        f"- Friction: {friction}",
        f"- Proof needed: {proof}",
        f"- Tone cue: {tone}",
        f"- Awareness stage: {awareness_stage}",
        f"- Concept angle: {concept_angle}",
        f"- Concept structure: {concept_structure}",
        "- Concept path is strategy only; do not render these labels on-image.",
    ])
    lines.append("")
    lines.append("EXACT ON-IMAGE COPY - DO NOT ALTER ANYTHING")
    lines.extend(copy_lines)
    lines.append("Render every character exactly as written. No paraphrasing, no punctuation changes, no autocorrection.")
    lines.append("")
    lines.append("NEGATIVE CONSTRAINTS")
    negative = [
        "- Do not recreate or redraw any product.",
        "- Do not blur, approximate, or rewrite any label text.",
        "- Do not use sale badges, burst graphics, or stickers.",
        "- Do not show body transformations or weight-loss visuals.",
        "- Do not use colors outside the defined palette.",
        "- Do not use more than 2 font weights.",
        "- Do not overcrowd the layout.",
        "- Do not add edge glows, tinted borders, amber/orange side casts, or decorative vignette frames.",
        "- Do not create inset-canvas look: no white frame border, no inner-card composition, no side fade bands, no outer padding effect.",
        "- Do not shrink main scene inside margins; composition must remain full-bleed to all four edges.",
        "- Do not place any translucent white panel behind or below the product cluster.",
        "- Do not make medical cure claims of any kind.",
        "- Do not use ring light, studio flash, or overproduced lighting.",
    ]
    if fmt == "UGC":
        negative.insert(8, "- Do not render unnatural or anatomically incorrect hands.")
        negative.insert(9, "- Do not show the creator holding, grabbing, lifting, opening, pinching, or touching any product, bottle, packet, jar, or kit box; products must rest independently on the surface at correct scale.")
    if fmt == "BA":
        negative.append("- Do not render literal BEFORE:/AFTER: words anywhere on-image.")
    lines.extend(negative)
    lines.append("")
    lines.append("QUALITY BAR - verify before accepting output")
    lines.extend([
        "- All 5 products present, correctly proportioned, and completely unmodified.",
        "- All on-image text sharp and readable at 375px mobile size.",
        "- Product labels accurate, unmodified, and fully readable.",
        "- Layout calm, balanced, and premium.",
        "- No clutter, no hype, and no forbidden elements.",
        "- Single clear focal hierarchy aligned with the selected visual archetype.",
        "- If any item above fails, regenerate immediately without compromise.",
    ])
    lines.append("")
    lines.append("VISUAL DIRECTION BLOCK")
    lines.extend([
        f"- Background slot: {bg['id']} - {bg.get('title', '').strip() or 'Catalog background'}",
        f"- Background seed: {bg_seed}",
        f"- Seeded background direction (single sentence, exact): {seeded_sentence}",
        f"- Subject: {subject_line}",
        f"- Action: {action_line}",
        f"- Camera: {camera_line}",
        f"- Lighting: Warm, soft, directional (top-left) and label-safe.",
        f"- Props: Minimal, non-competing; keep edge zones quiet.",
        f"- Surfaces: Premium, clean texture; avoid busy patterns under text.",
        "- Text panel treatment: if used, keep panel in upper text zone only with vertical fade to transparent before products; never place white panel behind or below product cluster.",
        "- Edge color control: keep border tones neutral and natural; no orange/amber edge tint and no ornamental frame glow.",
        "- Edge structure control: enforce full-bleed composition to edges; reject outputs with visible inset border, side matte strip, or inner-frame card look.",
        f"- Mood: Calm confidence and practical consistency; no hype.",
        f"- Realism: {realism_line}",
        f"- Selected visual archetype: {visual_archetype['id']} - {visual_archetype['label']}",
    ])
    if fmt == "HERO":
        lines.append("- HERO anti-convergence rule: obey the selected archetype literally. Do not fall back to generic centered headline + centered product composition unless the chosen archetype is the centered variant.")
    lines.extend(archetype_direction_lines)
    if fmt == "BA":
        lines.extend([
            "- BEFORE panel visual anchors: include at least 2 struggle cues (messy surface, unplanned snack context, low-clarity routine signals).",
            "- AFTER panel visual anchors: include at least 2 control cues (clean setup, fixed-step context, calmer brighter environment).",
            "- Identity continuity: if a person is shown, keep same person across both panels; change only mood/context, never fake body transformation.",
        ])
    lines.append("")
    lines.append("TYPOGRAPHY SHARPNESS BLOCK")
    lines.extend([
        "- Headline: Poppins Bold with high contrast against clean background area.",
        "- Support and CTA: Poppins Medium/Regular, same family.",
        "- Size: readable on a 375px mobile screen without zooming.",
        "- Placement: flat uncluttered zones only; avoid noisy textures.",
        "- Forbidden: thin fonts, script fonts, decorative typefaces, glow effects, outlined text, drop shadows on copy.",
        "- Mandatory: crisp hard text edges, zero softness, zero anti-alias blur on any character.",
        "- If any text is soft, blurry, or illegible, discard and regenerate immediately.",
    ])
    lines.append("")
    lines.append("Keep on-image text minimal and mobile-readable. Avoid dense copy blocks.")
    lines.append("Render CTA as a real filled button chip with rounded corners and strong contrast; never show CTA as standalone plain text.")
    lines.append("Typography must be pin-sharp. If any text appears soft, blurry, or smeared, regenerate.")
    lines.append("Keep text count minimal and increase font size rather than packing more copy.")
    lines.append("Use clean sans typography with strong stroke clarity; no thin/light weights for body text.")
    lines.append("Use Poppins only for on-image text: Headline in Poppins Bold, support/CTA in Poppins Medium/Regular.")
    lines.append("")
    lines.append(safezone_enforcement_block(aspect_ratio))
    lock_block = outpaint_lock_block(aspect_ratio)
    if lock_block:
        lines.append("")
        lines.append(lock_block)
    lines.append("")
    return "\n".join(lines).strip() + "\n"


def parse_xlsx(xlsx_path: Path) -> list[dict[str, Any]]:
    """Parse the xlsx and return rows as dicts."""
    wb = load_workbook(xlsx_path)
    ws = wb.active
    header = [str(cell.value or "").strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for excel_row in ws.iter_rows(min_row=2):
        values = [cell.value for cell in excel_row]
        row = {}
        for i, col in enumerate(header):
            row[col] = str(values[i]) if values[i] is not None else ""
        if row.get("prompt_id"):
            rows.append(row)
    return rows


def prompt_filename(fmt: str, persona_number: int, lang: str, creative_index: int = 1, creative_total: int = 1) -> str:
    suffix = f"_A{creative_index:02d}" if creative_total > 1 else ""
    return f"OUTPUT_{fmt}_P{persona_number:02d}_{lang}{suffix}.txt"


def aspect_ratio_folder(aspect_ratio: str) -> str:
    return "96" if aspect_ratio == "9:16" else "45"


def main() -> int:
    parser = argparse.ArgumentParser(description="Assemble prompt files from on-image-copy xlsx export")
    parser.add_argument("--xlsx", required=True, help="Path to on-image-copy xlsx file")
    parser.add_argument("--aspect-ratio", default="4:5", choices=["4:5", "9:16"], help="Aspect ratio for prompts")
    parser.add_argument("--batch", default="", help="Batch name like v99 (default: next available)")
    parser.add_argument("--lang", default="EN", choices=["EN", "HI"], help="Prompt language")
    parser.add_argument("--dry-run", action="store_true", help="Print plan without writing files")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx).expanduser()
    if not xlsx_path.exists():
        print(f"ERROR: xlsx file not found: {xlsx_path}")
        return 1

    aspect_ratio = args.aspect_ratio
    lang = args.lang
    batch_name = args.batch or next_batch_name()
    batch_dir = OUTPUT_DIR / batch_name
    ratio_dir = batch_dir / aspect_ratio_folder(aspect_ratio)

    rows = parse_xlsx(xlsx_path)
    if not rows:
        print("ERROR: No data rows found in xlsx")
        return 1

    # Group rows by (format, persona_number) to handle multipliers
    from collections import OrderedDict
    groups = OrderedDict()
    for row in rows:
        prompt_id = row.get("prompt_id", "").strip()
        persona_match = re.search(r"P(\d+)", prompt_id)
        persona_number = int(persona_match.group(1)) if persona_match else None
        fmt = row.get("format", "").strip().upper()
        key = (fmt, persona_number)
        groups.setdefault(key, []).append(row)

    # Flatten back with creative_index assigned per group
    deduped_rows = []
    for key, group_rows in groups.items():
        for idx, row in enumerate(group_rows):
            row["_creative_index"] = idx + 1
            row["_creative_total"] = len(group_rows)
            deduped_rows.append(row)
    rows = deduped_rows

    # Collect unique formats
    formats = set()
    for row in rows:
        fmt = row.get("format", "").strip().upper()
        if fmt:
            formats.add(fmt)
    if not formats:
        print("ERROR: No format values found in xlsx")
        return 1

    print(f"Batch: {batch_name}")
    print(f"Aspect ratio: {aspect_ratio}")
    print(f"Language: {lang}")
    print(f"Formats: {sorted(formats)}")
    print(f"Rows: {len(rows)}")
    print()

    # Pick one background per format
    backgrounds = load_backgrounds()
    fmt_bg = pick_one_background_per_format(backgrounds, formats)
    for fmt, bg in sorted(fmt_bg.items()):
        print(f"  Background for {fmt}: {bg['id']} - {bg.get('title', '')}")
    print()

    # Generate background sentences (one seed per format)
    fmt_bg_seed = {}
    fmt_seeded_sentence = {}
    for fmt in formats:
        bg_seed = random.randint(1, 2_147_483_647)
        fmt_bg_seed[fmt] = bg_seed
        fmt_seeded_sentence[fmt] = build_seeded_background_sentence(fmt_bg[fmt], bg_seed, aspect_ratio)

    if args.dry_run:
        print("Dry run complete. No files written.")
        return 0

    # Create directories
    ratio_dir.mkdir(parents=True, exist_ok=True)

    # Create run directory
    run_id = make_run_id()
    run_dir = RUNS_ROOT / run_id
    (run_dir / "context").mkdir(parents=True, exist_ok=True)
    (run_dir / "inputs").mkdir(parents=True, exist_ok=True)
    (run_dir / "logs").mkdir(parents=True, exist_ok=True)

    # Track persona numbers for manifest
    persona_numbers = set()
    prompt_files = []

    # Build copy_batch.json from xlsx data
    copy_batch_ads = []

    for idx, row in enumerate(rows):
        fmt = row.get("format", "").strip().upper()
        persona_name = row.get("persona_name", "").strip()
        persona_pain = row.get("persona_pain", "").strip()
        persona_desire = row.get("persona_desire", "").strip()
        persona_friction = row.get("persona_friction", "").strip()
        persona_proof = row.get("persona_proof", "").strip()
        persona_tone = row.get("persona_tone", "").strip()
        awareness_stage = row.get("persona_awareness_stage", "").strip() or row.get("awareness_stage", "").strip()
        concept_angle = row.get("concept_angle", "").strip()
        concept_structure = row.get("concept_structure", "").strip()
        exact_block = row.get("exact_on_image_copy_block", "").strip()
        headline_copy = row.get("headline_copy", "").strip()

        if not fmt:
            continue

        # Extract persona number from prompt_id
        prompt_id = row.get("prompt_id", "").strip()
        persona_match = re.search(r"P(\d+)", prompt_id)
        persona_number = int(persona_match.group(1)) if persona_match else (idx + 1)
        persona_numbers.add(persona_number)

        if not persona_name:
            persona_name = f"Persona {persona_number}"

        bg = fmt_bg[fmt]
        bg_seed = fmt_bg_seed[fmt]
        seeded_sentence = fmt_seeded_sentence[fmt]

        # Pick visual archetype
        visual_archetype = pick_visual_archetype(fmt, persona_number, headline_copy, bg_seed)

        # Render prompt
        prompt_text = render_prompt(
            fmt=fmt,
            lang=lang,
            aspect_ratio=aspect_ratio,
            persona_name=persona_name,
            persona_number=persona_number,
            persona_pain=persona_pain,
            persona_desire=persona_desire,
            persona_friction=persona_friction,
            persona_proof=persona_proof,
            persona_tone=persona_tone,
            awareness_stage=awareness_stage,
            concept_angle=concept_angle,
            concept_structure=concept_structure,
            exact_on_image_copy_block=exact_block,
            bg=bg,
            bg_seed=bg_seed,
            seeded_sentence=seeded_sentence,
            visual_archetype=visual_archetype,
        )

        creative_index = row.get("_creative_index", 1)
        creative_total = row.get("_creative_total", 1)

        filename = prompt_filename(fmt, persona_number, lang, creative_index, creative_total)
        out_path = ratio_dir / filename
        out_path.write_text(prompt_text, encoding="utf-8")

        # Write metadata sidecar
        prompt_meta = {
            "type": "ad_prompt",
            "format": fmt,
            "persona": f"P{persona_number:02d}",
            "persona_number": persona_number,
            "persona_name": persona_name,
            "language": lang,
            "aspect_ratio": aspect_ratio,
            "creative_index": creative_index,
            "creative_total": creative_total,
            "multiplier": creative_total,
            "background": {
                "slot": bg["id"],
                "name": bg.get("title", ""),
                "source": "catalog",
                "seed": bg_seed,
                "seeded_direction": seeded_sentence,
            },
            "visual_archetype": {
                "id": visual_archetype["id"],
                "label": visual_archetype["label"],
            },
        }
        out_path.with_suffix(".json").write_text(json.dumps(prompt_meta, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

        rel_path = str(out_path.relative_to(ROOT)).replace("\\", "/")
        prompt_files.append(rel_path)

        # Build copy_batch entry
        copy_batch_ads.append({
            "format": fmt,
            "persona": {
                "number": persona_number,
                "name": persona_name,
                "persona_name": persona_name,
            },
            "awareness_stage": awareness_stage,
            "concept_angle": concept_angle,
            "concept_structure": concept_structure,
            "hypothesis": {
                "type": row.get("hypothesis_type", "").strip(),
                "variant": row.get("hypothesis_variant", "").strip(),
            },
            "creative_index": creative_index,
            "creative_total": creative_total,
        })

        print(f"  Wrote: {rel_path}")

    # Write copy_batch.json
    copy_batch = {"ads": copy_batch_ads}
    copy_batch_path = run_dir / "context" / "copy_batch.json"
    copy_batch_path.write_text(json.dumps(copy_batch, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    # Write manifest
    manifest = {
        "run_id": run_id,
        "batch": batch_name,
        "prompt_files": prompt_files,
        "image_files": [],
        "image_generated": False,
        "copy_source": "xlsx import",
        "aspect_ratio": aspect_ratio,
        "language": lang,
        "created_at": now_iso(),
        "background_reuse": {
            "strategy": "one_per_format",
            "format_backgrounds": {fmt: {"id": bg["id"], "title": bg.get("title", "")} for fmt, bg in fmt_bg.items()},
        },
    }
    manifest_path = run_dir / "manifest.json"
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    print()
    print(f"Run ID: {run_id}")
    print(f"Batch: {batch_name}")
    print(f"Prompt files: {len(prompt_files)}")
    print(f"Manifest: {manifest_path}")
    print(f"Copy batch: {copy_batch_path}")
    print()
    print("Next: run gemini_web_automation.py or chatgpt_web_sutomation.py with:")
    print(f"  --prompt-dir {ratio_dir}")
    print(f"  --prompt-glob 'OUTPUT_*_P*_{lang}.txt'")
    print(f"  --out-dir {ROOT / 'generated_images' / batch_name}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
