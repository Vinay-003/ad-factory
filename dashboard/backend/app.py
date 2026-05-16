#!/usr/bin/env python3

from __future__ import annotations

import json
import os
import random
import re
import subprocess
import sys
import time
import urllib.request
import fcntl
import hashlib
import importlib.util
import mimetypes
import uuid
import psutil
from contextlib import contextmanager
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterator

from fastapi import Body, FastAPI, File, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles


ROOT = Path(__file__).resolve().parents[2]
STORAGE_ROOT = ROOT / "dashboard_storage"
RUNS_ROOT = STORAGE_ROOT / "runs"
RUNTIME_ROOT = ROOT / "runtime"
ENV_PATH = ROOT / ".env.dashboard"

DEFAULT_PRODUCT_MASTER = ROOT / "input" / "docs" / "product master doc.txt"
DEFAULT_PLAYBOOK = ROOT / "AD_CREATIVE_SYSTEM_PLAYBOOK.md"
DEFAULT_IMAGE_SOURCES_FILE = ROOT / "input" / "image_sources.txt"
LEGACY_ACTIVE_IMAGES_FILE = ROOT / "input" / "activeimages.txt"
INPUT_IMAGES_DIR = ROOT / "input" / "images"
GENERATED_IMAGES_ROOT = ROOT / "generated_images"
CONVERT_916_TEMPLATE_PATH = ROOT / "input" / "prompt_916_from_45.txt"
PERSONA_SEEDS_PATH = ROOT / "persona_seeds.json"
COPY_ARCH_PATH = ROOT / "dashboard" / "backend" / "copy_architecture.json"
COPY_PROMPTS_PATH = ROOT / "dashboard" / "backend" / "copy_prompt_templates.json"

DEFAULT_916_CONVERSION_PROMPT = """Convert the attached 4:5 ad creative into a 9:16 version for the same campaign.

Critical fidelity rules:
- Treat the attached 4:5 creative as the single source of truth.
- Product images are locked: preserve the exact same product packshots, label text, logos, illustrations, caps, and colors as the 4:5 reference.
- Allowed product changes are limited to orientation, placement, and scale needed for 9:16 composition.
- Do NOT redraw, reconstruct, relabel, retouch, or restyle any product.
- If any product differs from the source (text distortion, shape change, label change, color shift, missing/extra product), reject and regenerate automatically.
- Do NOT rewrite copy, do NOT invent new claims, do NOT add new offers, and do NOT add new objects.
- Do NOT perform a simple resize/crop/stretch of the original image.

Composition rules for 9:16:
- Extend canvas vertically to 9:16 by outpainting above/below the original composition.
- Keep the original 4:5 core composition stable in the center region.
- Add natural background continuation only where needed to fit 9:16.

Meta 9:16 safe-zone rules (for mobile feed/reels overlays):
- Canvas is 1080x1920 (9:16).
- STRICT META/REELS SAFE FIELD: x=86 to x=994, y=270 to y=1248. All headline text, support text, CTA buttons, offer text, logos, badges, and readable product-label text must sit fully inside this field.
- PRODUCT SAFE FIELD: all product packshots and human faces/hands must remain fully visible inside x=86 to x=994 and y=270 to y=1500. Do not cut off caps, labels, boxes, hands, or faces.
- RESTRICTED TOP UI BAND: y=0 to y=269 must contain background only. No text, CTA, logos, faces, or product details.
- RESTRICTED LOWER UI BAND: y=1249 to y=1919 must stay visually quiet for Meta/Reels overlays. No CTA, no headline/support copy, no logos, no badges, no readable product-label text, and no key product details in this area.
- The CTA must NEVER be placed near the bottom. Place it inside the strict safe field, preferably around y=1050 to y=1200, or omit it if it cannot fit safely.
- Keep all important content away from left/right edges with at least 8% horizontal padding on both sides.
- Before finalizing, perform a coordinate check. If any critical element crosses these safe fields, reject and regenerate automatically until it passes.

Quality bar:
- Final output must look like the same ad creative adapted to 9:16, not a new redesign.
- Maintain photoreal quality, product fidelity, and clean typography edges.
- If fidelity or safe-zone compliance fails, regenerate until all rules pass.
"""


FORMATS = ["HERO", "BA", "TEST", "FEAT", "UGC"]
DEFAULT_OPENCODE_API_URL = os.getenv("OPENCODE_API_URL", "http://127.0.0.1:4090")
OPENCODE_MAX_CONCURRENT = 2
OPENCODE_QUEUE_DIR = RUNTIME_ROOT / "opencode_queue"
OPENCODE_QUEUE_LOG = OPENCODE_QUEUE_DIR / "queue.log"


def load_format_visual_archetypes() -> dict[str, list[dict[str, str]]]:
    script_path = ROOT / "scripts" / "generate_ads.py"
    spec = importlib.util.spec_from_file_location("dashboard_generate_ads_patterns", script_path)
    if spec is None or spec.loader is None:
        return {}
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    raw = getattr(module, "FORMAT_VISUAL_ARCHETYPES", {})
    out: dict[str, list[dict[str, str]]] = {}
    if not isinstance(raw, dict):
        return out
    for fmt in FORMATS:
        items = raw.get(fmt) or []
        out[fmt] = [
            {"id": str(item.get("id") or ""), "label": str(item.get("label") or item.get("id") or "")}
            for item in items
            if isinstance(item, dict) and str(item.get("id") or "").strip()
        ]
    return out



def classify_hook_structure(headline: str) -> str:
    """Classify a headline opening pattern for hypothesis sanity checks.

    The classifier only verifies whether a pattern is present.
    It does not prescribe which pattern should be used — that's the LLM's job.
    A headline is a question_lead if it opens as a question or contains an
    early question mark. Everything else is classified by visible pattern cues.
    """
    text = (headline or "").strip().lower().replace("’", "'")
    if not text:
        return "proof_lead"
    if text.startswith(("why ", "what ", "how ", "when ", "can ", "could ", "will ", "want ", "need ", "tired ")) or "?" in text[:50]:
        return "question_lead"
    if text.startswith(("stop", "start", "try", "see", "check", "take")):
        return "command_lead"
    contrast_terms = ["before", "after", "without", "instead", " but ", " yet ", " still ", "doesn't have to", "even with"]
    if any(term in f" {text} " for term in contrast_terms):
        return "contrast_loop"
    if text.startswith(("i ", "my ")) or "felt" in text or "struggled" in text:
        return "confession_lead"
    if text.startswith(("finally", "trusted", "proven")) or "70,000" in text or "doctor" in text:
        return "proof_lead"
    return "proof_lead"


def headline_for_candidate(candidate: dict[str, Any], lang: str = "EN") -> str:
    copy = candidate.get("copy") if isinstance(candidate.get("copy"), dict) else {}
    block = copy.get(lang) if isinstance(copy.get(lang), dict) else {}
    return str(block.get("headline") or "").strip()


def copy_text_for_candidate(candidate: dict[str, Any], lang: str = "EN") -> str:
    copy = candidate.get("copy") if isinstance(candidate.get("copy"), dict) else {}
    block = copy.get(lang) if isinstance(copy.get(lang), dict) else {}
    parts: list[str] = []
    for key in ["headline", "support_line", "trust_line", "attribution", "cta"]:
        value = block.get(key)
        if isinstance(value, str) and value.strip():
            parts.append(value.strip())
    bullets = block.get("bullets")
    if isinstance(bullets, list):
        parts.extend(str(item).strip() for item in bullets if str(item).strip())
    return " ".join(parts)


def cta_for_candidate(candidate: dict[str, Any], lang: str = "EN") -> str:
    copy = candidate.get("copy") if isinstance(candidate.get("copy"), dict) else {}
    block = copy.get(lang) if isinstance(copy.get(lang), dict) else {}
    return str(block.get("cta") or "").strip()


def classify_proof_style_text(text: str) -> str:
    lower = (text or "").lower()
    if "70,000" in lower or "user" in lower or "people" in lower or "review" in lower or "testimonial" in lower or "trusted" in lower:
        return "social_proof"
    if "doctor" in lower or "ayurvedic" in lower or "dr." in lower or "formulated" in lower:
        return "authority_anchor"
    if "but" in lower or "skeptical" in lower or "doubt" in lower or "worried" in lower or "tried" in lower:
        return "objection_flip"
    if "simple" in lower or "clear" in lower or "5-minute" in lower or "easy" in lower or "low guesswork" in lower:
        return "routine_clarity"
    if "step" in lower or "routine" in lower or "morning" in lower or "night" in lower or "ok liquid" in lower or "craving" in lower or "fullness" in lower:
        return "mechanism_explainer"
    return "mechanism_explainer"


def classify_cta_voice_text(cta: str) -> str:
    text = (cta or "").strip().lower()
    if "test" in text or "challenge" in text or "15-day" in text or "15 day" in text:
        return "challenge_action"
    if "fit" in text or "risk" in text or "safe" in text or "suit" in text:
        return "reassurance_start"
    if "today" in text or "now" in text or "start" in text or "act" in text:
        return "urgent_start"
    if "learn" in text or "how" in text or "works" in text or "discover" in text:
        return "discovery_action"
    if "see" in text or "view" in text or "check" in text or "steps" in text or "details" in text or "plan" in text or "protocol" in text:
        return "guided_next_step"
    return "guided_next_step"


def proof_style_matches(expected: str, text: str) -> bool:
    lower = (text or "").lower()
    checks = {
        "authority_anchor": ["doctor", "ayurvedic", "dr.", "formulated"],
        "social_proof": ["70,000", "user", "people", "review", "testimonial", "trusted"],
        "mechanism_explainer": ["step", "routine", "morning", "night", "ok liquid", "craving", "fullness", "works"],
        "routine_clarity": ["simple", "clear", "5-minute", "easy", "low guesswork", "step", "routine"],
        "objection_flip": ["but", "skeptical", "doubt", "worried", "tried", "without"],
    }
    return any(term in lower for term in checks.get(expected, []))


def cta_voice_matches(expected: str, cta: str) -> bool:
    text = (cta or "").strip().lower()
    checks = {
        "urgent_start": ["today", "now", "start", "act"],
        "guided_next_step": ["see", "view", "check", "steps", "details", "plan", "protocol"],
        "reassurance_start": ["fit", "risk", "safe", "suit"],
        "challenge_action": ["test", "challenge", "15-day", "15 day"],
        "discovery_action": ["learn", "how", "works", "discover"],
    }
    return any(term in text for term in checks.get(expected, []))


def hook_structure_mismatch(candidate: dict[str, Any], planned_ad: dict[str, Any]) -> str | None:
    hypothesis = planned_ad.get("hypothesis") if isinstance(planned_ad.get("hypothesis"), dict) else {}
    if hypothesis.get("type") != "hook_structure":
        return None
    expected = str(hypothesis.get("variant") or "").strip()
    if not expected:
        return None
    headline = headline_for_candidate(candidate, "EN")
    actual = classify_hook_structure(headline)
    if actual != expected:
        return f"Expected hook_structure {expected}, but EN headline classified as {actual}: {headline!r}"
    return None


def hypothesis_mismatch(candidate: dict[str, Any], planned_ad: dict[str, Any]) -> str | None:
    hypothesis = planned_ad.get("hypothesis") if isinstance(planned_ad.get("hypothesis"), dict) else {}
    hyp_type = hypothesis.get("type")
    expected = str(hypothesis.get("variant") or "").strip()
    if not hyp_type or hyp_type == "none" or not expected:
        return None
    if hyp_type == "hook_structure":
        return hook_structure_mismatch(candidate, planned_ad)
    if hyp_type == "concept_angle":
        actual = str(candidate.get("concept_angle") or "").strip()
        if actual != expected:
            return f"Expected concept_angle {expected}, but candidate returned {actual or 'blank'}"
    if hyp_type == "awareness_stage":
        actual = str(candidate.get("awareness_stage") or "").strip()
        if actual != expected:
            return f"Expected awareness_stage {expected}, but candidate returned {actual or 'blank'}"
    if hyp_type == "concept_structure":
        actual = str(candidate.get("concept_structure") or "").strip()
        if actual != expected:
            return f"Expected concept_structure {expected}, but candidate returned {actual or 'blank'}"
    if hyp_type == "proof_style":
        copy_text = copy_text_for_candidate(candidate, "EN")
        if proof_style_matches(expected, copy_text):
            return None
        actual = classify_proof_style_text(copy_text)
        if actual != expected:
            return f"Expected proof_style {expected}, but EN copy classified as {actual}: {copy_text!r}"
    if hyp_type == "cta_voice":
        cta = cta_for_candidate(candidate, "EN")
        if cta_voice_matches(expected, cta):
            return None
        actual = classify_cta_voice_text(cta)
        if actual != expected:
            return f"Expected cta_voice {expected}, but EN CTA classified as {actual}: {cta!r}"
    return None


HYPOTHESIS_VARIABLES: dict[str, dict[str, Any]] = {}




def resolve_language_mode(config: dict[str, Any]) -> str:
    mode = str(config.get("language_mode") or "ALL").strip().upper()
    if mode in {"EN", "HI", "HINGLISH", "ALL"}:
        return mode
    return "ALL"


def assembler_language_mode(config: dict[str, Any]) -> str:
    mode = resolve_language_mode(config)
    if mode == "EN":
        return "EN"
    if mode == "HI":
        return "HI"
    if mode == "HINGLISH":
        return "HINGLISH"
    return "BOTH"


def now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def make_run_id() -> str:
    return f"run_{int(time.time())}_{random.randint(1000, 9999)}"


def ensure_dirs() -> None:
    RUNS_ROOT.mkdir(parents=True, exist_ok=True)
    RUNTIME_ROOT.mkdir(parents=True, exist_ok=True)
    INPUT_IMAGES_DIR.mkdir(parents=True, exist_ok=True)


def load_env_file(path: Path) -> None:
    if not path.exists() or not path.is_file():
        return
    for raw_line in path.read_text(encoding="utf-8", errors="ignore").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


def parse_persona_library(playbook_path: Path) -> list[dict[str, Any]]:
    path = PERSONA_SEEDS_PATH
    if not path.exists():
        return []
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    return [{"number": int(e["persona_number"]), "name": str(e["persona_name"])} for e in data]


def _load_persona_seeds() -> dict[int, dict[str, str]]:
    path = PERSONA_SEEDS_PATH
    if not path.exists():
        print(f"WARNING: {path} not found. Using empty persona seeds.", file=sys.stderr)
        return {}
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    seeds: dict[int, dict[str, str]] = {}
    for entry in data:
        pn = int(entry.get("persona_number", 0))
        if pn < 1:
            continue
        seeds[pn] = {
            "pain": str(entry.get("pain", "")),
            "desire": str(entry.get("desire", "")),
            "friction": str(entry.get("friction", "")),
            "proof": str(entry.get("proof", "")),
            "tone": str(entry.get("tone", "")),
            "awareness_stage": str(entry.get("awareness_stage", "unaware")),
        }
    return seeds


PERSONA_SEED_INPUTS = _load_persona_seeds()


def _load_copy_architecture() -> dict[str, Any]:
    path = COPY_ARCH_PATH
    if not path.exists():
        print(f"WARNING: {path} not found. Copy architecture templates disabled.", file=sys.stderr)
        return {"headline_architectures": {}, "support_line_architectures": {"rotation_order": [], "definitions": {}}}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception as exc:
        print(f"WARNING: Failed to load {path}: {exc}", file=sys.stderr)
        return {"headline_architectures": {}, "support_line_architectures": {"rotation_order": [], "definitions": {}}}


COPY_ARCH = _load_copy_architecture()


def _load_copy_prompts() -> dict[str, Any]:
    path = COPY_PROMPTS_PATH
    if not path.exists():
        print(f"WARNING: {path} not found. Prompt templates disabled.", file=sys.stderr)
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception as exc:
        print(f"WARNING: Failed to load {path}: {exc}", file=sys.stderr)
        return {}


def _hypothesis_variant_label(variant_id: str) -> str:
    acronyms = {"pas", "bab", "fab"}
    if variant_id in acronyms:
        return variant_id.upper()
    return variant_id.replace("_", " ").title()


def _build_hypothesis_variables() -> dict[str, dict[str, Any]]:
    hv: dict[str, dict[str, Any]] = {
        "none": {
            "label": "No hypothesis test",
            "description": "Generate ads normally without controlled A/B testing.",
            "options": [],
        }
    }

    arch_types = {
        "hook_structure": {
            "label": "Hook Structure (H1)",
            "description": "Test which headline opening pattern performs best: question vs. proof vs. contrast vs. confession vs. command.",
        },
        "concept_angle": {
            "label": "Concept Angle (H2)",
            "description": "Test which messaging angle drives better results: pain vs. outcome vs. proof vs. authority vs. curiosity.",
        },
        "awareness_stage": {
            "label": "Awareness Stage (H3)",
            "description": "Test whether matching the ad to the audience\u2019s funnel stage improves performance.",
        },
        "concept_structure": {
            "label": "Concept Structure (H6)",
            "description": "Test copy flow structure: PAS vs BAB vs FAB vs Four Us.",
        },
    }
    arch = COPY_ARCH.get("headline_architectures", {})
    for hyp_type, meta in arch_types.items():
        options = [{"id": vid, "label": _hypothesis_variant_label(vid)} for vid in arch.get(hyp_type, {})]
        hv[hyp_type] = {**meta, "options": options}

    cf_types = {
        "proof_style": {
            "label": "Proof Style (H4)",
            "description": "Test which trust framing works best for this persona: authority vs. social proof vs. mechanism explainer.",
        },
        "cta_voice": {
            "label": "CTA Voice (H5)",
            "description": "Test which call-to-action tone converts better: urgent vs. guided vs. reassuring vs. discovery.",
        },
    }
    cf = COPY_PROMPTS.get("concept_framework", {})
    for hyp_type, meta in cf_types.items():
        options = [{"id": vid, "label": _hypothesis_variant_label(vid)} for vid in cf.get(hyp_type, {})]
        hv[hyp_type] = {**meta, "options": options}

    return hv


COPY_PROMPTS = _load_copy_prompts()
HYPOTHESIS_VARIABLES = _build_hypothesis_variables()
CTA_VARIANTS = COPY_PROMPTS.get("cta_variants", {})


def _framework_item(group: str, item_id: str) -> dict[str, str]:
    items = COPY_PROMPTS.get("concept_framework", {}).get(group, [])
    if not items:
        return {"id": item_id, "direction": ""}
    for item in items:
        if item["id"] == item_id:
            return item
    return items[0]


# Feature-lane-driven headline concept selection removed.


def _select_headline_architecture(persona_number: int, fmt: str, concept_structure_id: str) -> dict[str, Any]:
    arch = COPY_ARCH.get("headline_architectures", {})
    structure_arch = arch.get("concept_structure", {}).get(concept_structure_id)
    if structure_arch:
        return {"source": "concept_structure", "variant": concept_structure_id, **structure_arch}
    hook_arch = arch.get("hook_structure", {})
    hook_keys = list(hook_arch.keys())
    if hook_keys:
        idx = (persona_number + sum(ord(c) for c in fmt)) % len(hook_keys)
        hook_id = hook_keys[idx]
        return {"source": "hook_structure", "variant": hook_id, **hook_arch[hook_id]}
    return {"source": "four_us", "variant": "four_us", "template": "", "examples": []}


def _select_support_line_architecture(persona_number: int, fmt: str, format_sequence_index: int) -> dict[str, Any]:
    sla = COPY_ARCH.get("support_line_architectures", {})
    rotation = sla.get("rotation_order", [])
    defs = sla.get("definitions", {})
    if not rotation:
        return {"variant": "none", "template": "", "examples": []}
    idx = (persona_number + format_sequence_index) % len(rotation)
    arch_id = rotation[idx]
    entry = defs.get(arch_id, {})
    return {"variant": arch_id, **entry}


def build_copy_requirements(persona_number: int, fmt: str, format_sequence_index: int, variation_seed: str = "") -> dict[str, Any]:
    persona_seed = PERSONA_SEED_INPUTS.get(persona_number, {})
    audience_id = persona_seed.get("awareness_stage", "unaware")

    fmt_defaults = COPY_PROMPTS.get("format_defaults", {})
    structure_by_fmt = fmt_defaults.get("structure_by_fmt", {"HERO": "four_us"})
    concept_structure_id = structure_by_fmt.get(fmt, "four_us")

    lead_angle = fmt_defaults.get("default_lead_angle", "desired_outcome")

    headline_arch = _select_headline_architecture(persona_number, fmt, concept_structure_id)
    support_line_arch = _select_support_line_architecture(persona_number, fmt, format_sequence_index)

    prompts = COPY_PROMPTS.get("copy_requirements", {})
    format_specific = prompts.get("format_specific", {})

    return {
        "must_mention": prompts.get("must_mention", "Headline or paired support must ensure the copy is clearly about weight loss \u2014 directly or indirectly."),
        "variation_rule": prompts.get("variation_rule", "Do not reuse the same headline skeleton, support-line skeleton, or persuasion angle as other ads in the same format for this batch."),
        "concept_variation": {
            "audience_stage": _framework_item("audience_stage", audience_id),
            "lead_angle": _framework_item("lead_angle", lead_angle),
            "message_structure": _framework_item("message_structure", concept_structure_id),
        },
        "hierarchy_rule": prompts.get("hierarchy_rule", "Use the assigned concept_structure flow to shape headline and support line. Do not output framework labels."),
        "format_specific_rule": format_specific.get(fmt, prompts.get("default_format_rule", "")),
        "headline_architecture": {
            "template": headline_arch.get("template", ""),
            "examples": headline_arch.get("examples", []),
            "source": headline_arch.get("source", ""),
            "variant": headline_arch.get("variant", ""),
        },
        "support_line_architecture": {
            "template": support_line_arch.get("template", ""),
            "examples": support_line_arch.get("examples", []),
            "variant": support_line_arch.get("variant", ""),
        },
    }


def compact_format_rules_for_copy(fmt: str, format_rules: dict[str, Any]) -> dict[str, Any]:
    fmt = fmt.strip().upper()
    prompts = COPY_PROMPTS
    wanted = (prompts.get("format_copy_keywords") or {}).get(fmt, [])
    blocked = (prompts.get("format_visual_keywords") or {}).get("blocklist", [])
    out: list[str] = []
    for raw_rule in format_rules.get("rules") or []:
        rule = str(raw_rule).strip()
        if not rule:
            continue
        lower = rule.lower()
        if any(b in lower for b in blocked):
            continue
        if wanted and not any(k in lower for k in wanted):
            continue
        out.append(rule)
        if len(out) >= 8:
            break
    return {"format": fmt or format_rules.get("format"), "rules": out}

def build_ad_copy_system_prompt(fmt: str) -> str:
    fmt = fmt.strip().upper()
    prompts = COPY_PROMPTS
    base_rules = prompts.get("system_prompt_base_rules", [])
    format_rules = prompts.get("system_prompt_format_rules", {})
    fmt_rules = format_rules.get(fmt, [])
    return " ".join(base_rules + fmt_rules)


def build_strict_schema_note(fmt: str) -> str:
    fmt = fmt.strip().upper()
    prompts = COPY_PROMPTS.get("strict_schema_note", {})
    field_map = prompts.get("field_map", {})
    copy_fields = field_map.get(fmt, prompts.get("default_fields", "headline, cta"))
    return " ".join([
        prompts.get("intro", ""),
        prompts.get("persona_fields_en", ""),
        prompts.get("language_extension", ""),
        prompts.get("format_closure_template", "").format(fmt=fmt or "this", copy_fields=copy_fields),
    ])


def build_ad_prompt_tail(fmt: str) -> str:
    fmt = fmt.strip().upper()
    tail = COPY_PROMPTS.get("prompt_tail", {})
    support_map = tail.get("support_target_map", {})
    support_target = support_map.get(fmt, tail.get("default_support_target", "support line"))
    lines = [line.format(fmt=fmt or "ad", support_target=support_target) for line in tail.get("lines", [])]
    return "\n".join(lines)


def build_generation_payload_for_llm(context: dict[str, Any]) -> dict[str, Any]:
    compact_ads: list[dict[str, Any]] = []
    for item in context.get("ads") or []:
        if not isinstance(item, dict):
            continue
        fmt = str(item.get("format") or "").strip().upper()
        persona = item.get("persona") if isinstance(item.get("persona"), dict) else {}
        format_rules = item.get("format_rules") if isinstance(item.get("format_rules"), dict) else {}
        copy_requirements = item.get("copy_requirements") if isinstance(item.get("copy_requirements"), dict) else {}
        compact_ads.append(
            {
                "format": fmt,
                "persona": persona,
                "format_rules": compact_format_rules_for_copy(fmt, format_rules),
                "copy_requirements": copy_requirements,
            }
        )

    requested_plan = []
    for item in compact_ads:
        persona = item.get("persona") if isinstance(item.get("persona"), dict) else {}
        requested_plan.append(
            {
                "format": item.get("format"),
                "persona_number": persona.get("persona_number"),
                "persona_name": persona.get("persona_name"),
            }
        )

    return {
        "generated_at": context.get("generated_at"),
        "run_id": context.get("run_id"),
        "language_mode": context.get("language_mode"),
        "context_source": context.get("context_source"),
        "requested_ad_count": len(compact_ads),
        "requested_plan": requested_plan,
        "product_doc": {
            "attached_in_session": True,
            "source_file": context.get("product_file_path"),
            "instruction": "Read and use the attached product master doc as source of truth for all product claims.",
        },
        "ads": compact_ads,
    }


def validate_generated_copy_payload(copy_json: dict[str, Any], planned_ads: list[dict[str, Any]]) -> str | None:
    ads = copy_json.get("ads") if isinstance(copy_json, dict) else None
    if not isinstance(ads, list):
        return "Generated payload did not include an ads array"
    if len(ads) < len(planned_ads):
        return f"Generated ads count {len(ads)} is lower than planned count {len(planned_ads)}"

    planned_keys = {
        (str(item.get("format") or "").strip().upper(), int((item.get("persona") or {}).get("persona_number") or 0))
        for item in planned_ads
        if isinstance(item, dict) and isinstance(item.get("persona"), dict)
    }
    seen_keys: set[tuple[str, int]] = set()
    for ad in ads:
        if not isinstance(ad, dict):
            return "Generated ads payload contains a non-object item"
        fmt = str(ad.get("format") or "").strip().upper()
        persona = ad.get("persona") if isinstance(ad.get("persona"), dict) else {}
        persona_number = persona.get("number")
        if not isinstance(persona_number, int):
            persona_number = persona.get("persona_number")
        if not isinstance(persona_number, int):
            return f"Generated ad for format {fmt or '?'} is missing persona number"
        seen_keys.add((fmt, persona_number))
        copy = ad.get("copy") if isinstance(ad.get("copy"), dict) else {}
        for lang in ["EN", "HI"]:
            block = copy.get(lang) if isinstance(copy.get(lang), dict) else {}
            if not str(block.get("headline") or "").strip():
                return f"Generated ad {fmt}/P{persona_number} is missing {lang} headline"
            if fmt in {"HERO", "UGC"} and not str(block.get("support_line") or "").strip():
                return f"Generated ad {fmt}/P{persona_number} is missing {lang} support line"
            if fmt in {"BA", "FEAT"}:
                bullets = block.get("bullets") if isinstance(block.get("bullets"), list) else []
                if len([item for item in bullets if isinstance(item, str) and item.strip()]) < 2:
                    return f"Generated ad {fmt}/P{persona_number} has insufficient {lang} bullets"
            if fmt == "TEST" and not str(block.get("attribution") or "").strip():
                return f"Generated ad {fmt}/P{persona_number} is missing {lang} attribution"

    missing = sorted(planned_keys - seen_keys)
    if missing:
        return "Generated payload is missing planned ads: " + ", ".join(f"{fmt}/P{persona}" for fmt, persona in missing)
    return None


def extract_generated_ad_candidate(payload: dict[str, Any]) -> dict[str, Any] | None:
    if not isinstance(payload, dict):
        return None

    def normalize_candidate(candidate: dict[str, Any]) -> dict[str, Any]:
        cloned = json.loads(json.dumps(candidate, ensure_ascii=False))
        persona = cloned.get("persona") if isinstance(cloned.get("persona"), dict) else {}
        if not isinstance(persona, dict):
            persona = {}
        if not isinstance(persona.get("persona_number"), int) and isinstance(cloned.get("persona_number"), int):
            persona["persona_number"] = cloned.get("persona_number")
        if not isinstance(persona.get("number"), int) and isinstance(cloned.get("persona_number"), int):
            persona["number"] = cloned.get("persona_number")
        if not isinstance(persona.get("persona_name"), str) and isinstance(cloned.get("persona_name"), str):
            persona["persona_name"] = cloned.get("persona_name")
        if not isinstance(persona.get("name"), str) and isinstance(cloned.get("persona_name"), str):
            persona["name"] = cloned.get("persona_name")
        if persona:
            cloned["persona"] = persona
        return cloned

    ads = payload.get("ads")
    if isinstance(ads, list):
        for item in ads:
            if isinstance(item, dict) and item.get("copy"):
                return normalize_candidate(item)
    if payload.get("format") and payload.get("copy"):
        return normalize_candidate(payload)
    return None


def hydrate_generated_ad_candidate(candidate: dict[str, Any], planned_ad: dict[str, Any]) -> dict[str, Any]:
    hydrated = json.loads(json.dumps(candidate, ensure_ascii=False))

    planned_format = str(planned_ad.get("format") or "").strip().upper()
    candidate_format = str(hydrated.get("format") or "").strip().upper()
    hydrated["format"] = candidate_format or planned_format

    planned_persona = planned_ad.get("persona") if isinstance(planned_ad.get("persona"), dict) else {}
    candidate_persona = hydrated.get("persona") if isinstance(hydrated.get("persona"), dict) else {}

    persona_number = candidate_persona.get("number")
    if not isinstance(persona_number, int):
        persona_number = candidate_persona.get("persona_number")
    if not isinstance(persona_number, int):
        planned_number = planned_persona.get("persona_number")
        if isinstance(planned_number, int):
            persona_number = planned_number

    persona_name = candidate_persona.get("name")
    if not isinstance(persona_name, str) or not persona_name.strip():
        persona_name = candidate_persona.get("persona_name")
    if not isinstance(persona_name, str) or not persona_name.strip():
        planned_name = planned_persona.get("persona_name")
        if isinstance(planned_name, str):
            persona_name = planned_name

    merged_persona = dict(candidate_persona)
    if isinstance(persona_number, int):
        merged_persona["number"] = persona_number
        merged_persona["persona_number"] = persona_number
    if isinstance(persona_name, str) and persona_name.strip():
        clean_name = persona_name.strip()
        merged_persona["name"] = clean_name
        merged_persona["persona_name"] = clean_name
    if merged_persona:
        hydrated["persona"] = merged_persona

    copy_payload = hydrated.get("copy") if isinstance(hydrated.get("copy"), dict) else {}
    normalized_copy: dict[str, Any] = {}
    for lang in ["EN", "HI"]:
        lang_block = copy_payload.get(lang)
        normalized_copy[lang] = lang_block if isinstance(lang_block, dict) else {}
    hydrated["copy"] = normalized_copy

    return hydrated


def _build_persona_payload_field(seed_field_value: Any, config: dict[str, Any]) -> Any:
    wrap = config.get("wrap_list", False)
    prefix = config.get("prefix")
    if wrap:
        val = str(seed_field_value) if seed_field_value else ""
        return [f"{prefix}{val}" if prefix and val else val]
    return seed_field_value

def build_persona_payload(persona_number: int, personas: list[dict[str, Any]]) -> dict[str, Any]:
    persona_name = f"Persona {persona_number}"
    for item in personas:
        if int(item.get("number") or 0) == persona_number:
            name = str(item.get("name") or "").strip()
            if name:
                persona_name = name
            break
    seed = PERSONA_SEED_INPUTS.get(persona_number, {})
    mapping = COPY_PROMPTS.get("persona_mapping", {})
    seed_to_payload = mapping.get("seed_to_payload", {})
    fallbacks = mapping.get("persona_fallbacks", {})

    payload: dict[str, Any] = {
        "persona_number": persona_number,
        "persona_name": persona_name,
    }

    for seed_key, field_cfg in seed_to_payload.items():
        raw = seed.get(seed_key, fallbacks.get(seed_key, ""))
        payload[field_cfg["field"]] = _build_persona_payload_field(raw, field_cfg)

    static = mapping.get("static_fields", {})
    for key, val in static.items():
        payload[key] = val

    hindi_default = mapping.get("hindi_ready_default", "")
    if hindi_default and payload.get("hindi_ready") in (None, []):
        payload["hindi_ready"] = [hindi_default]

    return payload


def read_active_images(path: Path) -> list[str]:
    if not path.exists():
        return []
    lines = [line.strip() for line in path.read_text(encoding="utf-8").splitlines()]
    return [line for line in lines if line and not line.startswith("#")]


def default_image_sources_file() -> Path:
    if DEFAULT_IMAGE_SOURCES_FILE.exists():
        return DEFAULT_IMAGE_SOURCES_FILE
    return LEGACY_ACTIVE_IMAGES_FILE


def list_input_images() -> list[str]:
    if not INPUT_IMAGES_DIR.exists():
        return []
    allowed = {".png", ".jpg", ".jpeg", ".webp", ".bmp", ".gif"}
    items = [
        p for p in sorted(INPUT_IMAGES_DIR.iterdir())
        if p.is_file() and p.suffix.lower() in allowed
    ]
    return [str(p.relative_to(ROOT)).replace("\\", "/") for p in items]


def default_product_doc_info() -> dict[str, Any]:
    return {
        "path": str(DEFAULT_PRODUCT_MASTER.relative_to(ROOT)).replace("\\", "/"),
        "name": DEFAULT_PRODUCT_MASTER.name,
        "exists": DEFAULT_PRODUCT_MASTER.exists(),
        "size_bytes": DEFAULT_PRODUCT_MASTER.stat().st_size if DEFAULT_PRODUCT_MASTER.exists() else 0,
    }


def store_uploaded_input_images(files: list[UploadFile], clear_existing: bool) -> list[str]:
    ensure_dirs()
    if clear_existing and INPUT_IMAGES_DIR.exists():
        for existing in INPUT_IMAGES_DIR.iterdir():
            if existing.is_file():
                existing.unlink(missing_ok=True)

    allowed = {".png", ".jpg", ".jpeg", ".webp", ".bmp", ".gif"}
    saved: list[str] = []
    for upload in files:
        filename = Path(upload.filename or "").name
        if not filename:
            continue
        ext = Path(filename).suffix.lower()
        if ext not in allowed:
            continue
        target = INPUT_IMAGES_DIR / filename
        counter = 1
        while target.exists():
            target = INPUT_IMAGES_DIR / f"{Path(filename).stem}_{counter}{ext}"
            counter += 1
        data = upload.file.read()
        target.write_bytes(data)
        saved.append(str(target.relative_to(ROOT)).replace("\\", "/"))
    return saved


def api_delete_input_image(payload: dict[str, Any] = Body(...)) -> dict[str, Any]:
    rel_path = str(payload.get("path") or "").strip().replace("\\", "/")
    if not rel_path.startswith("input/images/"):
        raise HTTPException(status_code=400, detail="path must be under input/images")
    target = (ROOT / rel_path).resolve()
    images_root = INPUT_IMAGES_DIR.resolve()
    if images_root not in target.parents:
        raise HTTPException(status_code=400, detail="Invalid image path")
    if not target.exists() or not target.is_file():
        raise HTTPException(status_code=404, detail="Input image not found")
    target.unlink()
    return {"status": "deleted", "path": rel_path}


def api_product_doc() -> dict[str, Any]:
    info = default_product_doc_info()
    content = DEFAULT_PRODUCT_MASTER.read_text(encoding="utf-8", errors="ignore") if DEFAULT_PRODUCT_MASTER.exists() else ""
    return {**info, "content": content}


def api_save_product_doc(payload: dict[str, Any] = Body(...)) -> dict[str, Any]:
    content = str(payload.get("content") or "")
    DEFAULT_PRODUCT_MASTER.parent.mkdir(parents=True, exist_ok=True)
    DEFAULT_PRODUCT_MASTER.write_text(content, encoding="utf-8")
    return {"status": "saved", **default_product_doc_info()}


def _is_opencode_run_cmd(cmd: list[str]) -> bool:
    return bool(cmd) and Path(cmd[0]).name == "opencode" and len(cmd) > 1 and cmd[1] == "run"


def _append_opencode_queue_log(message: str) -> None:
    try:
        OPENCODE_QUEUE_DIR.mkdir(parents=True, exist_ok=True)
        with OPENCODE_QUEUE_LOG.open("a", encoding="utf-8") as handle:
            handle.write(f"{now_iso()} {message}\n")
    except OSError:
        pass


def dashboard_subprocess_env() -> dict[str, str]:
    env = dict(os.environ)
    playwright_path = "/home/mylappy/.local/lib/python3.12/site-packages"
    if playwright_path not in env.get("PYTHONPATH", ""):
        env["PYTHONPATH"] = playwright_path + ((":" + env["PYTHONPATH"]) if env.get("PYTHONPATH") else "")
    return env


@contextmanager
def _opencode_queue_slot(label: str) -> Iterator[None]:
    OPENCODE_QUEUE_DIR.mkdir(parents=True, exist_ok=True)
    queued_at = time.time()
    logged_wait = False
    while True:
        for slot in range(OPENCODE_MAX_CONCURRENT):
            lock_path = OPENCODE_QUEUE_DIR / f"slot_{slot}.lock"
            lock_handle = lock_path.open("a+")
            try:
                fcntl.flock(lock_handle.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
            except BlockingIOError:
                lock_handle.close()
                continue
            wait_seconds = time.time() - queued_at
            if wait_seconds >= 0.25:
                _append_opencode_queue_log(f"{label} started slot={slot} wait_seconds={wait_seconds:.1f}")
            try:
                yield
                return
            finally:
                try:
                    fcntl.flock(lock_handle.fileno(), fcntl.LOCK_UN)
                finally:
                    lock_handle.close()
        if not logged_wait:
            _append_opencode_queue_log(f"{label} queued max_concurrent={OPENCODE_MAX_CONCURRENT}")
            logged_wait = True
        time.sleep(0.25)


def _run_opencode_queued(cmd: list[str], cwd: Path, env: dict[str, str]) -> subprocess.CompletedProcess[str]:
    with _opencode_queue_slot("command"):
        return subprocess.run(cmd, cwd=str(cwd), text=True, capture_output=True, check=False, env=env)


def run_cmd(cmd: list[str], cwd: Path) -> subprocess.CompletedProcess[str]:
    env = dashboard_subprocess_env()
    if _is_opencode_run_cmd(cmd):
        return _run_opencode_queued(cmd, cwd, env)
    return subprocess.run(cmd, cwd=str(cwd), text=True, capture_output=True, check=False, env=env)


def generated_image_roots() -> list[Path]:
    return [GENERATED_IMAGES_ROOT]


def ensure_916_conversion_template() -> Path:
    CONVERT_916_TEMPLATE_PATH.parent.mkdir(parents=True, exist_ok=True)
    if not CONVERT_916_TEMPLATE_PATH.exists():
        CONVERT_916_TEMPLATE_PATH.write_text(DEFAULT_916_CONVERSION_PROMPT, encoding="utf-8")
    return CONVERT_916_TEMPLATE_PATH


def build_916_conversion_prompt_job(fmt: str, persona_num: int, lang: str, index: int) -> str:
    fmt_clean = (fmt or "HERO").strip().upper() or "HERO"
    lang_clean = (lang or "EN").strip().upper() or "EN"
    persona_safe = max(0, int(persona_num or 0))
    if persona_safe > 0:
        return f"OUTPUT_{fmt_clean}_P{persona_safe:02d}_{lang_clean}_A{max(1, int(index)):02d}.txt"
    return f"OUTPUT_{fmt_clean}_P{index:02d}_{lang_clean}.txt"


def collect_45_reference_jobs_for_batch(batch: str) -> list[dict[str, Any]]:
    summary = load_batch_image_summary(batch)
    jobs: list[dict[str, Any]] = []
    seen_refs: set[str] = set()

    for entry in summary:
        prompt_file = str(entry.get("prompt_file") or "").strip().replace("\\", "/")
        saved_files = entry.get("saved_files") if isinstance(entry.get("saved_files"), list) else []
        if not prompt_file or not saved_files:
            continue

        parsed = parse_prompt_filename(prompt_file)
        if not parsed:
            continue
        fmt, lang, persona_num = parsed
        if persona_num is None:
            continue

        for candidate in saved_files:
            c = str(candidate or "").strip().replace("\\", "/")
            if not c:
                continue
            if c in seen_refs:
                continue
            image_abs = (ROOT / c).resolve()
            if not image_abs.exists() or not image_abs.is_file():
                continue

            seen_refs.add(c)
            jobs.append(
                {
                    "format": fmt.upper(),
                    "persona_number": int(persona_num),
                    "language": lang.upper(),
                    "image_rel": c,
                    "image_abs": str(image_abs),
                }
            )

    if jobs:
        return jobs

    # Fallback: derive from prompt files + filesystem scan under 4_5
    prompt_files = scan_prompt_files_for_batch(batch)
    for prompt_file in prompt_files:
        if "/45/" not in str(prompt_file):
            continue
        parsed = parse_prompt_filename(prompt_file)
        if not parsed:
            continue
        fmt, lang, persona_num = parsed
        if persona_num is None:
            continue
        base_name = f"p{persona_num:02d}"
        for img_root in generated_image_roots():
            ref_dir = img_root / batch / "4_5"
            if not ref_dir.exists():
                continue
            for ext in ("png", "jpg", "jpeg", "webp"):
                for f in sorted(ref_dir.glob(f"**/*{base_name}*.{ext}")):
                    found_rel = str(f.relative_to(ROOT)).replace("\\", "/")
                    if found_rel in seen_refs:
                        continue
                    image_abs = (ROOT / found_rel).resolve()
                    if not image_abs.exists() or not image_abs.is_file():
                        continue
                    seen_refs.add(found_rel)
                    jobs.append(
                        {
                            "format": fmt.upper(),
                            "persona_number": int(persona_num),
                            "language": lang.upper(),
                            "image_rel": found_rel,
                            "image_abs": str(image_abs),
                        }
                    )

    return jobs


def image_static_route_for_path(path: str) -> str:
    normalized = path.replace("\\", "/")
    if normalized.startswith("generated_images/"):
        return f"/generated_images/{normalized.removeprefix('generated_images/')}"
    return f"/generated_images/{normalized}"


def gemini_debugger_args() -> list[str]:
    address = resolve_gemini_debugger_address()
    return ["--attach-debugger-address", address]


def debugger_endpoint_reachable(address: str) -> bool:
    if not address:
        return False
    url = f"http://{address}/json/version"
    try:
        with urllib.request.urlopen(url, timeout=1.5) as resp:
            return resp.status == 200
    except Exception:
        return False


def resolve_gemini_debugger_address() -> str:
    configured = str(os.getenv("GEMINI_DEBUGGER_ADDRESS") or "").strip()
    candidates = [configured] if configured else []
    candidates.extend(["127.0.0.1:9222", "localhost:9222", "127.0.0.1:9223", "localhost:9223"])
    for candidate in candidates:
        if candidate and debugger_endpoint_reachable(candidate):
            return candidate
    # No reachable endpoint now: return preferred default so automation script
    # can auto-launch a debuggable Chrome session and continue.
    return configured or "127.0.0.1:9222"


def run_gemini_generation(
    *,
    batch: str,
    prompt_files: list[str],
    aspect_ratio: str,
    image_sources_file: str | None,
    prompt_reference_map: Path | None = None,
    headless: bool = False,
    run_dir: Path | None = None,
    prepend_starting_prompt: bool = True,
) -> subprocess.CompletedProcess[str]:
    aspect_folder = "9_16" if aspect_ratio == "9:16" else "4_5"
    prompt_work_dir = RUNTIME_ROOT / "gemini_selected_prompts" / f"{batch}_{aspect_folder}_{int(time.time())}_{uuid.uuid4().hex[:8]}"
    prompt_work_dir.mkdir(parents=True, exist_ok=True)

    starting_prompt = ""
    if prepend_starting_prompt:
        starting_prompt_path = ROOT / "input" / "startingprompt.txt"
        starting_prompt = starting_prompt_path.read_text(encoding="utf-8").strip() if starting_prompt_path.exists() else ""
    for prompt_file in prompt_files:
        source = Path(prompt_file)
        if not source.is_absolute():
            source = ROOT / source
        source = source.resolve()
        if not source.exists():
            raise RuntimeError(f"Prompt file not found: {source}")
        prompt_text = source.read_text(encoding="utf-8")
        combined = f"{starting_prompt}\n\n{prompt_text.strip()}\n" if starting_prompt else prompt_text
        (prompt_work_dir / source.name).write_text(combined, encoding="utf-8")
        sidecar = source.with_suffix(".json")
        if sidecar.exists():
            (prompt_work_dir / sidecar.name).write_text(sidecar.read_text(encoding="utf-8"), encoding="utf-8")

    out_dir = GENERATED_IMAGES_ROOT / batch / aspect_folder
    image_source_arg = image_sources_file
    if prompt_reference_map is not None:
        try:
            reference_payload = json.loads(prompt_reference_map.read_text(encoding="utf-8"))
        except Exception as exc:
            raise RuntimeError(f"Could not read prompt reference map: {exc}") from exc
        flattened_sources: list[str] = []
        if isinstance(reference_payload, dict):
            for value in reference_payload.values():
                if isinstance(value, list):
                    for item in value:
                        if isinstance(item, str) and item.strip() and item.strip() not in flattened_sources:
                            flattened_sources.append(item.strip())
        if flattened_sources:
            source_file = prompt_work_dir / "image_sources.txt"
            source_file.write_text("\n".join(flattened_sources) + "\n", encoding="utf-8")
            image_source_arg = str(source_file)

    cmd = [
        sys.executable,
        "scripts/gemini_web_automation.py",
        "--prompt-dir",
        str(prompt_work_dir),
        "--prompt-glob",
        "*.txt",
        "--out-dir",
        str(out_dir),
        "--timeout",
        str(int(os.getenv("GEMINI_GENERATION_TIMEOUT_SECONDS") or "420")),
        "--manual-login-timeout",
        str(int(os.getenv("GEMINI_MANUAL_LOGIN_TIMEOUT_SECONDS") or "180")),
        "--upload-dir",
        str(INPUT_IMAGES_DIR),
    ]
    if headless:
        cmd.append("--headless")
    if image_source_arg:
        cmd.extend(["--image-source-file", image_source_arg])
    if run_dir is not None:
        hyp_path = run_dir / "context" / "hypothesis_config.json"
        if hyp_path.exists():
            cmd.extend(["--hypothesis-config", str(hyp_path)])

    log_dir = RUNTIME_ROOT / "generation_logs"
    log_dir.mkdir(parents=True, exist_ok=True)
    log_path = log_dir / f"gen_{batch}_{aspect_folder}.log"

    env = dashboard_subprocess_env()

    with open(log_path, "w") as log_file:
        result = subprocess.run(cmd, cwd=str(ROOT), text=True, stdout=log_file, stderr=subprocess.STDOUT, check=False, env=env)

    full_output = log_path.read_text() if log_path.exists() else ""
    result.stdout = full_output
    result.stderr = ""
    return result


def run_chatgpt_generation(
    *,
    batch: str,
    prompt_files: list[str],
    aspect_ratio: str,
    image_sources_file: str | None,
    headless: bool = False,
    run_dir: Path | None = None,
    prepend_starting_prompt: bool = True,
) -> subprocess.CompletedProcess[str]:
    aspect_folder = "9_16" if aspect_ratio == "9:16" else "4_5"
    prompt_work_dir = RUNTIME_ROOT / "chatgpt_selected_prompts" / f"{batch}_{aspect_folder}_{int(time.time())}_{uuid.uuid4().hex[:8]}"
    prompt_work_dir.mkdir(parents=True, exist_ok=True)

    starting_prompt = ""
    if prepend_starting_prompt:
        starting_prompt_path = ROOT / "input" / "startingprompt.txt"
        starting_prompt = starting_prompt_path.read_text(encoding="utf-8").strip() if starting_prompt_path.exists() else ""
    for prompt_file in prompt_files:
        source = Path(prompt_file)
        if not source.is_absolute():
            source = ROOT / prompt_file
        source = source.resolve()
        if not source.exists():
            raise RuntimeError(f"Prompt file not found: {source}")
        prompt_text = source.read_text(encoding="utf-8")
        combined = f"{starting_prompt}\n\n{prompt_text.strip()}\n" if starting_prompt else prompt_text
        (prompt_work_dir / source.name).write_text(combined, encoding="utf-8")
        sidecar = source.with_suffix(".json")
        if sidecar.exists():
            (prompt_work_dir / sidecar.name).write_text(sidecar.read_text(encoding="utf-8"), encoding="utf-8")

    out_dir = GENERATED_IMAGES_ROOT / batch / aspect_folder
    out_dir.mkdir(parents=True, exist_ok=True)

    cmd = [
        sys.executable,
        "scripts/chatgpt_web_sutomation.py",
        "--prompt-dir",
        str(prompt_work_dir),
        "--prompt-glob",
        "*.txt",
        "--out-dir",
        str(out_dir),
        "--timeout",
        str(int(os.getenv("CHATGPT_GENERATION_TIMEOUT_SECONDS") or "420")),
        "--download-timeout",
        str(int(os.getenv("CHATGPT_DOWNLOAD_TIMEOUT_SECONDS") or "90")),
        "--manual-login-timeout",
        str(int(os.getenv("CHATGPT_MANUAL_LOGIN_TIMEOUT_SECONDS") or "180")),
        "--upload-dir",
        str(INPUT_IMAGES_DIR),
    ]
    if headless:
        cmd.append("--headless")
    if image_sources_file:
        cmd.extend(["--image-source-file", image_sources_file])

    log_dir = RUNTIME_ROOT / "generation_logs"
    log_dir.mkdir(parents=True, exist_ok=True)
    log_path = log_dir / f"gen_{batch}_{aspect_folder}_chatgpt.log"

    env = dashboard_subprocess_env()

    with open(log_path, "w") as log_file:
        result = subprocess.run(cmd, cwd=str(ROOT), text=True, stdout=log_file, stderr=subprocess.STDOUT, check=False, env=env)

    full_output = log_path.read_text() if log_path.exists() else ""
    result.stdout = full_output
    result.stderr = ""
    return result


def build_multipart_form(fields: dict[str, str], file_field: str, file_path: Path) -> tuple[bytes, str]:
    boundary = f"----dashboard{uuid.uuid4().hex}"
    lines: list[bytes] = []

    for key, value in fields.items():
        lines.append(f"--{boundary}\r\n".encode("utf-8"))
        lines.append(f'Content-Disposition: form-data; name="{key}"\r\n\r\n'.encode("utf-8"))
        lines.append(f"{value}\r\n".encode("utf-8"))

    mime_type = mimetypes.guess_type(file_path.name)[0] or "application/octet-stream"
    lines.append(f"--{boundary}\r\n".encode("utf-8"))
    lines.append(
        (
            f'Content-Disposition: form-data; name="{file_field}"; filename="{file_path.name}"\r\n'
            f"Content-Type: {mime_type}\r\n\r\n"
        ).encode("utf-8")
    )
    lines.append(file_path.read_bytes())
    lines.append(b"\r\n")
    lines.append(f"--{boundary}--\r\n".encode("utf-8"))

    body = b"".join(lines)
    content_type = f"multipart/form-data; boundary={boundary}"
    return body, content_type


def upload_image_to_cloudinary(image_path: Path, cloud_name: str, api_key: str, api_secret: str) -> str:
    if not image_path.exists() or not image_path.is_file():
        raise RuntimeError(f"Image not found for upload: {image_path}")

    timestamp = str(int(time.time()))
    signature_base = f"timestamp={timestamp}{api_secret}"
    signature = hashlib.sha1(signature_base.encode("utf-8")).hexdigest()

    fields = {
        "api_key": api_key,
        "timestamp": timestamp,
        "signature": signature,
    }
    body, content_type = build_multipart_form(fields, "file", image_path)
    upload_url = f"https://api.cloudinary.com/v1_1/{cloud_name}/image/upload"
    req = urllib.request.Request(
        url=upload_url,
        data=body,
        method="POST",
        headers={"Content-Type": content_type},
    )
    with urllib.request.urlopen(req, timeout=180) as response:
        raw = response.read().decode("utf-8")
    payload = json.loads(raw)
    secure_url = str(payload.get("secure_url") or "").strip()
    if not secure_url:
        raise RuntimeError(f"Cloudinary upload did not return secure_url: {payload}")
    return secure_url


def load_batch_image_summary(batch: str) -> list[dict[str, Any]]:
    summary_path = GENERATED_IMAGES_ROOT / batch / "batch_run_summary.json"
    if not summary_path.exists():
        jobs_by_prompt: dict[str, dict[str, Any]] = {}
        for generated_root in generated_image_roots():
            generated_batch_dir = generated_root / batch
            if not generated_batch_dir.exists():
                continue
            for meta_file in sorted(generated_batch_dir.glob("**/*.json")):
                try:
                    payload = json.loads(meta_file.read_text(encoding="utf-8"))
                except Exception:
                    continue
                if not isinstance(payload, dict):
                    continue
                rec_type = str(payload.get("type") or payload.get("record_type") or "").strip()
                if rec_type not in ("ad_image", "generated_image"):
                    continue

                prompt_file = str(payload.get("prompt_file_relative") or payload.get("prompt_file") or "").strip().replace("\\", "/")
                saved_file = str(payload.get("saved_file") or "").strip().replace("\\", "/")
                if not prompt_file or not saved_file:
                    continue

                existing = jobs_by_prompt.get(prompt_file)
                if not existing:
                    fmt = payload.get("format") or payload.get("format_id") or ""
                    lang = payload.get("language") or payload.get("lang_id") or ""
                    existing = {
                        "prompt_file": prompt_file,
                        "saved_files": [],
                        "format": fmt,
                        "language": lang,
                        "variation": payload.get("variation"),
                        "task_id": payload.get("task_id"),
                        "prompt_metadata": payload.get("prompt_metadata") or {},
                    }
                    jobs_by_prompt[prompt_file] = existing
                saved_files = existing.get("saved_files")
                if not isinstance(saved_files, list):
                    saved_files = []
                    existing["saved_files"] = saved_files
                if saved_file not in saved_files:
                    saved_files.append(saved_file)

        return list(jobs_by_prompt.values())
    try:
        summary = json.loads(summary_path.read_text(encoding="utf-8"))
    except Exception:
        return []
    jobs = summary.get("jobs")
    if isinstance(jobs, list):
        return [job for job in jobs if isinstance(job, dict)]
    return []


def strip_ansi(text: str) -> str:
    return re.sub(r"\x1B\[[0-?]*[ -/]*[@-~]", "", text)


def opencode_discovery_env() -> dict[str, str]:
    env = os.environ.copy()
    default_xdg = Path.home() / ".local" / "share"
    default_auth = default_xdg / "opencode" / "auth.json"

    raw_xdg = env.get("XDG_DATA_HOME", "").strip()
    current_xdg = Path(raw_xdg).expanduser() if raw_xdg else default_xdg
    current_auth = current_xdg / "opencode" / "auth.json"

    if default_auth.exists() and not current_auth.exists():
        env["XDG_DATA_HOME"] = str(default_xdg)
    return env


def run_opencode_discovery_cmd(cmd: list[str]) -> subprocess.CompletedProcess[str]:
    return subprocess.run(cmd, cwd=str(ROOT), text=True, capture_output=True, check=False, env=opencode_discovery_env())


def list_opencode_models() -> list[str]:
    result = run_opencode_discovery_cmd(["opencode", "models"])
    if result.returncode != 0:
        return []
    lines = [line.strip() for line in strip_ansi(result.stdout).splitlines()]
    return [line for line in lines if line and "/" in line]


def list_opencode_provider_labels() -> list[str]:
    result = run_opencode_discovery_cmd(["opencode", "providers", "list"])
    if result.returncode != 0:
        return []
    lines = [line.strip() for line in strip_ansi(result.stdout).splitlines()]
    labels: list[str] = []
    for line in lines:
        match = re.search(r"[●•]\s+(.+?)\s+(oauth|api|token|key)\b", line, flags=re.IGNORECASE)
        if match:
            value = match.group(1).strip()
        else:
            fallback = re.search(r"^[│\s]*[●•]\s+(.+)$", line)
            if not fallback:
                continue
            value = re.sub(r"\s+(oauth|api|token|key)\b.*$", "", fallback.group(1), flags=re.IGNORECASE).strip()
        if value:
            labels.append(value)
    return labels


def provider_id_from_label(label: str) -> str:
    known = {
        "github copilot": "github-copilot",
        "github-copilot": "github-copilot",
        "opencode": "opencode",
    }
    key = label.strip().lower()
    if key in known:
        return known[key]
    return re.sub(r"[^a-z0-9]+", "-", key).strip("-")


def list_models_for_provider(provider: str) -> list[str]:
    result = run_opencode_discovery_cmd(["opencode", "models", provider])
    if result.returncode != 0:
        return []
    lines = [line.strip() for line in strip_ansi(result.stdout).splitlines()]
    return [line for line in lines if line and line.startswith(provider + "/")]


def choose_openai_gpt52(models: list[str]) -> str:
    if not models:
        return ""
    preferred = "openai/gpt-5.2"
    if preferred in models:
        return preferred
    for model in models:
        lower = model.lower()
        if lower.startswith("openai/") and "gpt-5.2" in lower:
            return model
    for model in models:
        if model.lower().startswith("openai/"):
            return model
    non_copilot = [m for m in models if not m.lower().startswith("github-copilot/")]
    if non_copilot:
        return non_copilot[0]
    return models[0]


def sanitize_dashboard_model(selected: str, models: list[str]) -> str:
    chosen = (selected or "").strip()
    if chosen and (not models or chosen in models):
        return chosen
    return choose_openai_gpt52(models)


def build_opencode_catalog() -> dict[str, Any]:
    models = list_opencode_models()
    provider_labels = list_opencode_provider_labels()
    provider_ids = {line.split("/", 1)[0] for line in models}

    known_providers = ["opencode", "openai"]
    for provider in known_providers:
        provider_ids.add(provider)
    for label in provider_labels:
        pid = provider_id_from_label(label)
        if pid:
            provider_ids.add(pid)

    for provider in sorted(provider_ids):
        if any(model.startswith(provider + "/") for model in models):
            continue
        models.extend(list_models_for_provider(provider))

    providers = sorted(provider for provider in provider_ids if provider.lower() != "github-copilot")
    grouped: dict[str, list[str]] = {provider: [] for provider in providers}
    for model in models:
        provider = model.split("/", 1)[0]
        if provider.lower() == "github-copilot":
            continue
        grouped.setdefault(provider, []).append(model)
    for provider in grouped:
        grouped[provider] = sorted(grouped[provider])
    providers_with_models = [provider for provider, values in grouped.items() if values]
    copilot_models = [model for model in models if model.lower().startswith("github-copilot/")]
    if providers_with_models:
        providers = sorted(providers_with_models)
    elif copilot_models:
        providers = ["github-copilot"]
        grouped = {"github-copilot": sorted(copilot_models)}
    default_model = ""
    if models:
        default_model = choose_openai_gpt52(models)
    return {
        "api_url": DEFAULT_OPENCODE_API_URL,
        "providers": providers,
        "provider_labels": provider_labels,
        "models_by_provider": grouped,
        "default_model": default_model,
    }


def parse_json_stdout(result: subprocess.CompletedProcess[str], context: str) -> Any:
    if result.returncode != 0:
        raise RuntimeError(f"{context} failed: {result.stderr.strip() or result.stdout.strip()}")
    try:
        return json.loads(result.stdout)
    except json.JSONDecodeError as exc:
        raise RuntimeError(f"{context} returned invalid JSON") from exc


def save_upload(target: Path, upload: UploadFile | None) -> Path | None:
    if upload is None or not upload.filename:
        return None
    target.parent.mkdir(parents=True, exist_ok=True)
    data = upload.file.read()
    target.write_bytes(data)
    return target


def coalesce_path(uploaded: Path | None, default_path: Path) -> Path:
    return uploaded if uploaded and uploaded.exists() else default_path


def resolve_safe_path(relative_path: str) -> Path:
    candidate = (ROOT / relative_path).resolve()
    if str(candidate).startswith(str(ROOT.resolve())):
        return candidate
    raise HTTPException(status_code=400, detail="Invalid path")


def choose_text(items: list[str], fallback: str) -> str:
    for item in items:
        clean = item.strip()
        if clean:
            return clean
    return fallback


def shorten_copy_line(text: str) -> str:
    return " ".join((text or "").split()).strip()


def strip_internal_marker(text: str) -> str:
    if not isinstance(text, str):
        return ""
    cleaned = re.sub(r"\s*\b\d{4}-\d{2}-(hero|ba|test|feat|ugc)\.?\b", "", text, flags=re.IGNORECASE)
    cleaned = re.sub(r"\s*\(\s*\d+[_-]\d+\s*\)", "", cleaned)
    cleaned = re.sub(r"\s{2,}", " ", cleaned).strip()
    return cleaned


def strip_price_tokens(text: str) -> str:
    if not isinstance(text, str):
        return ""
    cleaned = text
    cleaned = re.sub(r"\bINR\b\s*\d+[\d,]*(?:\.\d+)?", "", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"[₹$]\s*\d+[\d,]*(?:\.\d+)?", "", cleaned)
    cleaned = re.sub(r"\b\d+[\d,]*(?:\.\d+)?\s*(?:INR|Rs\.?|rupees?)\b", "", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\b(?:price|only|discount|off|mrp)\b", "", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\s{2,}", " ", cleaned).strip(" ,;:-")
    return cleaned


def strip_ba_panel_label(text: str) -> str:
    if not isinstance(text, str):
        return ""
    cleaned = text.strip()
    cleaned = re.sub(r"^\s*(?:before|after)\s*[:\-]\s*", "", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"^\s*(?:पहले|बाद|पहले\s*में|बाद\s*में)\s*[:\-]\s*", "", cleaned)
    cleaned = re.sub(r"\s{2,}", " ", cleaned).strip()
    return cleaned


def strip_internal_markers_from_payload(payload: dict[str, Any]) -> dict[str, Any]:
    ads = payload.get("ads")
    if not isinstance(ads, list):
        return payload

    for ad in ads:
        if not isinstance(ad, dict):
            continue
        copy = ad.get("copy")
        if not isinstance(copy, dict):
            continue
        for lang in ["EN", "HI"]:
            block = copy.get(lang)
            if not isinstance(block, dict):
                continue
            for key in ["headline", "support_line", "cta", "trust_line", "attribution"]:
                if key in block and isinstance(block.get(key), str):
                    value = strip_internal_marker(block[key])
                    block[key] = strip_price_tokens(value)
            if isinstance(block.get("context_line"), str):
                context_line = strip_price_tokens(strip_internal_marker(block["context_line"]))
                if re.search(r"\bneeds\b|proof needed|tone cue|persona", context_line, flags=re.IGNORECASE):
                    block["context_line"] = ""
                else:
                    block["context_line"] = context_line
            if isinstance(block.get("bullets"), list):
                cleaned_bullets = []
                for item in block["bullets"]:
                    if not isinstance(item, str):
                        continue
                    value = strip_price_tokens(strip_internal_marker(item))
                    if value:
                        cleaned_bullets.append(value)
                block["bullets"] = cleaned_bullets
    return payload


CTA_VARIANTS: dict[str, dict[str, list[str]]] = {}


def registry_banlist_values(context: dict[str, Any]) -> set[str]:
    banlist = context.get("banlist") if isinstance(context.get("banlist"), dict) else {}
    buckets = banlist.get("buckets") if isinstance(banlist.get("buckets"), dict) else {}
    values: set[str] = set()
    for arr in buckets.values():
        if not isinstance(arr, list):
            continue
        for item in arr:
            if isinstance(item, str) and item.strip():
                values.add(item.strip())
    registry_path = ROOT / "AD_GENERATION_REGISTRY.JSON"
    if registry_path.exists():
        try:
            registry = json.loads(registry_path.read_text(encoding="utf-8"))
        except Exception:
            registry = {}
        used_text = ((registry.get("indexes") or {}).get("used_text") or {}) if isinstance(registry, dict) else {}
        if isinstance(used_text, dict):
            for arr in used_text.values():
                if not isinstance(arr, list):
                    continue
                for item in arr:
                    if isinstance(item, str) and item.strip():
                        values.add(item.strip())
    return values


def enforce_unique_ctas(payload: dict[str, Any], context: dict[str, Any]) -> dict[str, Any]:
    ads = payload.get("ads") if isinstance(payload.get("ads"), list) else []
    blocked = registry_banlist_values(context)
    seen: set[str] = set()
    for ad in ads:
        if not isinstance(ad, dict):
            continue
        fmt = _clean_str(ad.get("format")).upper()
        copy = ad.get("copy") if isinstance(ad.get("copy"), dict) else {}
        for lang in ["EN", "HI"]:
            block = copy.get(lang) if isinstance(copy.get(lang), dict) else None
            if not block:
                continue
            current = _clean_str(block.get("cta"))
            if current and current not in seen and current not in blocked:
                seen.add(current)
                continue
            variants = CTA_VARIANTS.get(lang, {}).get(fmt, [])
            chosen = ""
            for candidate in variants:
                if candidate not in seen and candidate not in blocked:
                    chosen = candidate
                    break
            if not chosen:
                # Keep CTA natural if uniqueness pool is exhausted.
                chosen = current or (variants[0] if variants else "See Details")
            block["cta"] = chosen
            seen.add(chosen)
    return payload


PROOF_NOTE_MARKERS = [
    "needs",
    "proof needed",
    "tone cue",
    "persona",
    "non-cure",
    "compliant",
    "weight-support framing",
]


def scrub_on_image_copy(payload: dict[str, Any]) -> dict[str, Any]:
    ads = payload.get("ads") if isinstance(payload.get("ads"), list) else []
    for ad in ads:
        if not isinstance(ad, dict):
            continue
        copy = ad.get("copy") if isinstance(ad.get("copy"), dict) else {}
        for lang in ["EN", "HI"]:
            block = copy.get(lang)
            if not isinstance(block, dict):
                continue
            ctx = block.get("context_line")
            if isinstance(ctx, str) and ctx.strip():
                lowered = ctx.lower()
                if any(marker in lowered for marker in PROOF_NOTE_MARKERS):
                    block.pop("context_line", None)
    return payload


def parse_uniqueness_collisions(error_text: str) -> list[dict[str, Any]]:
    collisions: list[dict[str, Any]] = []
    for raw_line in error_text.splitlines():
        line = raw_line.strip()
        match = re.search(r"ads\[(\d+)\]\.copy\.(EN|HI)\.([a-z_]+)", line)
        if not match:
            continue
        collisions.append(
            {
                "ad_index": int(match.group(1)),
                "language": match.group(2),
                "field": match.group(3),
                "line": line,
            }
        )
    return collisions


def parse_json_object_from_text(content: str) -> dict[str, Any] | None:
    text = (content or "").strip()
    if not text:
        return None

    try:
        parsed = json.loads(text)
        if isinstance(parsed, dict):
            return parsed
    except json.JSONDecodeError:
        pass

    fence = re.search(r"```(?:json)?\s*(\{.*?\})\s*```", text, flags=re.DOTALL | re.IGNORECASE)
    if fence:
        try:
            parsed = json.loads(fence.group(1))
            if isinstance(parsed, dict):
                return parsed
        except json.JSONDecodeError:
            pass

    decoder = json.JSONDecoder()
    best: dict[str, Any] | None = None
    best_span = -1
    for match in re.finditer(r"\{", text):
        start = match.start()
        try:
            parsed, end = decoder.raw_decode(text[start:])
        except json.JSONDecodeError:
            continue
        if not isinstance(parsed, dict):
            continue
        span = end
        if span > best_span:
            best = parsed
            best_span = span
    return best


def parse_opencode_json_output(stdout: str) -> dict[str, Any] | None:
    text_chunks: list[str] = []
    for raw_line in (stdout or "").splitlines():
        line = raw_line.strip()
        if not line:
            continue
        try:
            event = json.loads(line)
        except json.JSONDecodeError:
            continue
        if event.get("type") != "text":
            continue
        part = event.get("part") or {}
        text = part.get("text")
        if isinstance(text, str) and text.strip():
            text_chunks.append(text.strip())

    if text_chunks:
        parsed = parse_json_object_from_text("\n".join(text_chunks).strip())
        if parsed is not None:
            return parsed

    return parse_json_object_from_text((stdout or "").strip())


def _find_session_id(value: Any, session_scoped: bool = False) -> str | None:
    if isinstance(value, dict):
        event_type = str(value.get("type") or "").lower()
        scoped = session_scoped or "session" in event_type
        for key in ("sessionID", "sessionId", "session_id"):
            candidate = value.get(key)
            if isinstance(candidate, str) and candidate.strip():
                return candidate.strip()
        if scoped:
            candidate = value.get("id")
            if isinstance(candidate, str) and candidate.strip():
                return candidate.strip()
        for key, nested in value.items():
            nested_scoped = scoped or "session" in str(key).lower()
            found = _find_session_id(nested, nested_scoped)
            if found:
                return found
    elif isinstance(value, list):
        for item in value:
            found = _find_session_id(item, session_scoped)
            if found:
                return found
    return None


def parse_opencode_session_id(stdout: str) -> str | None:
    for raw_line in (stdout or "").splitlines():
        line = raw_line.strip()
        if not line:
            continue
        try:
            event = json.loads(line)
        except json.JSONDecodeError:
            continue
        found = _find_session_id(event)
        if found:
            return found

    match = re.search(r'"session(?:ID|Id|_id)"\s*:\s*"([^"]+)"', stdout or "")
    if match:
        return match.group(1).strip()
    return None


def build_product_doc_bootstrap_prompt() -> str:
    return COPY_PROMPTS.get("product_doc_bootstrap_prompt", "Read the attached product master doc completely. Return only valid JSON: {\"status\":\"product_doc_loaded\"}.")


def append_run_log(run_dir: Path, filename: str, message: str) -> None:
    log_path = run_dir / "logs" / filename
    log_path.parent.mkdir(parents=True, exist_ok=True)
    with log_path.open("a", encoding="utf-8") as handle:
        handle.write(message.rstrip() + "\n")


def call_opencode_repair_copy(
    config: dict[str, Any],
    context: dict[str, Any],
    current_copy: dict[str, Any],
    collisions: list[dict[str, Any]],
    run_dir: Path,
) -> dict[str, Any] | None:
    api_url = (config.get("opencode_api_url") or "").strip()
    model = sanitize_dashboard_model((config.get("opencode_model") or "").strip(), list_opencode_models())
    if not api_url:
        return None

    payload = {
        "task": "Repair uniqueness collisions only",
        "rules": [
            "Return valid JSON only",
            "Keep existing structure and fields",
            "Only change collided fields",
            "Do not use generic repeated support lines",
            "Do not add internal tags or IDs",
        ],
        "collisions": collisions,
        "current_copy": current_copy,
        "context": build_generation_payload_for_llm(context),
    }
    prompt = (
        "You are fixing ad copy JSON after uniqueness collisions. "
        "Return only corrected JSON object with keys default_aspect_ratio and ads.\n\n"
        + json.dumps(payload, ensure_ascii=False)
    )

    password = (config.get("opencode_api_key") or "").strip() or os.getenv("OPENCODE_SERVER_PASSWORD", "").strip()
    cmd = [
        "opencode",
        "run",
        "--pure",
        "--attach",
        api_url,
        "--model",
        model,
        "--format",
        "json",
        prompt,
    ]
    if password:
        cmd.extend(["--password", password])
    try:
        result = run_cmd(cmd, cwd=ROOT)
    except OSError as exc:
        (run_dir / "logs" / "opencode_repair_error.txt").write_text(
            f"Repair command launch failed: {exc}", encoding="utf-8"
        )
        return None
    if result.returncode != 0:
        (run_dir / "logs" / "opencode_repair_error.txt").write_text(
            f"Repair command failed\nSTDOUT:\n{result.stdout}\n\nSTDERR:\n{result.stderr}", encoding="utf-8"
        )
        return None

    return parse_opencode_json_output(result.stdout)


def _clean_str(value: Any) -> str:
    return value.strip() if isinstance(value, str) else ""


def _clean_bullets(value: Any) -> list[str]:
    if not isinstance(value, list):
        return []
    out: list[str] = []
    for item in value:
        if isinstance(item, str) and item.strip():
            out.append(item.strip())
    return out


def concept_ids_from_requirements(copy_req: dict[str, Any]) -> dict[str, str]:
    concept = copy_req.get("concept_variation") if isinstance(copy_req.get("concept_variation"), dict) else {}

    def nested_id(key: str, fallback: str) -> str:
        item = concept.get(key) if isinstance(concept.get(key), dict) else {}
        value = item.get("id") if isinstance(item, dict) else ""
        return _clean_str(value) or fallback

    return {
        "awareness_stage": nested_id("audience_stage", "problem_aware"),
        "concept_angle": nested_id("lead_angle", "desired_outcome"),
        "concept_structure": nested_id("message_structure", "four_us"),
    }


def ensure_testimonial_headline(headline: str, lang: str, persona: dict[str, Any]) -> str:
    clean = shorten_copy_line(headline)
    guidance = COPY_PROMPTS.get("testimonial_headline_guidance", {})
    cfg = guidance.get(lang, guidance.get("EN", {}))
    first_pat = cfg.get("first_person_pattern", "")
    weight_pat = cfg.get("weight_pattern", "")
    suffix = cfg.get("suffix", "")
    desire_template = cfg.get("desire_template", "")
    fallback_text = cfg.get("fallback", "")
    desire_field = cfg.get("desire_field", "")

    if lang == "EN":
        if first_pat and re.search(first_pat, clean, flags=re.IGNORECASE):
            if weight_pat and re.search(weight_pat, clean, flags=re.IGNORECASE):
                return clean
            return shorten_copy_line(f'{clean.rstrip(".")}. {suffix}')
        desire = _clean_str(persona.get(desire_field)).rstrip(".")
        if desire:
            desire_phrase = desire[:1].lower() + desire[1:] if len(desire) > 1 else desire.lower()
            return shorten_copy_line(desire_template.format(desire_phrase=desire_phrase))
        return fallback_text

    if first_pat and re.search(first_pat, clean):
        if weight_pat and re.search(weight_pat, clean):
            return clean
        return shorten_copy_line(f'{clean.rstrip("।")}। {suffix}')
    desire = _clean_str(persona.get(desire_field)).rstrip("।")
    if desire:
        return shorten_copy_line(desire_template.format(desire_phrase=desire))
    return fallback_text


def ensure_testimonial_attribution(attribution: str, lang: str, persona: dict[str, Any], headline: str, trust_line: str) -> str:
    variant_lists = COPY_PROMPTS.get("testimonial_attribution_variants", {})
    variants = variant_lists.get(lang, variant_lists.get("EN", []))
    if not variants:
        return attribution

    current = _clean_str(attribution)
    seed_input = (
        f"{persona.get('number', '')}|{persona.get('name', '')}|{headline}|{trust_line}|{persona.get('pain_en', '')}|{persona.get('pain_hi', '')}"
    )
    digest = hashlib.sha1(seed_input.encode("utf-8", errors="ignore")).hexdigest()
    idx = int(digest[:8], 16) % len(variants)
    chosen = variants[idx]

    if not current:
        return chosen

    # For TEST attribution we avoid personal names and generic repeated labels.
    if re.search(r"\brepresentative\s+user\s+review\b", current, flags=re.IGNORECASE):
        return chosen
    if re.search(r"\b(user|customer)\s+review\b", current, flags=re.IGNORECASE):
        return chosen
    if re.search(r"\b[A-Z][a-z]+\s+[A-Z][a-z]+\b", current):
        return chosen
    if re.search(r"[\u0900-\u097F]{2,}\s+[\u0900-\u097F]{2,}", current):
        return chosen
    return shorten_copy_line(current)


def _persona_number_from_candidate(candidate: dict[str, Any]) -> int | None:
    persona = candidate.get("persona") if isinstance(candidate, dict) else None
    if not isinstance(persona, dict):
        return None
    val = persona.get("number")
    if isinstance(val, int):
        return val
    val = persona.get("persona_number")
    if isinstance(val, int):
        return val
    if isinstance(val, str) and val.strip().isdigit():
        return int(val.strip())
    return None


def _persona_name_from_candidate(candidate: dict[str, Any]) -> str:
    persona = candidate.get("persona") if isinstance(candidate, dict) else None
    if not isinstance(persona, dict):
        return ""
    return _clean_str(persona.get("name") or persona.get("persona_name") or "")


def normalize_generated_copy(
    generated: dict[str, Any] | None,
    context: dict[str, Any],
    run_id: str,
) -> dict[str, Any]:
    base = build_template_copy(context, run_id)
    ads_generated = generated.get("ads") if isinstance(generated, dict) else None
    candidates = ads_generated if isinstance(ads_generated, list) else []
    used_indices: set[int] = set()

    def pick_candidate(fmt: str, persona_num: int, persona_name: str) -> dict[str, Any] | None:
        for idx, cand in enumerate(candidates):
            if idx in used_indices or not isinstance(cand, dict):
                continue
            cand_fmt = _clean_str(cand.get("format")).upper()
            if cand_fmt != fmt:
                continue
            cand_num = _persona_number_from_candidate(cand)
            cand_name = _persona_name_from_candidate(cand)
            if cand_num == persona_num or (cand_name and cand_name.lower() == persona_name.lower()):
                used_indices.add(idx)
                return cand

        for idx, cand in enumerate(candidates):
            if idx in used_indices or not isinstance(cand, dict):
                continue
            cand_fmt = _clean_str(cand.get("format")).upper()
            if cand_fmt == fmt:
                used_indices.add(idx)
                return cand
        return None

    for ad in base.get("ads", []):
        fmt = _clean_str(ad.get("format")).upper()
        persona = ad.get("persona") or {}
        persona_num = int(persona.get("number"))
        persona_name = _clean_str(persona.get("name"))
        candidate = pick_candidate(fmt, persona_num, persona_name)
        if not candidate:
            continue

        hypothesis = candidate.get("hypothesis") if isinstance(candidate.get("hypothesis"), dict) else None
        if hypothesis:
            ad["hypothesis"] = hypothesis

        angle = _clean_str(candidate.get("headline_angle"))
        if angle:
            ad["headline_angle"] = angle

        for key in ["awareness_stage", "concept_angle", "concept_structure"]:
            value = _clean_str(candidate.get(key))
            if value:
                ad[key] = value

        cand_copy = candidate.get("copy") if isinstance(candidate.get("copy"), dict) else {}
        for lang in ["EN", "HI"]:
            base_lang = ad["copy"][lang]
            src_lang = cand_copy.get(lang) if isinstance(cand_copy.get(lang), dict) else {}

            headline = _clean_str(src_lang.get("headline"))
            cta = _clean_str(src_lang.get("cta"))
            if headline:
                base_lang["headline"] = shorten_copy_line(headline)
            if cta:
                base_lang["cta"] = cta

            if fmt == "TEST":
                base_lang["headline"] = ensure_testimonial_headline(base_lang.get("headline", ""), lang, persona)
                base_lang["attribution"] = ensure_testimonial_attribution(
                    _clean_str(src_lang.get("attribution")),
                    lang,
                    persona,
                    base_lang.get("headline", ""),
                    _clean_str(src_lang.get("trust_line")) or base_lang.get("trust_line", ""),
                )

            if fmt in {"HERO", "UGC"}:
                support = _clean_str(src_lang.get("support_line"))
                if support:
                    base_lang["support_line"] = shorten_copy_line(support)
            elif fmt in {"BA", "FEAT"}:
                bullets = _clean_bullets(src_lang.get("bullets"))
                if len(bullets) >= 2:
                    if fmt == "BA":
                        bullets = [strip_ba_panel_label(b) for b in bullets]
                    base_lang["bullets"] = [shorten_copy_line(b) for b in bullets]
            else:
                trust = _clean_str(src_lang.get("trust_line"))
                if trust:
                    base_lang["trust_line"] = shorten_copy_line(trust)

    return base


def _template_copy(primary_key: str, secondary_key: str, concept_angle: str, pain: str, lang: str) -> str:
    templates = COPY_PROMPTS.get("template_copy_headline_sentence", {})
    template = templates.get(lang, templates.get("EN", ""))
    return template.format(primary_key=primary_key, secondary_key=secondary_key, pain=pain)


def template_headline(primary_key: str, concept_angle: str, pain: str, lang: str) -> str:
    return _template_copy(primary_key, concept_angle, concept_angle, pain, lang)


def template_support(primary_key: str, secondary_key: str, lang: str) -> str:
    templates = COPY_PROMPTS.get("template_copy_support_sentence", {})
    template = templates.get(lang, templates.get("EN", ""))
    return template.format(primary_key=primary_key, secondary_key=secondary_key)


def feature_template(key: str) -> dict[str, str]:
    templates = COPY_PROMPTS.get("feature_templates", {})
    entry = templates.get(key)
    if entry:
        return {"support_en": entry.get("EN", ""), "support_hi": entry.get("HI", "")}
    default = templates.get("_default", {})
    return {"support_en": default.get("EN", "Structured system for consistent progress."), "support_hi": default.get("HI", "लगातार प्रगति के लिए व्यवस्थित सिस्टम।")}


def build_template_copy(context: dict[str, Any], run_id: str) -> dict[str, Any]:
    ads: list[dict[str, Any]] = []
    token = run_id[-4:]
    for idx, item in enumerate(context["ads"], start=1):
        persona = item["persona"]
        fmt = item["format"]
        persona_num = int(persona["persona_number"])
        persona_name = persona["persona_name"]
        unique = f"{token}-{idx:02d}-{fmt.lower()}"
        copy_req = item.get("copy_requirements") if isinstance(item.get("copy_requirements"), dict) else {}
        concept_ids = concept_ids_from_requirements(copy_req)
        primary_key = _clean_str(copy_req.get("primary_feature_key")) or "structured_system"
        secondary_key = _clean_str(copy_req.get("secondary_feature_key")) or "cravings_down"

        pain_en = choose_text(persona.get("pain_points", []), f"Daily routine feels heavy and hard to sustain for persona {persona_num}.")
        desire_en = choose_text(persona.get("core_message", []), "A practical routine that feels easy to follow.")
        friction_en = choose_text(persona.get("objections", []), "Past plans felt too strict and difficult to maintain.")
        proof_en = choose_text(persona.get("trust_anchors", []), "Needs proof through clear structure and believable support.")
        en_fallbacks = COPY_PROMPTS.get("template_copy_en_fallbacks", {})
        tone_en = en_fallbacks.get("tone", "Practical, empathetic, and confidence-building")

        hi_fallbacks = COPY_PROMPTS.get("template_copy_hi_fallbacks", {})
        pain_hi = hi_fallbacks.get("pain", "रोज की वजन-घटाने की दिनचर्या टूटना आसान है।")
        desire_hi = hi_fallbacks.get("desire", "ऐसा आसान सिस्टम चाहिए जो रोज निभ सके।")
        friction_hi = hi_fallbacks.get("friction", "पहले के प्लान बहुत सख्त और मुश्किल थे।")
        proof_hi = hi_fallbacks.get("proof", "साफ कदम, भरोसेमंद सपोर्ट और व्यावहारिक प्रमाण चाहिए।")
        tone_hi = hi_fallbacks.get("tone", "सरल, भरोसेमंद, और व्यावहारिक")

        concept_angle = concept_ids["concept_angle"]
        headline_en = template_headline(primary_key, concept_angle, pain_en, "EN")
        headline_hi = template_headline(primary_key, concept_angle, pain_en, "HI")
        fmt_overrides = COPY_PROMPTS.get("template_copy_format_overrides", {})
        fo = fmt_overrides.get(fmt, {})
        if fo.get("headline"):
            headline_en = fo["headline"].get("EN", headline_en)
            headline_hi = fo["headline"].get("HI", headline_hi)

        cta_map = COPY_PROMPTS.get("template_copy_cta_map", {})
        default_cta = cta_map.get("_default", {"EN": "Start Today", "HI": "आज शुरू करें"})
        fmt_cta = cta_map.get(fmt, default_cta)
        cta_en = fmt_cta.get("EN", "Start Today")
        cta_hi = fmt_cta.get("HI", "आज शुरू करें")

        copy_en: dict[str, Any]
        copy_hi: dict[str, Any]
        if fmt in {"HERO", "UGC"}:
            support_en = template_support(primary_key, secondary_key, "EN")
            support_hi = template_support(primary_key, secondary_key, "HI")
            if fo.get("support_override"):
                support_en = fo["support_override"].get("EN", support_en)
                support_hi = fo["support_override"].get("HI", support_hi)
            copy_en = {"headline": headline_en, "support_line": support_en, "cta": cta_en}
            copy_hi = {"headline": headline_hi, "support_line": support_hi, "cta": cta_hi}
        elif fmt == "BA":
            bullets_en = [
                pain_en.rstrip("."),
                friction_en.rstrip("."),
                feature_template(primary_key)["support_en"].rstrip("."),
                feature_template(secondary_key)["support_en"].rstrip("."),
            ]
            bullets_hi = [
                pain_hi.rstrip("।"),
                friction_hi.rstrip("।"),
                feature_template(primary_key)["support_hi"].rstrip("।"),
                feature_template(secondary_key)["support_hi"].rstrip("।"),
            ]
            copy_en = {"headline": headline_en, "bullets": bullets_en, "cta": cta_en}
            copy_hi = {"headline": headline_hi, "bullets": bullets_hi, "cta": cta_hi}
        elif fmt == "FEAT":
            feat_bullets_en = fo.get("bullets", {}).get("EN", [])
            feat_bullets_hi = fo.get("bullets", {}).get("HI", [])
            bullets_en = feat_bullets_en or [
                "Morning OK Liquid helps reduce hunger and random snacking for weight loss.",
                "Night Tablet + Powder support digestion and lighter mornings in obesity routine.",
                "Built for visible 15-day weight-loss support without crash-diet pressure.",
            ]
            bullets_hi = feat_bullets_hi or [
                "सुबह का OK Liquid वजन घटाने के लिए भूख और अचानक खाने की आदत कम करने में सहायक है।",
                "रात का Tablet + Powder मोटापा-नियंत्रण दिनचर्या में पाचन-सपोर्ट देता है।",
                "कठोर डाइट दबाव के बिना 15 दिन के वजन-सपोर्ट के लिए बनाया गया।",
            ]
            copy_en = {"headline": headline_en, "bullets": bullets_en, "cta": cta_en}
            copy_hi = {"headline": headline_hi, "bullets": bullets_hi, "cta": cta_hi}
        else:
            attribution_en = fo.get("attribution", {}).get("EN", "Doctor-formulated Ayurvedic obesity and weight-loss protocol")
            attribution_hi = fo.get("attribution", {}).get("HI", "डॉक्टर-फॉर्मुलेटेड आयुर्वेदिक मोटापा और वजन-घटाने का प्रोटोकॉल")
            trust_en = fo.get("trust_line", {}).get("EN", "Structured morning-night steps for visible weight-loss progress and obesity control.")
            trust_hi = fo.get("trust_line", {}).get("HI", "सुबह-रात के स्पष्ट कदमों से वजन घटाने और मोटापा नियंत्रण का भरोसेमंद सपोर्ट।")
            copy_en = {
                "headline": headline_en,
                "attribution": attribution_en,
                "trust_line": trust_en,
                "cta": cta_en,
            }
            copy_hi = {
                "headline": headline_hi,
                "attribution": attribution_hi,
                "trust_line": trust_hi,
                "cta": cta_hi,
            }

        ad_payload = {
            "format": fmt,
            "headline_angle": "mechanism",
            "awareness_stage": concept_ids["awareness_stage"],
            "concept_angle": concept_ids["concept_angle"],
            "concept_structure": concept_ids["concept_structure"],
            "persona": {
                "number": persona_num,
                "name": persona_name,
                "pain_en": pain_en,
                "desire_en": desire_en,
                "friction_en": friction_en,
                "proof_needed_en": proof_en,
                "tone_cue_en": tone_en,
                "pain_hi": pain_hi,
                "desire_hi": desire_hi,
                "friction_hi": friction_hi,
                "proof_needed_hi": proof_hi,
                "tone_cue_hi": tone_hi,
            },
            "copy": {"EN": copy_en, "HI": copy_hi},
        }
        hypothesis = item.get("hypothesis") if isinstance(item.get("hypothesis"), dict) else None
        if hypothesis:
            ad_payload["hypothesis"] = hypothesis
        for key in [
            "visual_archetype",
            "visual_pattern_reused_from_run_id",
            "visual_pattern_reuse_key",
            "creative_index",
            "creative_total",
            "background_group_key",
        ]:
            if key in item:
                ad_payload[key] = item[key]
        ads.append(ad_payload)

    return {"default_aspect_ratio": "4:5", "ads": ads}


def call_opencode_compatible(config: dict[str, Any], context: dict[str, Any], run_dir: Path) -> dict[str, Any] | None:
    api_url = (config.get("opencode_api_url") or "").strip()
    api_key = (config.get("opencode_api_key") or "").strip() or os.getenv("OPENCODE_SERVER_PASSWORD", "").strip()
    model = sanitize_dashboard_model((config.get("opencode_model") or "").strip(), list_opencode_models())
    if not api_url:
        return None

    print(f"[call_opencode_compatible] api_url={api_url}, model={model}", file=sys.stderr)

    language_mode = resolve_language_mode(config)
    cli_password = api_key or os.getenv("OPENCODE_SERVER_PASSWORD", "").strip()
    product_file = Path(str(context.get("product_file_path") or DEFAULT_PRODUCT_MASTER))
    generated_ads: list[dict[str, Any]] = []
    errors: list[str] = []
    warnings: list[str] = []
    session_id: str | None = None
    session_fallback_used = False

    if not product_file.exists() or not product_file.is_file():
        errors.append(f"Product master doc missing: {product_file}")
        (run_dir / "logs" / "opencode_error.txt").write_text("\n\n---\n\n".join(errors), encoding="utf-8")
        return None

    env = dashboard_subprocess_env()

    def build_cmd(prompt: str, *, use_session: bool, attach_product_doc: bool) -> list[str]:
        cmd = [
            "opencode",
            "run",
            "--pure",
            "--attach",
            api_url,
            "--model",
            model,
            "--format",
            "json",
        ]
        if use_session and session_id:
            cmd.extend(["--session", session_id])
        if attach_product_doc:
            cmd.extend(["--file", str(product_file)])
        if cli_password:
            cmd.extend(["--password", cli_password])
        cmd.extend(["--", prompt])
        return cmd

    def run_opencode(prompt: str, *, force_file: bool = False) -> tuple[dict[str, Any] | None, str, str, int]:
        use_session = bool(session_id) and not force_file
        cmd = build_cmd(prompt, use_session=use_session, attach_product_doc=force_file or not use_session)
        result = subprocess.run(cmd, cwd=str(ROOT), text=True, capture_output=True, check=False, env=env)
        stdout = result.stdout or ""
        stderr = result.stderr or ""
        if result.returncode != 0:
            return None, stdout, stderr, result.returncode
        parsed = parse_opencode_json_output(stdout)
        return (extract_generated_ad_candidate(parsed) if parsed else None), stdout, stderr, result.returncode

    with _opencode_queue_slot(f"copy_session {run_dir.name}"):
        append_run_log(
            run_dir,
            "opencode_session.log",
            f"{now_iso()} Starting OpenCode product-doc session with file: {product_file}",
        )
        bootstrap_cmd = build_cmd(build_product_doc_bootstrap_prompt(), use_session=False, attach_product_doc=True)
        bootstrap = subprocess.run(bootstrap_cmd, cwd=str(ROOT), text=True, capture_output=True, check=False, env=env)
        append_run_log(
            run_dir,
            "opencode_session.log",
            (
                f"{now_iso()} Bootstrap return_code={bootstrap.returncode}\n"
                f"STDOUT:\n{bootstrap.stdout or ''}\nSTDERR:\n{bootstrap.stderr or ''}"
            ),
        )
        if bootstrap.returncode == 0:
            session_id = parse_opencode_session_id(bootstrap.stdout or "")
            if session_id:
                append_run_log(run_dir, "opencode_session.log", f"{now_iso()} Reusing OpenCode session: {session_id}")
            else:
                session_fallback_used = True
                warning = "OpenCode did not expose a session id; falling back to attaching product master doc on every ad request."
                warnings.append(warning)
                append_run_log(run_dir, "opencode_session.log", f"{now_iso()} FALLBACK: {warning}")
        else:
            session_fallback_used = True
            warning = "OpenCode product-doc session bootstrap failed; falling back to attaching product master doc on every ad request."
            warnings.append(warning)
            append_run_log(run_dir, "opencode_session.log", f"{now_iso()} FALLBACK: {warning}")

        for index, ad_item in enumerate(context.get("ads") or [], start=1):
            previous_same_format: list[dict[str, Any]] = []
            target_format = str(ad_item.get("format") or "").strip().upper()
            for prev in generated_ads:
                if not isinstance(prev, dict):
                    continue
                if str(prev.get("format") or "").strip().upper() != target_format:
                    continue
                prev_copy = prev.get("copy") if isinstance(prev.get("copy"), dict) else {}
                prev_en = prev_copy.get("EN") if isinstance(prev_copy.get("EN"), dict) else {}
                previous_same_format.append(
                    {
                        "persona": (prev.get("persona") or {}).get("name") if isinstance(prev.get("persona"), dict) else "",
                        "headline_angle": prev.get("headline_angle"),
                        "headline": prev_en.get("headline"),
                        "support_line": prev_en.get("support_line"),
                        "bullets": prev_en.get("bullets") if isinstance(prev_en.get("bullets"), list) else [],
                    }
                )
            single_context = {
                **context,
                "ads": [ad_item],
            }
            target_langs = {"EN": ["EN"], "HI": ["HI"], "HINGLISH": ["HINGLISH"], "ALL": ["EN", "HI", "HINGLISH"]}
            user_payload = {
                "task": "Generate fresh ad copy JSON for provided context.",
                "context": build_generation_payload_for_llm(single_context),
                "already_used_ads_DO_NOT_REUSE": previous_same_format,
                "constraints": {
                    "language": target_langs.get(language_mode, ["EN", "HI"]),
                    "language_mode": language_mode,
                    "format": target_format,
                    "return_json_only": True,
                },
            }
            cli_prompt = (
                "SYSTEM:\n"
                f"{build_ad_copy_system_prompt(target_format)}\n\n"
                "USER_PAYLOAD_JSON:\n"
                f"{json.dumps(user_payload, ensure_ascii=False)}\n\n"
                f"{build_ad_prompt_tail(target_format)}"
            )

            try:
                candidate, last_stdout, last_stderr, last_code = run_opencode(cli_prompt)
            except OSError as exc:
                errors.append(f"Ad {index}: launch failed: {exc}")
                continue

            if last_code != 0 and session_id:
                session_id = None
                session_fallback_used = True
                warning = f"OpenCode reusable session failed on ad {index}; falling back to product-doc file attachment for remaining requests."
                warnings.append(f"{warning}\nSTDOUT:\n{last_stdout}\nSTDERR:\n{last_stderr}")
                append_run_log(run_dir, "opencode_session.log", f"{now_iso()} FALLBACK: {warning}")
                candidate, last_stdout, last_stderr, last_code = run_opencode(cli_prompt, force_file=True)

            if not candidate:
                retry_prompt = f"{cli_prompt}\n\n{build_strict_schema_note(target_format)}\n"
                candidate, last_stdout, last_stderr, last_code = run_opencode(retry_prompt, force_file=session_fallback_used and not session_id)

            mismatch = hypothesis_mismatch(candidate, ad_item) if candidate else None
            should_retry_hypothesis = bool(config.get("retry_hypothesis_mismatch"))
            if mismatch and should_retry_hypothesis:
                retry_prompt = (
                    f"{cli_prompt}\n\n"
                    f"REVISION_REQUIRED: {mismatch}\n"
                    "Rewrite only this ad so it satisfies the requested hypothesis while keeping schema valid.\n"
                )
                candidate, last_stdout, last_stderr, last_code = run_opencode(retry_prompt, force_file=session_fallback_used and not session_id)
                mismatch_after = hypothesis_mismatch(candidate, ad_item) if candidate else None
                if mismatch_after:
                    warnings.append(f"Ad {index}: hypothesis retry mismatch persisted; accepting generated copy: {mismatch_after}\nSTDOUT:\n{last_stdout}\nSTDERR:\n{last_stderr}")
            elif mismatch:
                warnings.append(f"Ad {index}: hypothesis mismatch accepted without retry to avoid extra LLM token spend: {mismatch}")

            if not candidate:
                errors.append(f"Ad {index}: returned no usable ad JSON; return code {last_code}\nSTDOUT:\n{last_stdout}\nSTDERR:\n{last_stderr}")
                continue
            generated_ads.append(hydrate_generated_ad_candidate(candidate, ad_item))

    if errors or warnings:
        (run_dir / "logs" / "opencode_error.txt").write_text("\n\n---\n\n".join(errors + warnings), encoding="utf-8")

    if not generated_ads:
        return None

    result_payload: dict[str, Any] = {"default_aspect_ratio": "4:5", "ads": generated_ads}
    if errors:
        result_payload["_opencode_failures"] = errors
    if warnings:
        result_payload["_opencode_warnings"] = warnings
    if session_fallback_used:
        result_payload["_opencode_session_fallback"] = True
    return result_payload


def collect_run_result(run_dir: Path, batch_name: str, image_generated: bool) -> dict[str, Any]:
    output_dir = ROOT / "output" / batch_name
    prompt_files = []
    if output_dir.exists():
        for file in sorted(output_dir.glob("**/OUTPUT_*.txt")):
            prompt_files.append(str(file.relative_to(ROOT)))

    image_files: list[str] = []
    if image_generated:
        for generated_root in generated_image_roots():
            image_dir = generated_root / batch_name
            if not image_dir.exists():
                continue
            for ext in ("*.png", "*.jpg", "*.jpeg", "*.webp"):
                for file in sorted(image_dir.glob(f"**/{ext}")):
                    image_files.append(str(file.relative_to(ROOT)))

    result = {
        "run_id": run_dir.name,
        "batch": batch_name,
        "prompt_files": prompt_files,
        "image_files": image_files,
        "image_generated": image_generated,
        "updated_at": now_iso(),
    }
    (run_dir / "manifest.json").write_text(json.dumps(result, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return result


def _collect_aspect_ratio_images(batch_name: str, aspect_ratio: str) -> list[str]:
    """Collect generated image paths for a specific aspect ratio.

    Searches both legacy and new unified output roots, looking under
    generated_images/{batch}/{aspect}/ for image files recursively.
    """
    aspect_folder = "4_5" if aspect_ratio == "4:5" else "9_16"
    image_files: list[str] = []
    for generated_root in generated_image_roots():
        image_dir = generated_root / batch_name / aspect_folder
        if not image_dir.exists():
            continue
        for ext in ("*.png", "*.jpg", "*.jpeg", "*.webp"):
            for file in sorted(image_dir.glob(f"**/{ext}")):
                rel = str(file.relative_to(ROOT)).replace("\\", "/")
                if "/debug/" in rel or "/.browser_downloads/" in rel:
                    continue
                image_files.append(rel)
    return image_files


def scan_prompt_files_for_batch(batch_name: str) -> list[str]:
    output_dir = ROOT / "output" / batch_name
    prompt_files: list[str] = []
    if not output_dir.exists():
        return prompt_files
    for file in sorted(output_dir.glob("**/OUTPUT_*.txt")):
        prompt_files.append(str(file.relative_to(ROOT)))
    return prompt_files


def scan_image_files_for_batch(batch_name: str) -> list[str]:
    image_files: list[str] = []
    seen: set[str] = set()
    for generated_root in generated_image_roots():
        image_dir = generated_root / batch_name
        if not image_dir.exists():
            continue
        for ext in ("*.png", "*.jpg", "*.jpeg", "*.webp"):
            for file in sorted(image_dir.glob(f"**/{ext}")):
                rel = str(file.relative_to(ROOT)).replace("\\", "/")
                if "/debug/" in rel or "/.browser_downloads/" in rel:
                    continue
                if rel in seen:
                    continue
                seen.add(rel)
                image_files.append(rel)
    return image_files


def refresh_manifest_file_state(run_dir: Path, manifest: dict[str, Any]) -> dict[str, Any]:
    batch_name = str(manifest.get("batch") or "").strip()
    if not batch_name:
        return manifest

    prompt_files = scan_prompt_files_for_batch(batch_name)
    image_files = scan_image_files_for_batch(batch_name)
    image_generated = bool(image_files) or bool(manifest.get("image_generated", False))
    previous_prompt_files = list(manifest.get("prompt_files") or [])
    previous_image_files = list(manifest.get("image_files") or [])
    if (
        previous_prompt_files == prompt_files
        and previous_image_files == image_files
        and bool(manifest.get("image_generated", False)) == image_generated
    ):
        return manifest

    newest_mtime = 0.0
    for rel in prompt_files + image_files:
        try:
            path = ROOT / rel
            if path.exists():
                newest_mtime = max(newest_mtime, path.stat().st_mtime)
        except Exception:
            pass
    updated_at = (
        datetime.fromtimestamp(newest_mtime, tz=timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")
        if newest_mtime > 0
        else now_iso()
    )
    refreshed = {
        "run_id": run_dir.name,
        "batch": batch_name,
        "prompt_files": prompt_files,
        "image_files": image_files,
        "image_generated": image_generated,
        "updated_at": updated_at,
    }
    merged = {**manifest, **refreshed}
    (run_dir / "manifest.json").write_text(json.dumps(merged, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return merged


def force_aspect_ratio(copy_json: dict[str, Any], aspect_ratio: str) -> dict[str, Any]:
    cloned = json.loads(json.dumps(copy_json, ensure_ascii=False))
    cloned["default_aspect_ratio"] = aspect_ratio
    ads = cloned.get("ads")
    if isinstance(ads, list):
        for ad in ads:
            if isinstance(ad, dict):
                ad["aspect_ratio"] = aspect_ratio
    return cloned


def _parse_prompt_field(prompt_text: str, label: str) -> str:
    match = re.search(rf"^\s*-\s*{re.escape(label)}:\s*(.+)$", prompt_text, flags=re.MULTILINE)
    return (match.group(1).strip() if match else "")


def parse_background_lock_from_prompt(prompt_text: str) -> tuple[str, int] | None:
    slot_match = re.search(r"^\s*-\s*Background\s+slot:\s*(BG-\d{3})\b", prompt_text, flags=re.MULTILINE | re.IGNORECASE)
    seed_match = re.search(r"^\s*-\s*Background\s+seed:\s*(\d+)\s*$", prompt_text, flags=re.MULTILINE | re.IGNORECASE)
    if not slot_match or not seed_match:
        return None
    return (slot_match.group(1).upper(), int(seed_match.group(1)))


def parse_prompt_filename(prompt_path: str) -> tuple[str, str, int | None] | None:
    name = Path(prompt_path).name
    match = re.match(r"^OUTPUT_([A-Z]+)(?:_P(\d+))?_(EN|HI|HINGLISH)(?:_(?:V|A)\d+)?\.txt$", name)
    if not match:
        return None
    persona_raw = match.group(2)
    persona_number = int(persona_raw) if persona_raw else None
    return (match.group(1), match.group(3), persona_number)


def parse_prompt_creative_index(prompt_path: str) -> int:
    match = re.search(r"_A(\d+)\.txt$", Path(prompt_path).name, flags=re.IGNORECASE)
    return int(match.group(1)) if match else 1


def parse_persona_number_from_prompt(prompt_text: str) -> int | None:
    match = re.search(r"\(\s*Persona\s*(\d+)\s*\)", prompt_text, flags=re.IGNORECASE)
    if not match:
        return None
    return int(match.group(1))


EXACT_COPY_BLOCK_RE = re.compile(
    r"EXACT ON-IMAGE COPY - DO NOT ALTER ANYTHING\s*\n(?P<block>.+?)\n\s*Render every character exactly as written",
    flags=re.DOTALL,
)


def extract_on_image_copy_lines(prompt_text: str) -> list[dict[str, str]]:
    """
    Legacy-ish extractor used by the dashboard editor.

    It DOES NOT preserve exact spacing/linebreaks; it trims lines into {label,value}.
    Keep this for backward compatibility.
    """
    block = EXACT_COPY_BLOCK_RE.search(prompt_text)
    if not block:
        return []

    out: list[dict[str, str]] = []
    for line in block.group("block").splitlines():
        raw = line.strip()
        if not raw:
            continue
        parsed = re.match(r"^-\s*([^:]+):\s*(.*)$", raw)
        if not parsed:
            continue
        out.append({"label": parsed.group(1).strip(), "value": parsed.group(2).strip()})
    return out


def extract_exact_on_image_copy_block(prompt_text: str, *, warn_log_path: Path | None = None) -> str | None:
    """
    Task 5: Extract ONLY the content inside:
      EXACT ON-IMAGE COPY - DO NOT ALTER ANYTHING
      ...
      Render every character exactly as written

    Rules:
    - preserve exact text including punctuation/case/spacing/line breaks
    - no normalization (no strip, no join)
    - if block missing: optionally log warning; return None
    """
    pattern = (
        r"EXACT ON-IMAGE COPY - DO NOT ALTER ANYTHING\s*\n"
        r"(?P<block>.+?)\n\s*Render every character exactly as written"
    )
    m = re.search(pattern, prompt_text, flags=re.DOTALL)
    if not m:
        if warn_log_path is not None:
            warn_log_path.parent.mkdir(parents=True, exist_ok=True)
            warn_log_path.write_text(
                "WARNING: EXACT ON-IMAGE COPY block missing; skipping this prompt.\n",
                encoding="utf-8",
            )
        return None

    # Return exactly what was captured: no strip().
    return m.group("block")


def load_run_language_mode(run_dir: Path) -> str:
    run_context_path = run_dir / "context" / "run_context.json"
    assembler_mode = "BOTH"
    if not run_context_path.exists():
        return assembler_mode
    try:
        run_context = json.loads(run_context_path.read_text(encoding="utf-8"))
        lang_mode = str(run_context.get("language_mode") or "ALL").upper()
        if lang_mode == "EN":
            return "EN"
        if lang_mode == "HI":
            return "HI"
    except Exception:
        return assembler_mode
    return assembler_mode


def rerender_prompts_for_run(run_dir: Path, batch: str, copy_file: Path, language_mode: str) -> None:
    result = run_cmd(
        [
            "python3",
            "scripts/generate_ads.py",
            "--copy-file",
            str(copy_file),
            "--batch",
            batch,
            "--language-mode",
            language_mode,
            "--no-registry-write",
            "--skip-uniqueness-check",
        ],
        cwd=ROOT,
    )
    if result.returncode != 0:
        error_text = result.stderr or result.stdout
        (run_dir / "logs" / "assembler_edit_error.txt").write_text(error_text, encoding="utf-8")
        short_error = "\n".join([line for line in error_text.splitlines() if line.strip()][-12:])
        raise HTTPException(status_code=500, detail=f"Prompt regeneration failed: {short_error}")


def merge_manifest(run_dir: Path, previous_manifest: dict[str, Any], refreshed: dict[str, Any]) -> dict[str, Any]:
    merged = {**previous_manifest, **refreshed}
    (run_dir / "manifest.json").write_text(json.dumps(merged, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return merged


def generate_916_for_run(run_dir: Path, manifest: dict[str, Any]) -> dict[str, Any]:
    copy_path = run_dir / "context" / "copy_batch.json"
    if not copy_path.exists():
        raise HTTPException(status_code=404, detail="copy_batch.json not found for run")

    batch = (manifest.get("batch") or "").strip()
    if not batch:
        raise HTTPException(status_code=400, detail="Run has no batch folder")

    copy_json = json.loads(copy_path.read_text(encoding="utf-8"))
    copy_916 = force_aspect_ratio(copy_json, "9:16")
    visual_locks = collect_45_visual_locks(batch)
    if visual_locks:
        copy_916 = apply_visual_locks(copy_916, visual_locks)
    copy_916_path = run_dir / "context" / "copy_batch_916.json"
    copy_916_path.write_text(json.dumps(copy_916, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    assembler_mode = load_run_language_mode(run_dir)
    result = run_cmd(
        [
            "python3",
            "scripts/generate_ads.py",
            "--copy-file",
            str(copy_916_path),
            "--batch",
            batch,
            "--language-mode",
            assembler_mode,
            "--no-registry-write",
            "--skip-uniqueness-check",
        ],
        cwd=ROOT,
    )

    if result.returncode != 0:
        error_text = result.stderr or result.stdout
        (run_dir / "logs" / "assembler_916_error.txt").write_text(error_text, encoding="utf-8")
        short_error = "\n".join([line for line in error_text.splitlines() if line.strip()][-12:])
        raise HTTPException(status_code=500, detail=f"9:16 generation failed: {short_error}")

    refreshed = collect_run_result(run_dir, batch, bool(manifest.get("image_generated", False)))
    refreshed["generated_variant"] = "9:16"
    return merge_manifest(run_dir, manifest, refreshed)


def collect_45_visual_locks(batch: str) -> dict[str, dict[str, Any]]:
    out: dict[str, dict[str, Any]] = {}
    ratio_dir = ROOT / "output" / batch / "45"
    if not ratio_dir.exists():
        return out
    for prompt_file in sorted(ratio_dir.glob("OUTPUT_*_EN.txt")) + sorted(ratio_dir.glob("OUTPUT_*_HI.txt")):
        parsed = parse_prompt_filename(prompt_file.name)
        if not parsed:
            continue
        fmt, _lang, persona_number = parsed
        key = f"{fmt}::P{persona_number}" if isinstance(persona_number, int) else fmt
        current = out.get(key, {})
        text = prompt_file.read_text(encoding="utf-8", errors="ignore")
        if persona_number is None:
            inferred = parse_persona_number_from_prompt(text)
            if isinstance(inferred, int):
                persona_number = inferred
                key = f"{fmt}::P{persona_number}"
                current = out.get(key, current)
        lock = parse_background_lock_from_prompt(text)
        if lock:
            current["background_slot"] = lock[0]
            current["background_seed"] = lock[1]

        visual_lock = {
            "seeded_background_direction": _parse_prompt_field(text, "Seeded background direction (single sentence, exact)"),
            "subject": _parse_prompt_field(text, "Subject"),
            "action": _parse_prompt_field(text, "Action"),
            "camera": _parse_prompt_field(text, "Camera"),
            "lighting": _parse_prompt_field(text, "Lighting"),
            "props": _parse_prompt_field(text, "Props"),
            "surfaces": _parse_prompt_field(text, "Surfaces"),
            "mood": _parse_prompt_field(text, "Mood"),
            "realism": _parse_prompt_field(text, "Realism"),
        }
        visual_lock = {k: v for k, v in visual_lock.items() if v}
        if visual_lock:
            current["visual_lock"] = visual_lock

        if current:
            out[key] = current
    return out


def apply_visual_locks(copy_json: dict[str, Any], locks: dict[str, dict[str, Any]]) -> dict[str, Any]:
    cloned = json.loads(json.dumps(copy_json, ensure_ascii=False))
    ads = cloned.get("ads")
    if not isinstance(ads, list):
        return cloned
    for ad in ads:
        if not isinstance(ad, dict):
            continue
        fmt = str(ad.get("format") or "").strip().upper()
        persona_no = None
        persona = ad.get("persona")
        if isinstance(persona, dict):
            raw_no = persona.get("number")
            if isinstance(raw_no, int):
                persona_no = raw_no
        lock_key = f"{fmt}::P{persona_no}" if isinstance(persona_no, int) else ""
        lock = (locks.get(lock_key) if lock_key else None) or locks.get(fmt) or {}
        if not lock:
            continue
        if isinstance(lock.get("background_slot"), str):
            ad["background_slot"] = lock["background_slot"]
        if isinstance(lock.get("background_seed"), int):
            ad["background_seed"] = lock["background_seed"]
        if isinstance(lock.get("visual_lock"), dict):
            ad["visual_lock"] = lock["visual_lock"]
    return cloned


def _background_reuse_keys(fmt: str, persona_no: int | None, visual_archetype: str, share_across_personas: bool) -> list[str]:
    fmt = fmt.strip().upper()
    persona = f"P{persona_no:02d}" if isinstance(persona_no, int) else ""
    arch = visual_archetype.strip()
    if share_across_personas:
        return [key for key in [f"{fmt}::{arch}" if arch else "", fmt] if key]
    return [key for key in [f"{fmt}::{persona}::{arch}" if persona and arch else "", f"{fmt}::{persona}" if persona else ""] if key]


def collect_background_reuse_locks(source_run_id: str) -> dict[str, dict[str, Any]]:
    source_run_id = str(source_run_id or "").strip()
    if not source_run_id:
        return {}
    _source_dir, manifest, _has_storage_manifest = load_manifest_for_run(source_run_id)
    locks: dict[str, dict[str, Any]] = {}
    for rel_path in manifest.get("prompt_files") or []:
        rel = str(rel_path).replace("\\", "/")
        if "/916/" in rel or "/96/" in rel:
            continue
        parsed = parse_prompt_filename(rel)
        if not parsed:
            continue
        fmt, _lang, persona_no = parsed
        prompt_path = ROOT / rel
        if not prompt_path.exists():
            continue

        slot = ""
        seed: int | None = None
        visual_archetype = ""
        sidecar = prompt_path.with_suffix(".json")
        if sidecar.exists():
            try:
                meta = json.loads(sidecar.read_text(encoding="utf-8"))
            except Exception:
                meta = {}
            bg = meta.get("background") if isinstance(meta.get("background"), dict) else {}
            slot = str(bg.get("slot") or "").strip()
            raw_seed = bg.get("seed")
            if isinstance(raw_seed, int):
                seed = raw_seed
            visual = meta.get("visual_archetype") if isinstance(meta.get("visual_archetype"), dict) else {}
            visual_archetype = str(visual.get("id") or "").strip()

        if not slot or not isinstance(seed, int):
            text = prompt_path.read_text(encoding="utf-8", errors="ignore")
            lock = parse_background_lock_from_prompt(text)
            if lock:
                slot, seed = lock
            if not visual_archetype:
                visual_archetype = _parse_prompt_field(text, "Selected visual archetype").split(" - ", 1)[0].strip()

        if not slot or not isinstance(seed, int):
            continue

        lock_payload = {
            "background_slot": slot,
            "background_seed": seed,
            "background_reused_from_run_id": source_run_id,
        }
        for key in _background_reuse_keys(fmt, persona_no, visual_archetype, False):
            locks.setdefault(key, lock_payload)
        for key in _background_reuse_keys(fmt, persona_no, visual_archetype, True):
            locks.setdefault(key, lock_payload)
    return locks


def apply_background_reuse_locks(
    copy_json: dict[str, Any],
    locks: dict[str, dict[str, Any]],
    *,
    share_across_personas: bool,
) -> tuple[dict[str, Any], int]:
    cloned = json.loads(json.dumps(copy_json, ensure_ascii=False))
    ads = cloned.get("ads")
    if not isinstance(ads, list) or not locks:
        return cloned, 0
    applied = 0
    for ad in ads:
        if not isinstance(ad, dict):
            continue
        fmt = str(ad.get("format") or "").strip().upper()
        persona_no = None
        persona = ad.get("persona")
        if isinstance(persona, dict) and isinstance(persona.get("number"), int):
            persona_no = int(persona["number"])
        visual_archetype = str(ad.get("visual_archetype") or "").strip()
        lock = None
        reuse_key = ""
        for key in _background_reuse_keys(fmt, persona_no, visual_archetype, share_across_personas):
            if key in locks:
                lock = locks[key]
                reuse_key = key
                break
        if not lock:
            continue
        ad["background_slot"] = lock["background_slot"]
        ad["background_seed"] = lock["background_seed"]
        ad["background_reused_from_run_id"] = lock.get("background_reused_from_run_id", "")
        ad["background_reuse_key"] = reuse_key
        applied += 1
    return cloned, applied


def collect_visual_pattern_reuse_locks(source_run_id: str) -> dict[str, dict[str, Any]]:
    source_run_id = str(source_run_id or "").strip()
    if not source_run_id:
        return {}
    _source_dir, manifest, _has_storage_manifest = load_manifest_for_run(source_run_id)
    locks: dict[str, dict[str, Any]] = {}
    for rel_path in manifest.get("prompt_files") or []:
        rel = str(rel_path).replace("\\", "/")
        if "/916/" in rel or "/96/" in rel:
            continue
        parsed = parse_prompt_filename(rel)
        if not parsed:
            continue
        fmt, _lang, persona_no = parsed
        prompt_path = ROOT / rel
        if not prompt_path.exists():
            continue

        visual_archetype = ""
        sidecar = prompt_path.with_suffix(".json")
        if sidecar.exists():
            try:
                meta = json.loads(sidecar.read_text(encoding="utf-8"))
            except Exception:
                meta = {}
            visual = meta.get("visual_archetype") if isinstance(meta.get("visual_archetype"), dict) else {}
            visual_archetype = str(visual.get("id") or "").strip()

        if not visual_archetype:
            text = prompt_path.read_text(encoding="utf-8", errors="ignore")
            visual_archetype = _parse_prompt_field(text, "Selected visual archetype").split(" - ", 1)[0].strip()

        if not visual_archetype:
            continue

        lock_payload = {
            "visual_archetype": visual_archetype,
            "visual_pattern_reused_from_run_id": source_run_id,
        }
        for key in _background_reuse_keys(fmt, persona_no, visual_archetype, False):
            locks.setdefault(key, lock_payload)
        for key in _background_reuse_keys(fmt, persona_no, visual_archetype, True):
            locks.setdefault(key, lock_payload)
    return locks


def apply_visual_pattern_reuse_to_plan(
    plan: list[dict[str, Any]],
    locks: dict[str, dict[str, Any]],
    *,
    share_across_personas: bool,
) -> tuple[list[dict[str, Any]], int]:
    if not locks:
        return plan, 0
    out: list[dict[str, Any]] = []
    applied = 0
    for item in plan:
        entry = dict(item)
        fmt = str(entry.get("format") or "").strip().upper()
        persona_no = int(entry.get("persona")) if entry.get("persona") is not None else None
        lock = None
        reuse_key = ""
        keys = []
        if share_across_personas:
            keys.append(fmt)
        else:
            keys.append(f"{fmt}::P{persona_no:02d}" if isinstance(persona_no, int) else fmt)
        for key in keys:
            if key in locks:
                lock = locks[key]
                reuse_key = key
                break
        if lock:
            entry["visual_archetype"] = lock["visual_archetype"]
            entry["visual_pattern_reused_from_run_id"] = lock.get("visual_pattern_reused_from_run_id", "")
            entry["visual_pattern_reuse_key"] = reuse_key
            applied += 1
        out.append(entry)
    return out, applied


def resolve_format_plan(config: dict[str, Any]) -> list[dict[str, Any]]:
    personas = config.get("selected_personas") or []
    if not personas:
        raise RuntimeError("selected_personas is required")

    all_formats = [fmt for fmt in (config.get("global_formats") or []) if fmt in FORMATS]
    format_map = config.get("formats_by_persona") or {}
    archetype_map = config.get("visual_archetypes_by_format") or {}
    share_bg_across_personas = bool(config.get("share_background_across_personas"))
    try:
        multiplier = max(1, min(20, int(config.get("multiplier") or 1)))
    except (TypeError, ValueError):
        multiplier = 1

    out: list[dict[str, Any]] = []
    for raw_persona in personas:
        persona_num = int(raw_persona)
        per_formats = [fmt for fmt in (format_map.get(str(persona_num)) or format_map.get(persona_num) or []) if fmt in FORMATS]
        formats = per_formats if per_formats else all_formats
        if not formats:
            formats = ["HERO"]
        for fmt in formats:
            forced_archetype = str(archetype_map.get(fmt) or "").strip()
            background_group_key = fmt if share_bg_across_personas else f"{fmt}::P{persona_num:02d}"
            for creative_index in range(1, multiplier + 1):
                item = {
                    "persona": persona_num,
                    "format": fmt,
                    "creative_index": creative_index,
                    "creative_total": multiplier,
                    "background_group_key": background_group_key,
                    "share_background_across_personas": share_bg_across_personas,
                }
                if forced_archetype:
                    item["visual_archetype"] = forced_archetype
                out.append(item)
    return out


def expand_plan_with_hypothesis(plan: list[dict[str, Any]], hypothesis_cfg: dict[str, Any]) -> list[dict[str, Any]]:
    """Expand ad plan to include hypothesis style.

    When a hypothesis is active, generates ads using that specific style/variant.
    """
    hyp_type = str(hypothesis_cfg.get("type") or "none").strip().lower()
    if hyp_type == "none" or hyp_type not in HYPOTHESIS_VARIABLES:
        return plan

    variable_def = HYPOTHESIS_VARIABLES[hyp_type]
    selected_variant = str(hypothesis_cfg.get("variant") or "").strip()
    available_options = [opt["id"] for opt in variable_def.get("options", [])]

    if not available_options:
        return plan

    # Use the selected variant if valid, otherwise use first available
    variant_to_use = selected_variant if selected_variant in available_options else available_options[0]

    out: list[dict[str, Any]] = []
    for item in plan:
        entry = dict(item)
        entry["hypothesis"] = {
            "type": hyp_type,
            "variable_label": variable_def["label"],
            "variant": variant_to_use,
            "hypothesis_id": f"{hyp_type}-{variant_to_use}",
        }
        base_group_key = str(entry.get("background_group_key") or f"{entry.get('format')}::P{int(entry.get('persona')):02d}")
        entry["background_group_key"] = f"{base_group_key}::{hyp_type}::{variant_to_use}"
        out.append(entry)
    return out


app = FastAPI(title="Ad Dashboard API", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://127.0.0.1:4090", "http://localhost:4090"],
    allow_credentials=False,
    allow_methods=["GET", "POST"],
    allow_headers=["Content-Type", "Authorization"],
)


@app.on_event("startup")
def startup() -> None:
    load_env_file(ENV_PATH)
    ensure_dirs()


def api_defaults() -> dict[str, Any]:
    personas = parse_persona_library(DEFAULT_PLAYBOOK)
    opencode = build_opencode_catalog()
    return {
        "personas": personas,
        "formats": FORMATS,
        "format_patterns": load_format_visual_archetypes(),
        "image_sources": read_active_images(default_image_sources_file()),
        "input_images": list_input_images(),
        "product_doc": default_product_doc_info(),
        "default_files": {
            "product_info": str(DEFAULT_PRODUCT_MASTER.relative_to(ROOT)),
            "playbook": str(DEFAULT_PLAYBOOK.relative_to(ROOT)),
        },
        "opencode": opencode,
        "hypothesis": {
            "variables": HYPOTHESIS_VARIABLES,
            "default": {"type": "none", "variant": ""},
        },
    }


def api_progress(batch_key: str) -> dict[str, Any]:
    batch_key_clean = str(batch_key).strip()
    for root in generated_image_roots():
        log_path = root / batch_key_clean / "_headless_progress.json"
        if not log_path.exists():
            continue
        lines = log_path.read_text(encoding="utf-8", errors="ignore").splitlines()
        entries = []
        for line in lines:
            line = line.strip()
            if not line:
                continue
            try:
                entries.append(json.loads(line))
            except json.JSONDecodeError:
                pass
        if not entries:
            continue
        latest = entries[-1]
        return {
            "batch_key": batch_key_clean,
            "step": latest.get("step", ""),
            "message": latest.get("message", ""),
            "time": latest.get("time", 0),
            "entries": entries,
        }
    raise HTTPException(status_code=404, detail=f"No progress found for batch: {batch_key_clean}")


def api_opencode_catalog() -> dict[str, Any]:
    return build_opencode_catalog()


def _extract_backfill_batch(run_id: str) -> str | None:
    match = re.match(r"^batch_(v\d+)$", str(run_id or "").strip(), flags=re.IGNORECASE)
    if not match:
        return None
    return match.group(1)


def _build_backfill_manifest(run_id: str, batch: str) -> dict[str, Any]:
    prompt_files = scan_prompt_files_for_batch(batch)
    if not prompt_files:
        raise HTTPException(status_code=404, detail=f"No prompt files found in output/{batch}")
    image_files = scan_image_files_for_batch(batch)
    batch_dir = ROOT / "output" / batch
    updated_at = now_iso()
    if batch_dir.exists():
        updated_at = datetime.fromtimestamp(batch_dir.stat().st_mtime, tz=timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")
    return {
        "run_id": run_id,
        "batch": batch,
        "prompt_files": prompt_files,
        "image_files": image_files,
        "image_generated": bool(image_files),
        "updated_at": updated_at,
        "source": "output_backfill",
    }


def load_manifest_for_run(run_id: str) -> tuple[Path | None, dict[str, Any], bool]:
    run_dir = RUNS_ROOT / run_id
    manifest_path = run_dir / "manifest.json"
    if manifest_path.exists():
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        refreshed = refresh_manifest_file_state(run_dir, manifest)
        return run_dir, refreshed, True

    backfill_batch = _extract_backfill_batch(run_id)
    if backfill_batch:
        return None, _build_backfill_manifest(run_id, backfill_batch), False

    raise HTTPException(status_code=404, detail="Run not found")


def collect_backfill_result(run_id: str, batch: str) -> dict[str, Any]:
    manifest = _build_backfill_manifest(run_id, batch)
    manifest["generated_variant"] = "4:5"
    return manifest


def api_runs() -> dict[str, Any]:
    ensure_dirs()
    runs: list[dict[str, Any]] = []
    seen_run_ids: set[str] = set()
    seen_batches: set[str] = set()
    for run_dir in sorted(RUNS_ROOT.glob("run_*"), reverse=True):
        manifest = run_dir / "manifest.json"
        if manifest.exists():
            payload = json.loads(manifest.read_text(encoding="utf-8"))
            refreshed = refresh_manifest_file_state(run_dir, payload)
            run_id = str(refreshed.get("run_id") or run_dir.name)
            batch = str(refreshed.get("batch") or "").strip()
            if run_id in seen_run_ids:
                continue
            seen_run_ids.add(run_id)
            if batch:
                seen_batches.add(batch)
            runs.append(refreshed)

    # Backfill batches that exist on disk but have no run manifest
    # (e.g., older/generated output imported from another machine).
    output_root = ROOT / "output"
    if output_root.exists():
        for batch_dir in sorted(output_root.glob("v*"), reverse=True):
            if not batch_dir.is_dir():
                continue
            batch_name = batch_dir.name
            if batch_name in seen_batches:
                continue
            prompt_files = scan_prompt_files_for_batch(batch_name)
            if not prompt_files:
                continue
            image_files = scan_image_files_for_batch(batch_name)
            updated_at = datetime.fromtimestamp(batch_dir.stat().st_mtime, tz=timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")
            runs.append(
                {
                    "run_id": f"batch_{batch_name}",
                    "batch": batch_name,
                    "prompt_files": prompt_files,
                    "image_files": image_files,
                    "image_generated": bool(image_files),
                    "updated_at": updated_at,
                    "source": "output_backfill",
                }
            )

    def batch_sort_key(run: dict[str, Any]) -> tuple[int, float]:
        batch = str(run.get("batch") or "").strip().lower()
        match = re.match(r"^v(\d+)$", batch)
        batch_num = int(match.group(1)) if match else -1
        updated = str(run.get("updated_at") or "")
        ts = 0.0
        if updated:
            try:
                ts = datetime.fromisoformat(updated.replace("Z", "+00:00")).timestamp()
            except Exception:
                ts = 0.0
        return (batch_num, ts)

    runs.sort(key=batch_sort_key, reverse=True)
    return {"runs": runs}


def api_run(run_id: str) -> dict[str, Any]:
    _run_dir, manifest, _has_storage_manifest = load_manifest_for_run(run_id)
    return manifest


def api_run_prompt_copies(run_id: str) -> dict[str, Any]:
    _run_dir, manifest, _has_storage_manifest = load_manifest_for_run(run_id)
    prompt_files_all = manifest.get("prompt_files") or []
    prompt_files = [path for path in prompt_files_all if "/45/" in str(path)] or prompt_files_all
    records: list[dict[str, Any]] = []
    for rel_path in prompt_files:
        prompt_path = ROOT / rel_path
        if not prompt_path.exists() or not prompt_path.is_file():
            continue
        text = prompt_path.read_text(encoding="utf-8", errors="ignore")
        parsed_name = parse_prompt_filename(rel_path)
        persona_number = parsed_name[2] if parsed_name else None
        if persona_number is None:
            persona_number = parse_persona_number_from_prompt(text)
        records.append(
            {
                "prompt_file": rel_path,
                "format": parsed_name[0] if parsed_name else "",
                "language": parsed_name[1] if parsed_name else "",
                "persona_number": persona_number,
                "review_url": "/output/" + rel_path.replace("output/", ""),
                "copy_lines": extract_on_image_copy_lines(text),
            }
        )

    return {"run_id": run_id, "batch": manifest.get("batch"), "prompts": records}


def api_run_update_prompt_copies(run_id: str, payload: dict[str, Any] = Body(...)) -> dict[str, Any]:
    run_dir = RUNS_ROOT / run_id
    manifest_path = run_dir / "manifest.json"
    copy_path = run_dir / "context" / "copy_batch.json"

    if not manifest_path.exists():
        raise HTTPException(status_code=404, detail="Run not found")
    if not copy_path.exists():
        raise HTTPException(status_code=404, detail="copy_batch.json not found for run")

    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    batch = (manifest.get("batch") or "").strip()
    if not batch:
        raise HTTPException(status_code=400, detail="Run has no batch folder")

    edits = payload.get("edits")
    if not isinstance(edits, list) or not edits:
        raise HTTPException(status_code=400, detail="edits must be a non-empty array")

    copy_json = json.loads(copy_path.read_text(encoding="utf-8"))
    ads = copy_json.get("ads")
    if not isinstance(ads, list) or not ads:
        raise HTTPException(status_code=400, detail="Invalid copy batch payload")

    updated_count = 0
    for entry in edits:
        if not isinstance(entry, dict):
            continue
        prompt_file = str(entry.get("prompt_file") or "").strip()
        if not prompt_file:
            continue
        parsed_name = parse_prompt_filename(prompt_file)
        if not parsed_name:
            continue
        fmt, lang, parsed_persona = parsed_name
        persona_number = entry.get("persona_number")
        if not isinstance(persona_number, int):
            persona_number = parsed_persona
        line_items = entry.get("copy_lines")
        if not isinstance(line_items, list) or not line_items:
            continue

        target_ad = None
        for ad in ads:
            if not isinstance(ad, dict):
                continue
            if str(ad.get("format") or "").strip().upper() != fmt:
                continue
            if isinstance(persona_number, int):
                persona = ad.get("persona")
                ad_persona_no = None
                if isinstance(persona, dict) and isinstance(persona.get("number"), int):
                    ad_persona_no = int(persona.get("number"))
                if ad_persona_no != persona_number:
                    continue
            target_ad = ad
            break
        if not isinstance(target_ad, dict):
            continue

        ad_copy = target_ad.setdefault("copy", {})
        if not isinstance(ad_copy, dict):
            continue
        lang_copy = ad_copy.setdefault(lang, {})
        if not isinstance(lang_copy, dict):
            continue

        for line_item in line_items:
            if not isinstance(line_item, dict):
                continue
            label = str(line_item.get("label") or "").strip()
            value = str(line_item.get("value") or "").strip()
            if not label:
                continue
            key = label.lower()

            if key == "headline":
                lang_copy["headline"] = value
            elif key == "support line":
                lang_copy["support_line"] = value
            elif key == "context line":
                lang_copy["context_line"] = value
            elif key == "cta":
                lang_copy["cta"] = value
            elif key == "attribution":
                lang_copy["attribution"] = value
            elif key == "trust line":
                lang_copy["trust_line"] = value
            elif key.startswith("bullet "):
                match = re.match(r"^bullet\s+(\d+)$", key)
                if not match:
                    continue
                index = int(match.group(1)) - 1
                if index < 0:
                    continue
                bullets = lang_copy.get("bullets")
                if not isinstance(bullets, list):
                    bullets = []
                while len(bullets) <= index:
                    bullets.append("")
                bullets[index] = value
                lang_copy["bullets"] = bullets
            elif key.startswith("left situation ") or key.startswith("right shift "):
                match = re.match(r"^(left situation|right shift)\s+(\d+)$", key)
                if not match:
                    continue
                side = match.group(1)
                ordinal = int(match.group(2))
                if ordinal <= 0:
                    continue
                if side == "left situation":
                    index = ordinal - 1
                else:
                    index = ordinal + 1
                bullets = lang_copy.get("bullets")
                if not isinstance(bullets, list):
                    bullets = []
                while len(bullets) <= index:
                    bullets.append("")
                bullets[index] = value
                lang_copy["bullets"] = bullets

        updated_count += 1

    if updated_count == 0:
        raise HTTPException(status_code=400, detail="No valid prompt edits were provided")

    copy_path.write_text(json.dumps(copy_json, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    rerender_prompts_for_run(run_dir, batch, copy_path, load_run_language_mode(run_dir))

    _append_audit_log(
        run_dir,
        "prompt_updates",
        {
            "run_id": run_id,
            "batch": batch,
            "updated_count": updated_count,
        },
    )

    has_916 = any("/96/" in str(path) for path in (manifest.get("prompt_files") or []))
    if has_916:
        manifest = generate_916_for_run(run_dir, manifest)

    refreshed = collect_run_result(run_dir, batch, bool(manifest.get("image_generated", False)))
    refreshed["copy_edits_applied"] = updated_count
    merged = merge_manifest(run_dir, manifest, refreshed)
    return merged


# ──────────────────────────────────────────────────────────────────────────────
# Task 6/7/8: Export/Import on-image copy (EXACT ON-IMAGE COPY block)
# ──────────────────────────────────────────────────────────────────────────────

EXACT_COPY_SHEET_COLUMNS = [
    "prompt_id",
    "vn",
    "hypothesis_id",
    "hypothesis_name",
    "persona",
    "angle",
    "headline_copy",
    "exact_on_image_copy_block",
    "created_at",
]

def _extract_vn_from_prompt_rel_path(prompt_rel_path: str) -> str:
    # Expected pattern: output/v{N}/...
    # Keep backward compatible: if not found, return empty string.
    m = re.search(r"/output/(v\d+)(/|$)", prompt_rel_path.replace("\\", "/"))
    return m.group(1) if m else ""


def _extract_vn_from_image_path(image_path: str) -> str:
    # Expected pattern: generated_images/v{N}/...
    m = re.search(r"/generated_images/(v\d+)(/|$)", image_path.replace("\\", "/"))
    return m.group(1) if m else ""


def _extract_aspect_from_image_path(image_path: str) -> str:
    # Extract aspect folder e.g. 4_5 or 9_16 from generated_images/v77/4_5/...
    m = re.search(r"/(\d+_\d+)/", image_path.replace("\\", "/"))
    return m.group(1) if m else ""


def _extract_created_at_iso_from_file(file_path: Path) -> str:
    try:
        ts = file_path.stat().st_mtime
        return datetime.fromtimestamp(ts, tz=timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")
    except Exception:
        return ""


def _parse_exact_block_headline_value(block_text: str) -> str | None:
    """
    Preserve EXACT headline value text as written in the exact block.

    We intentionally do NOT trim/normalize:
    - keep any spaces immediately after the colon
    - keep punctuation/case/capitalization
    """
    for raw_line in (block_text or "").splitlines():
        # Allow optional whitespace before "-" and around "-", but preserve everything after "Headline:"
        m = re.match(r"^\s*-\s*Headline:(.*)$", raw_line)
        if not m:
            continue
        return m.group(1)
    return None


def _replace_exact_copy_block(prompt_text: str, new_block_text: str) -> str | None:
    m = EXACT_COPY_BLOCK_RE.search(prompt_text or "")
    if not m:
        return None
    start_idx = m.start("block")
    end_idx = m.end("block")
    return (prompt_text[:start_idx] + new_block_text + prompt_text[end_idx:])


def _load_run_prompt_files(run_id: str, aspect_ratios: list[str] | None = None) -> list[str]:
    _run_dir, manifest, _has_storage_manifest = load_manifest_for_run(run_id)
    prompt_files_all = manifest.get("prompt_files") or []
    if not aspect_ratios:
        return prompt_files_all
    result: list[str] = []
    for p in prompt_files_all:
        for ar in aspect_ratios:
            if f"/{ar}/" in str(p):
                result.append(p)
                break
    return result or prompt_files_all


def _extract_prompt_row_metadata(run_id: str, copy_batch: dict[str, Any], prompt_rel_path: str, batch_vn: str = "") -> dict[str, Any]:
    prompt_path = ROOT / prompt_rel_path
    text = prompt_path.read_text(encoding="utf-8", errors="ignore")

    block = extract_exact_on_image_copy_block(text)
    headline_copy = None
    exact_block = ""
    if block is not None:
        headline_copy = _parse_exact_block_headline_value(block)
        exact_block = block.strip()
    if headline_copy is None:
        headline_copy = ""

    vn = _extract_vn_from_prompt_rel_path(prompt_rel_path)
    if not vn and batch_vn:
        vn = batch_vn
    created_at = _extract_created_at_iso_from_file(prompt_path)

    parsed = parse_prompt_filename(prompt_rel_path)
    fmt = parsed[0] if parsed else ""
    persona_number = parsed[2] if parsed else None
    creative_index = parse_prompt_creative_index(prompt_rel_path)
    if persona_number is None:
        persona_number = parse_persona_number_from_prompt(text)

    persona = ""
    angle = ""
    hypothesis_id = ""
    hypothesis_name = ""

    if isinstance(persona_number, int):
        for ad in copy_batch.get("ads") or []:
            if not isinstance(ad, dict):
                continue
            if str(ad.get("format") or "").strip().upper() != fmt:
                continue
            persona = ""
            persona_obj = ad.get("persona")
            if isinstance(persona_obj, dict):
                if isinstance(persona_obj.get("number"), int) and int(persona_obj.get("number")) == persona_number:
                    if int(ad.get("creative_index") or 1) != creative_index:
                        continue
                    persona = str(persona_obj.get("persona_name") or persona_obj.get("name") or f"Persona {persona_number}")
            if persona:
                angle = str(ad.get("headline_angle") or ad.get("concept_angle") or "")
                hyp = ad.get("hypothesis") if isinstance(ad.get("hypothesis"), dict) else {}
                if hyp:
                    hypothesis_id = str(hyp.get("hypothesis_id") or "")
                    hypothesis_name = str(hyp.get("variant") or hyp.get("variable_label") or "")
                break

    return {
        "prompt_id": prompt_rel_path,
        "vn": vn,
        "hypothesis_id": hypothesis_id,
        "hypothesis_name": hypothesis_name,
        "persona": persona,
        "angle": angle,
        "headline_copy": headline_copy,
        "exact_on_image_copy_block": exact_block,
        "created_at": created_at,
    }


def _append_audit_log(run_dir: Path, event_type: str, payload: dict[str, Any]) -> None:
    audit_dir = run_dir / "audit"
    audit_dir.mkdir(parents=True, exist_ok=True)
    path = audit_dir / "audit_log.jsonl"
    entry = {"ts": now_iso(), "event_type": event_type, "payload": payload}
    with path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(entry, ensure_ascii=False) + "\n")


def api_export_on_image_copy(run_id: str) -> StreamingResponse:
    from openpyxl import Workbook
    import io

    run_dir, manifest, has_storage_manifest = load_manifest_for_run(run_id)

    copy_batch: dict[str, Any] = {"ads": []}
    if has_storage_manifest and run_dir is not None:
        copy_path = run_dir / "context" / "copy_batch.json"
        if copy_path.exists():
            copy_batch = json.loads(copy_path.read_text(encoding="utf-8"))

    prompt_files = _load_run_prompt_files(run_id)

    unique_vns = set()
    for rel in prompt_files:
        prompt_rel_path = str(rel).replace("\\", "/")
        vn = _extract_vn_from_prompt_rel_path(prompt_rel_path)
        if vn:
            unique_vns.add(vn)

    batch = manifest.get("batch", "")

    if not unique_vns:
        if batch and batch.startswith("v"):
            unique_vns.add(batch)

    if unique_vns:
        vn_suffix = "-".join(sorted(unique_vns))
    else:
        vn_suffix = None

    wb = Workbook()
    ws = wb.active
    ws.title = "on-image-copy"

    ws.append(EXACT_COPY_SHEET_COLUMNS)
    for rel in prompt_files:
        prompt_rel_path = str(rel).replace("\\", "/")
        if not (ROOT / prompt_rel_path).exists():
            continue
        row = _extract_prompt_row_metadata(run_id, copy_batch, prompt_rel_path, batch)
        ws.append([row.get(col, "") for col in EXACT_COPY_SHEET_COLUMNS])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    if run_dir is not None:
        _append_audit_log(run_dir, "export_on_image_copy", {"run_id": run_id, "prompt_rows": len(prompt_files)})

    if vn_suffix:
        filename = f"on-image-copy-{vn_suffix}.xlsx"
    else:
        filename = f"on-image-copy-{run_id}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


async def api_import_on_image_copy(
    run_id: str,
    file: UploadFile = File(...),
    confirm: bool = Form(False),
) -> dict[str, Any]:
    from openpyxl import load_workbook

    run_dir, manifest, has_storage_manifest = load_manifest_for_run(run_id)

    # Parse xlsx (no prompt regeneration; only exact-block replacement)
    if not file.filename:
        raise HTTPException(status_code=400, detail="Missing upload filename")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Empty upload")

    import_root = (run_dir / "imports") if run_dir is not None else (RUNTIME_ROOT / "imports")
    tmp_path = import_root / f"upload-{int(time.time())}-{file.filename}"
    tmp_path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path.write_bytes(content)

    wb = load_workbook(tmp_path)
    ws = wb.active

    # Build column index
    header = [str(cell.value or "").strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col_idx = {name: i for i, name in enumerate(header) if name}
    missing_cols = [c for c in EXACT_COPY_SHEET_COLUMNS if c not in col_idx]
    if missing_cols:
        raise HTTPException(status_code=400, detail=f"Missing required columns: {missing_cols}")

    seen_prompt_ids: set[str] = set()
    rows: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []

    for excel_row in ws.iter_rows(min_row=2):
        values = [cell.value for cell in excel_row]
        prompt_id = str(values[col_idx["prompt_id"]] or "").replace("\\", "/").strip()
        if not prompt_id:
            continue
        if prompt_id in seen_prompt_ids:
            errors.append({"prompt_id": prompt_id, "error": "duplicate_prompt_id"})
            continue
        seen_prompt_ids.add(prompt_id)

        headline_copy = str(values[col_idx["headline_copy"]] or "").strip()
        full_block = str(values[col_idx.get("exact_on_image_copy_block", -1)] or "").strip() if "exact_on_image_copy_block" in col_idx else ""

        if not headline_copy.strip() and not full_block:
            errors.append({"prompt_id": prompt_id, "error": "empty_headline_copy_and_block"})
            continue

        rows.append(
            {
                "prompt_id": prompt_id,
                "headline_copy": headline_copy,
                "full_block": full_block,
            }
        )

    # Validate prompt_id exists
    for r in rows:
        p = ROOT / r["prompt_id"]
        if not p.exists() or not p.is_file():
            errors.append({"prompt_id": r["prompt_id"], "error": "prompt_id_not_found"})
    if errors:
        if run_dir is not None:
            _append_audit_log(run_dir, "import_on_image_copy_validation_failed", {"run_id": run_id, "errors": errors, "confirm": confirm})
        raise HTTPException(status_code=400, detail={"validation_errors": errors})

    # Preview diffs
    preview_items: list[dict[str, Any]] = []
    applied_count = 0
    skipped_count = 0

    for r in rows:
        prompt_rel_path = r["prompt_id"]
        prompt_path = ROOT / prompt_rel_path
        old_text = prompt_path.read_text(encoding="utf-8", errors="ignore")

        old_block = extract_exact_on_image_copy_block(old_text, warn_log_path=None)
        if old_block is None:
            skipped_count += 1
            preview_items.append({"prompt_id": prompt_rel_path, "status": "skipped_missing_exact_block"})
            continue

        full_block = r.get("full_block", "")
        new_block = None

        if full_block:
            new_block = full_block
            old_copy = old_block.strip()
            new_copy = new_block.strip()
        else:
            headline_copy = r.get("headline_copy", "")
            new_lines: list[str] = []
            headline_replaced = False
            for line in old_block.splitlines():
                m = re.match(r"^(\s*-\s*Headline:)(.*)$", line)
                if m:
                    new_lines.append(m.group(1) + headline_copy)
                    headline_replaced = True
                else:
                    new_lines.append(line)

            if not headline_replaced:
                skipped_count += 1
                preview_items.append({"prompt_id": prompt_rel_path, "status": "skipped_headline_line_not_found"})
                continue

            new_block = "\n".join(new_lines)
            old_copy = _parse_exact_block_headline_value(old_block) or ""
            new_copy = _parse_exact_block_headline_value(new_block) or ""

        preview_items.append(
            {
                "prompt_id": prompt_rel_path,
                "status": "ready_to_apply" if confirm else "preview",
                "old_copy": old_copy[:100] + "..." if len(old_copy) > 100 else old_copy,
                "new_copy": new_copy[:100] + "..." if len(new_copy) > 100 else new_copy,
            }
        )

        if confirm:
            updated_text = _replace_exact_copy_block(old_text, new_block)
            if updated_text is None:
                skipped_count += 1
                preview_items[-1]["status"] = "skipped_replace_failed"
                continue
            prompt_path.write_text(updated_text, encoding="utf-8")
            applied_count += 1

    if run_dir is not None:
        _append_audit_log(
            run_dir,
            "import_on_image_copy",
            {"run_id": run_id, "confirm": confirm, "rows": len(rows), "applied": applied_count, "skipped": skipped_count},
        )

    if not confirm:
        return {
            "run_id": run_id,
            "preview": True,
            "changed_rows_count": applied_count,
            "skipped_rows": skipped_count,
            "failed_rows": len(errors),
            "items": preview_items,
        }

    # Re-assemble prompts side-effects: since we directly edited prompt text,
    # we do not mutate copy_batch.json metadata (per requirements).
    # However, manifest/prompt_files state should be refreshed.
    merged: dict[str, Any] | None = None
    if run_dir is not None and has_storage_manifest:
        manifest = json.loads((run_dir / "manifest.json").read_text(encoding="utf-8"))
        batch = (manifest.get("batch") or "").strip()
        refreshed = collect_run_result(run_dir, batch, bool(manifest.get("image_generated", False)))
        refreshed["on_image_copy_import_applied"] = applied_count
        merged = merge_manifest(run_dir, manifest, refreshed)
    else:
        batch = str(manifest.get("batch") or "")
        merged = collect_backfill_result(run_id, batch) if batch else manifest
        merged["on_image_copy_import_applied"] = applied_count

    return {
        "run_id": run_id,
        "preview": False,
        "changed_rows_count": applied_count,
        "skipped_rows": skipped_count,
        "failed_rows": len(errors),
        "items": preview_items,
        "manifest": merged,
    }


def api_run_generate_916(run_id: str) -> dict[str, Any]:
    run_dir, manifest, has_storage_manifest = load_manifest_for_run(run_id)
    if not has_storage_manifest or run_dir is None:
        raise HTTPException(status_code=400, detail="This endpoint requires run context in dashboard_storage. Use generate-images-916-from-45 for output-only batches.")
    return generate_916_for_run(run_dir, manifest)


def api_run_generate_916_selected(run_id: str, payload: dict[str, Any] = Body(...)) -> dict[str, Any]:
    run_dir, manifest, has_storage_manifest = load_manifest_for_run(run_id)
    if not has_storage_manifest or run_dir is None:
        raise HTTPException(status_code=400, detail="This endpoint requires run context in dashboard_storage for copy_batch filtering.")
    copy_path = run_dir / "context" / "copy_batch.json"
    if not copy_path.exists():
        raise HTTPException(status_code=404, detail="copy_batch.json not found for run")
    batch = str(manifest.get("batch") or "").strip()
    if not batch:
        raise HTTPException(status_code=400, detail="Run has no batch folder")

    prompt_files = payload.get("prompt_files")
    if not isinstance(prompt_files, list) or not prompt_files:
        raise HTTPException(status_code=400, detail="prompt_files must be a non-empty array")

    selected_45 = validate_selected_45_prompts(batch, prompt_files)
    if not selected_45:
        raise HTTPException(status_code=400, detail="No valid 4:5 prompt files selected")

    selected_keys = extract_selected_ad_keys_from_45_prompts(selected_45)
    if not selected_keys:
        raise HTTPException(status_code=400, detail="Could not resolve selected persona/format keys")

    copy_json = json.loads(copy_path.read_text(encoding="utf-8"))
    selected_copy = filter_copy_json_for_selected_ads(copy_json, selected_keys)
    ads = selected_copy.get("ads")
    if not isinstance(ads, list) or not ads:
        raise HTTPException(status_code=400, detail="No ads matched selected prompts")

    copy_916 = force_aspect_ratio(selected_copy, "9:16")
    visual_locks = collect_45_visual_locks(batch)
    if visual_locks:
        copy_916 = apply_visual_locks(copy_916, visual_locks)
    copy_916_path = run_dir / "context" / "copy_batch_916_selected.json"
    copy_916_path.write_text(json.dumps(copy_916, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    result = run_cmd(
        [
            "python3",
            "scripts/generate_ads.py",
            "--copy-file",
            str(copy_916_path),
            "--batch",
            batch,
            "--language-mode",
            load_run_language_mode(run_dir),
            "--no-registry-write",
            "--skip-uniqueness-check",
        ],
        cwd=ROOT,
    )
    if result.returncode != 0:
        error_text = result.stderr or result.stdout
        (run_dir / "logs" / "assembler_916_selected_error.txt").write_text(error_text, encoding="utf-8")
        short_error = "\n".join([line for line in error_text.splitlines() if line.strip()][-12:])
        raise HTTPException(status_code=500, detail=f"Selective 9:16 generation failed: {short_error}")

    refreshed = collect_run_result(run_dir, batch, bool(manifest.get("image_generated", False)))
    refreshed["generated_variant"] = "9:16"
    refreshed["generated_916_for_prompts"] = selected_45
    return merge_manifest(run_dir, manifest, refreshed)


def validate_selected_45_prompts(batch: str, prompt_files: list[Any]) -> list[str]:
    valid_prompt_files: list[str] = []
    for prompt_file in prompt_files:
        rel = str(prompt_file or "").strip().replace("\\", "/")
        if not rel or not rel.startswith("output/"):
            continue
        if "/45/" not in rel:
            continue
        candidate = ROOT / rel
        if not candidate.exists() or not candidate.is_file():
            continue
        if f"output/{batch}/" not in rel:
            continue
        valid_prompt_files.append(rel)
    return valid_prompt_files


def map_45_to_96_prompts(selected_45: list[str]) -> list[str]:
    out: list[str] = []
    for rel in selected_45:
        rel_96 = rel.replace("/45/", "/96/")
        file_96 = ROOT / rel_96
        if file_96.exists() and file_96.is_file():
            out.append(rel_96)
    return out


def extract_selected_ad_keys_from_45_prompts(selected_45: list[str]) -> set[tuple[str, int | None]]:
    keys: set[tuple[str, int | None]] = set()
    for rel in selected_45:
        parsed = parse_prompt_filename(rel)
        if not parsed:
            continue
        fmt, _lang, persona_number = parsed
        keys.add((fmt, persona_number))
    return keys


def filter_copy_json_for_selected_ads(copy_json: dict[str, Any], selected_keys: set[tuple[str, int | None]]) -> dict[str, Any]:
    ads = copy_json.get("ads")
    if not isinstance(ads, list):
        return copy_json
    selected_ads: list[dict[str, Any]] = []
    for ad in ads:
        if not isinstance(ad, dict):
            continue
        fmt = str(ad.get("format") or "").strip().upper()
        persona_number = None
        persona = ad.get("persona")
        if isinstance(persona, dict) and isinstance(persona.get("number"), int):
            persona_number = int(persona.get("number"))
        if (fmt, persona_number) in selected_keys or (fmt, None) in selected_keys:
            selected_ads.append(ad)
    cloned = json.loads(json.dumps(copy_json, ensure_ascii=False))
    cloned["ads"] = selected_ads
    return cloned


def api_run_generate_images_45(run_id: str, payload: dict[str, Any] = Body(...)) -> dict[str, Any]:
    run_dir, manifest, has_storage_manifest = load_manifest_for_run(run_id)
    batch = str(manifest.get("batch") or "").strip()
    if not batch:
        raise HTTPException(status_code=400, detail="Run has no batch folder")

    prompt_files = payload.get("prompt_files")
    if not isinstance(prompt_files, list) or not prompt_files:
        raise HTTPException(status_code=400, detail="prompt_files must be a non-empty array")

    selected_45 = validate_selected_45_prompts(batch, prompt_files)
    if not selected_45:
        raise HTTPException(status_code=400, detail="No valid 4:5 prompt files selected")

    headless = bool(payload.get("headless", False))
    engine = str(payload.get("engine") or "gemini").strip().lower()
    try:
        if engine == "chatgpt":
            result = run_chatgpt_generation(
                batch=batch,
                prompt_files=selected_45,
                aspect_ratio="4:5",
                image_sources_file=None,
                headless=headless,
                run_dir=run_dir,
            )
        else:
            result = run_gemini_generation(
                batch=batch,
                prompt_files=selected_45,
                aspect_ratio="4:5",
                image_sources_file=None,
                headless=headless,
                run_dir=run_dir,
            )
    except RuntimeError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    if result.returncode != 0:
        error_text = result.stderr or result.stdout
        engine_label = "ChatGPT" if engine == "chatgpt" else "Gemini"
        log_path = RUNTIME_ROOT / "generation_logs" / f"gen_{batch}_4_5{'_chatgpt' if engine == 'chatgpt' else ''}.log"
        if run_dir is not None:
            (run_dir / "logs" / f"image_generation_45_error{'_chatgpt' if engine == 'chatgpt' else ''}.txt").write_text(error_text, encoding="utf-8")
        short_error = "\n".join([line for line in error_text.splitlines() if line.strip()][-6:])
        raise HTTPException(status_code=500, detail=f"{engine_label} image generation failed (4:5). Log: {log_path}\n{short_error}")

    if not has_storage_manifest or run_dir is None:
        refreshed = collect_backfill_result(run_id, batch)
        refreshed["generated_images_for_prompts_45"] = selected_45
        return refreshed

    refreshed = collect_run_result(run_dir, batch, True)
    refreshed["generated_images_for_prompts_45"] = selected_45
    merged = merge_manifest(run_dir, manifest, refreshed)
    return merged


def api_run_generate_images_916_from_45(run_id: str, payload: dict[str, Any] = Body(...)) -> dict[str, Any]:
    run_dir, manifest, has_storage_manifest = load_manifest_for_run(run_id)
    batch = str(manifest.get("batch") or "").strip()
    if not batch:
        raise HTTPException(status_code=400, detail="Run has no batch folder")

    prompt_files = payload.get("prompt_files")
    if not isinstance(prompt_files, list) or not prompt_files:
        raise HTTPException(status_code=400, detail="prompt_files must be a non-empty array")

    selected_45 = validate_selected_45_prompts(batch, prompt_files)
    if not selected_45:
        raise HTTPException(status_code=400, detail="No valid 4:5 prompt files for 9:16 generation")

    selected_keys = extract_selected_ad_keys_from_45_prompts(selected_45)
    all_jobs = collect_45_reference_jobs_for_batch(batch)
    selected_jobs = [
        job
        for job in all_jobs
        if (job["format"], int(job["persona_number"])) in selected_keys or (job["format"], None) in selected_keys
    ]
    if not selected_jobs:
        raise HTTPException(status_code=400, detail="No usable 4:5 reference images matched selected prompts")

    headless = bool(payload.get("headless", False))
    engine = str(payload.get("engine") or "gemini").strip().lower()
    if engine not in {"gemini", "chatgpt"}:
        raise HTTPException(status_code=400, detail="engine must be gemini or chatgpt")
    result = run_916_conversion_from_45_for_batch(batch=batch, headless=headless, run_dir=run_dir, engine=engine, jobs=selected_jobs)

    if not has_storage_manifest or run_dir is None:
        refreshed = collect_backfill_result(run_id, batch)
        refreshed["generated_images_for_prompts_916"] = result.get("prompt_files_used", [])
        refreshed["generated_variant"] = "9:16"
        refreshed["conversion_failures"] = result.get("failures", [])
        return refreshed

    refreshed = collect_run_result(run_dir, batch, True)
    refreshed["generated_images_for_prompts_916"] = result.get("prompt_files_used", [])
    refreshed["generated_variant"] = "9:16"
    refreshed["conversion_failures"] = result.get("failures", [])
    merged = merge_manifest(run_dir, manifest, refreshed)
    return merged


def api_batch_generate_images_45(payload: dict[str, Any] = Body(...)) -> dict[str, Any]:
    run_ids = payload.get("run_ids")
    if not isinstance(run_ids, list) or not run_ids:
        raise HTTPException(status_code=400, detail="run_ids must be a non-empty array")

    all_prompt_files: list[str] = []
    run_info: list[dict[str, Any]] = []

    primary_run_dir: Path | None = None
    for run_id in run_ids:
        try:
            run_dir, manifest, has_storage_manifest = load_manifest_for_run(run_id)
        except HTTPException:
            continue
        batch = str(manifest.get("batch") or "").strip()
        if not batch:
            continue
        prompt_files_all = manifest.get("prompt_files") or []
        prompt_files_45 = [path for path in prompt_files_all if "/45/" in str(path)]
        if not prompt_files_45:
            continue
        all_prompt_files.extend(prompt_files_45)
        if has_storage_manifest and run_dir is not None and primary_run_dir is None:
            primary_run_dir = run_dir
        run_info.append({
            "run_id": run_id,
            "batch": batch,
            "prompt_count": len(prompt_files_45),
        })

    if not all_prompt_files:
        raise HTTPException(status_code=400, detail="No 4:5 prompt files found for any run")

    batch_names = sorted({r["batch"] for r in run_info})
    batch_name = batch_names[0] if len(batch_names) == 1 else "_".join(batch_names)
    work_id = f"{int(time.time())}_{uuid.uuid4().hex[:8]}"
    engine = str(payload.get("engine") or "gemini").strip().lower()
    engine_label = "ChatGPT" if engine == "chatgpt" else "Gemini"
    prompt_work_dir = RUNTIME_ROOT / f"{engine.lower()}_selected_prompts" / f"{batch_name}_{work_id}"
    prompt_work_dir.mkdir(parents=True, exist_ok=True)
    starting_prompt = ""
    starting_prompt_path = ROOT / "input" / "startingprompt.txt"
    if starting_prompt_path.exists():
        starting_prompt = starting_prompt_path.read_text(encoding="utf-8").strip()
    prompt_files_created: list[str] = []
    for src_pf in all_prompt_files:
        src = Path(src_pf)
        if not src.is_absolute():
            src = ROOT / src
        src = src.resolve()
        if not src.exists():
            continue
        prompt_text = src.read_text(encoding="utf-8")
        combined = f"{starting_prompt}\n\n{prompt_text.strip()}\n" if starting_prompt else prompt_text
        dest = prompt_work_dir / src.name
        dest.write_text(combined, encoding="utf-8")
        sidecar = src.with_suffix(".json")
        if sidecar.exists():
            (prompt_work_dir / sidecar.name).write_text(sidecar.read_text(encoding="utf-8"), encoding="utf-8")
        prompt_files_created.append(str(dest))
    headless = bool(payload.get("headless", False))
    out_dir = GENERATED_IMAGES_ROOT / batch_name / "4_5"
    out_dir.mkdir(parents=True, exist_ok=True)

    if engine == "chatgpt":
        cmd = [
            sys.executable,
            "scripts/chatgpt_web_sutomation.py",
            "--prompt-dir",
            str(prompt_work_dir),
            "--prompt-glob",
            "*.txt",
            "--out-dir",
            str(out_dir),
            "--timeout",
            str(int(os.getenv("CHATGPT_GENERATION_TIMEOUT_SECONDS") or "420")),
            "--download-timeout",
            str(int(os.getenv("CHATGPT_DOWNLOAD_TIMEOUT_SECONDS") or "90")),
            "--manual-login-timeout",
            str(int(os.getenv("CHATGPT_MANUAL_LOGIN_TIMEOUT_SECONDS") or "180")),
            "--upload-dir",
            str(INPUT_IMAGES_DIR),
        ]
    else:
        cmd = [
            sys.executable,
            "scripts/gemini_web_automation.py",
            "--prompt-dir",
            str(prompt_work_dir),
            "--prompt-glob",
            "*.txt",
            "--out-dir",
            str(out_dir),
            "--timeout",
            str(int(os.getenv("GEMINI_GENERATION_TIMEOUT_SECONDS") or "420")),
            "--manual-login-timeout",
            str(int(os.getenv("GEMINI_MANUAL_LOGIN_TIMEOUT_SECONDS") or "180")),
            "--upload-dir",
            str(INPUT_IMAGES_DIR),
        ]
    if headless:
        cmd.append("--headless")

    result = run_cmd(cmd, cwd=ROOT)
    if result.returncode != 0:
        error_text = result.stderr or result.stdout
        short_error = "\n".join([line for line in error_text.splitlines() if line.strip()][-30:])
        raise HTTPException(status_code=500, detail=f"Batch 4:5 generation failed ({engine_label}):\n{short_error}")

    return {
        "status": "completed",
        "batch_key": batch_name,
        "total_prompts": len(prompt_files_created),
        "run_count": len(run_ids),
    }


def _resolve_916_generation_for_run(run_dir: Path, manifest: dict[str, Any]) -> list[dict[str, Any]]:
    """For a single run, build the list of {prompt_96, image_sources} entries for 9:16 generation.

    Checks manifest for existing 9:16 prompt files, falls back to deriving from 4:5 prompts.
    Uses load_batch_image_summary to find existing 4:5 images as references.
    """
    batch = (manifest.get("batch") or "").strip()
    if not batch:
        return []

    prompt_files_all = manifest.get("prompt_files") or []

    # First try to use existing 9:16 prompt files from the manifest
    prompt_files_96 = [p for p in prompt_files_all if "/96/" in str(p)]

    if prompt_files_96:
        # We have 9:16 prompts already; find their corresponding 4:5 images
        image_summary = load_batch_image_summary(batch)
        prompt_to_images: dict[str, list[str]] = {}
        for entry in image_summary:
            pf = entry.get("prompt_file") or ""
            saved = entry.get("saved_files") or []
            if pf and saved:
                prompt_to_images[pf] = saved

        entries: list[dict[str, Any]] = []
        for pf96 in prompt_files_96:
            rel_96 = str(pf96).replace("\\", "/")
            parsed = parse_prompt_filename(rel_96)
            if not parsed:
                continue
            fmt, lang, persona_num = parsed

            # Look for 4:5 image by matching format+persona
            base_name = f"p{persona_num:02d}"
            image_sources: list[str] = []
            for pf45, imgs in prompt_to_images.items():
                if f"OUTPUT_{fmt}_P{persona_num:02d}" in str(pf45).upper():
                    image_sources = list(imgs)
                    break

            # Fallback: search image roots directly for the 4:5 image
            if not image_sources:
                for img_root in generated_image_roots():
                    ref_dir = img_root / batch / "4_5"
                    if not ref_dir.exists():
                        continue
                    for ext in ("png", "jpg", "jpeg", "webp"):
                        for f in sorted(ref_dir.glob(f"**/*{base_name}*.{ext}")):
                            rel = str(f.relative_to(ROOT))
                            if rel not in image_sources:
                                image_sources.append(rel)
                        if image_sources:
                            break
                    if image_sources:
                        break

            if not image_sources:
                continue

            pf96_path = f"output/{batch}/96/{Path(pf96).name}"
            entries.append({
                "prompt_96": pf96_path,
                "image_sources": image_sources,
            })
        return entries

    # Fallback: derive 9:16 prompts from 4:5 prompts (if 96 outputs exist on disk)
    prompt_files_45 = [p for p in prompt_files_all if "/45/" in str(p)]
    image_summary = load_batch_image_summary(batch)
    prompt_to_images: dict[str, list[str]] = {}
    for entry in image_summary:
        pf = entry.get("prompt_file") or ""
        saved = entry.get("saved_files") or []
        if pf and saved:
            prompt_to_images[pf] = saved

    entries = []
    for pf in prompt_files_45:
        rel_45 = str(pf).replace("\\", "/")
        parsed = parse_prompt_filename(rel_45)
        if not parsed:
            continue
        fmt, lang, persona_num = parsed

        # 9:16 prompt expected at output/{batch}/96/
        pf_filename = f"OUTPUT_{fmt}_P{persona_num:02d}_{lang}.txt"
        prompt_96 = f"output/{batch}/96/{pf_filename}"
        prompt_96_path = ROOT / prompt_96
        if not prompt_96_path.exists():
            continue

        image_sources = list(prompt_to_images.get(rel_45, []))

        # Fallback: search image roots directly
        if not image_sources:
            base_name = f"p{persona_num:02d}"
            for img_root in generated_image_roots():
                ref_dir = img_root / batch / "4_5"
                if not ref_dir.exists():
                    continue
                for ext in ("png", "jpg", "jpeg", "webp"):
                    for f in sorted(ref_dir.glob(f"**/*{base_name}*.{ext}")):
                        rel = str(f.relative_to(ROOT))
                        if rel not in image_sources:
                            image_sources.append(rel)
                        break
                    if image_sources:
                        break
                if image_sources:
                    break

        if not image_sources:
            continue

        entries.append({
            "prompt_96": prompt_96,
            "image_sources": image_sources,
        })

    return entries


def run_916_conversion_from_45_for_batch(
    *,
    batch: str,
    headless: bool,
    run_dir: Path | None,
    engine: str = "gemini",
    jobs: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    resolved_jobs = jobs if isinstance(jobs, list) else collect_45_reference_jobs_for_batch(batch)
    if not resolved_jobs:
        raise HTTPException(status_code=400, detail=f"No usable 4:5 reference images found for batch {batch}")

    template_path = ensure_916_conversion_template()
    template_text = template_path.read_text(encoding="utf-8").strip()
    prompt_root = RUNTIME_ROOT / "conversion_916_prompts" / f"{batch}_{int(time.time())}_{uuid.uuid4().hex[:8]}"
    prompt_root.mkdir(parents=True, exist_ok=True)

    failures: list[str] = []
    completed = 0
    prompt_files_used: list[str] = []

    for index, job in enumerate(resolved_jobs, start=1):
        prompt_name = build_916_conversion_prompt_job(job["format"], int(job["persona_number"]), job["language"], index)
        prompt_path = prompt_root / prompt_name
        prompt_path.write_text(template_text + "\n", encoding="utf-8")

        source_file = prompt_root / f"{prompt_path.stem}.images.txt"
        source_file.write_text(str(job["image_abs"]) + "\n", encoding="utf-8")

        if engine == "chatgpt":
            result = run_chatgpt_generation(
                batch=batch,
                prompt_files=[str(prompt_path)],
                aspect_ratio="9:16",
                image_sources_file=str(source_file),
                headless=headless,
                run_dir=run_dir,
                prepend_starting_prompt=False,
            )
        else:
            result = run_gemini_generation(
                batch=batch,
                prompt_files=[str(prompt_path)],
                aspect_ratio="9:16",
                image_sources_file=str(source_file),
                headless=headless,
                run_dir=run_dir,
                prepend_starting_prompt=False,
            )

        if result.returncode != 0:
            failures.append(f"{prompt_name}: {(result.stderr or result.stdout or '').strip()[:300]}")
            continue

        completed += 1
        prompt_files_used.append(str(prompt_path))

    if completed == 0:
        short = "\n".join(failures[:3])
        engine_label = "ChatGPT" if engine == "chatgpt" else "Gemini"
        raise HTTPException(status_code=500, detail=f"9:16 conversion failed for batch {batch} ({engine_label}). {short}")

    return {
        "batch": batch,
        "completed": completed,
        "attempted": len(resolved_jobs),
        "failures": failures,
        "prompt_files_used": prompt_files_used,
    }


def api_batch_generate_images_916(payload: dict[str, Any] = Body(...)) -> dict[str, Any]:
    run_ids = payload.get("run_ids")
    if not isinstance(run_ids, list) or not run_ids:
        raise HTTPException(status_code=400, detail="run_ids must be a non-empty array")

    headless = bool(payload.get("headless", False))
    engine = str(payload.get("engine") or "gemini").strip().lower()
    if engine not in {"gemini", "chatgpt"}:
        raise HTTPException(status_code=400, detail="engine must be gemini or chatgpt")
    batch_to_run_dir: dict[str, Path | None] = {}
    for run_id in run_ids:
        try:
            run_dir, manifest, has_storage_manifest = load_manifest_for_run(run_id)
        except HTTPException:
            continue
        batch = str(manifest.get("batch") or "").strip()
        if not batch:
            continue
        if has_storage_manifest and run_dir is not None:
            batch_to_run_dir[batch] = run_dir
        elif batch not in batch_to_run_dir:
            batch_to_run_dir[batch] = None

    if not batch_to_run_dir:
        raise HTTPException(status_code=400, detail="No valid batches found for selected runs")

    total_attempted = 0
    total_completed = 0
    processed_batches: list[str] = []
    batch_errors: list[str] = []

    for batch, run_dir in sorted(batch_to_run_dir.items()):
        try:
            result = run_916_conversion_from_45_for_batch(batch=batch, headless=headless, run_dir=run_dir, engine=engine)
        except HTTPException as exc:
            batch_errors.append(f"{batch}: {exc.detail}")
            continue

        processed_batches.append(batch)
        total_attempted += int(result.get("attempted") or 0)
        total_completed += int(result.get("completed") or 0)

        if run_dir is not None:
            manifest_path = run_dir / "manifest.json"
            if manifest_path.exists():
                manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
                refreshed = collect_run_result(run_dir, batch, True)
                refreshed["generated_variant"] = "9:16"
                refreshed["generated_images_for_prompts_916"] = result.get("prompt_files_used", [])
                merge_manifest(run_dir, manifest, refreshed)

    if total_completed == 0:
        detail = "No 9:16 conversions succeeded"
        if batch_errors:
            detail += ": " + " | ".join(batch_errors[:3])
        raise HTTPException(status_code=400, detail=detail)

    return {
        "status": "completed",
        "batch_key": ",".join(processed_batches),
        "total_prompts": total_completed,
        "attempted_prompts": total_attempted,
        "run_count": len(processed_batches),
        "errors": batch_errors,
    }


def extract_persona_input_block(prompt_text: str) -> str:
    markers = ["EXACT ON-IMAGE COPY", "PERSONA INPUT", "PERSONA:", "INPUT:"]
    for marker in markers:
        if marker in prompt_text.upper():
            start = prompt_text.upper().find(marker)
            if start != -1:
                return prompt_text[start:].strip()
    if len(prompt_text) > 50:
        return prompt_text.strip()
    return ""


def api_file_content(path: str, max_lines: int = 400) -> dict[str, Any]:
    file_path = resolve_safe_path(path)
    if not file_path.exists() or not file_path.is_file():
        raise HTTPException(status_code=404, detail="File not found")
    lines = file_path.read_text(encoding="utf-8", errors="ignore").splitlines()
    clipped = lines[:max_lines]
    return {
        "path": str(file_path.relative_to(ROOT)),
        "total_lines": len(lines),
        "shown_lines": len(clipped),
        "content": "\n".join(clipped),
    }


async def api_run_execute(
    config: str = Form(...),
    product_info_file: UploadFile | None = File(None),
    mechanism_file: UploadFile | None = File(None),
    faq_file: UploadFile | None = File(None),
    image_source_file: UploadFile | None = File(None),
    input_image_files: list[UploadFile] | None = File(None),
    clear_input_images: bool = Form(False),
) -> dict[str, Any]:
    ensure_dirs()
    batch = "v0"
    try:
        cfg = json.loads(config)
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail="Invalid config JSON") from exc
    if str((cfg.get("server_type") or "opencode")).strip().lower() != "opencode":
        raise HTTPException(status_code=400, detail="Unsupported server type. Use OpenCode.")

    run_id = make_run_id()
    run_dir = RUNS_ROOT / run_id
    (run_dir / "inputs").mkdir(parents=True, exist_ok=True)
    (run_dir / "logs").mkdir(parents=True, exist_ok=True)
    (run_dir / "context").mkdir(parents=True, exist_ok=True)

    product_path = save_upload(run_dir / "inputs" / "product master doc.txt", product_info_file)
    mechanism_path = None
    faq_path = None
    image_sources_path = save_upload(run_dir / "inputs" / "image_sources.txt", image_source_file)
    saved_input_images = store_uploaded_input_images(input_image_files or [], clear_input_images)

    product_file = coalesce_path(product_path, DEFAULT_PRODUCT_MASTER)
    mechanism_file_path = ROOT / "__empty__.txt"
    faq_file_path = ROOT / "__empty__.txt"

    image_sources_file_path = coalesce_path(image_sources_path, default_image_sources_file())

    try:
        base_plan = resolve_format_plan(cfg)
    except RuntimeError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    hypothesis_cfg = cfg.get("hypothesis") or {}
    plan = expand_plan_with_hypothesis(base_plan, hypothesis_cfg)

    reuse_visual_patterns_from_run_id = str(cfg.get("reuse_visual_patterns_from_run_id") or "").strip()
    if reuse_visual_patterns_from_run_id:
        pattern_locks = collect_visual_pattern_reuse_locks(reuse_visual_patterns_from_run_id)
        plan, applied_patterns = apply_visual_pattern_reuse_to_plan(
            plan,
            pattern_locks,
            share_across_personas=bool(cfg.get("share_background_across_personas")),
        )
        (run_dir / "context" / "visual_pattern_reuse.json").write_text(
            json.dumps(
                {
                    "source_run_id": reuse_visual_patterns_from_run_id,
                    "available_locks": len(pattern_locks),
                    "applied_ads": applied_patterns,
                    "share_across_personas": bool(cfg.get("share_background_across_personas")),
                },
                ensure_ascii=False,
                indent=2,
            )
            + "\n",
            encoding="utf-8",
        )

    # Save hypothesis config to run dir for reference
    if hypothesis_cfg:
        (run_dir / "context" / "hypothesis_config.json").write_text(
            json.dumps(hypothesis_cfg, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
        )

    product_ctx_source = "attached_product_master_doc"
    extractor_model = "none"
    (run_dir / "context" / "product_doc_source.json").write_text(
        json.dumps(
            {
                "source": product_ctx_source,
                "product_file": str(product_file),
                "note": "Canonical extraction is disabled; the full product master doc is attached to the OpenCode copy session unchanged.",
            },
            ensure_ascii=False,
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )

    persona_library = parse_persona_library(DEFAULT_PLAYBOOK)
    ads_context: list[dict[str, Any]] = []
    format_seen_counts: dict[str, int] = {}
    for item in plan:
        persona_no = item["persona"]
        fmt = item["format"]
        format_seen_counts[fmt] = format_seen_counts.get(fmt, 0) + 1
        persona_payload = build_persona_payload(persona_no, persona_library)

        format_result = run_cmd(
            [
                "python3",
                "scripts/extract_format_rules.py",
                "--playbook",
                str(DEFAULT_PLAYBOOK),
                "--format",
                fmt,
                "--json",
            ],
            cwd=ROOT,
        )
        format_payload = parse_json_stdout(format_result, f"extract_format_rules({fmt})")
        copy_req = build_copy_requirements(persona_no, fmt, format_seen_counts[fmt], run_id)

        # Inject hypothesis directive if active
        hyp_meta = item.get("hypothesis")
        concept = {}
        if isinstance(hyp_meta, dict) and hyp_meta.get("type") != "none":
            copy_req["hypothesis"] = hyp_meta
            concept = copy_req.get("concept_variation") or {}
            hyp_type = hyp_meta.get("type")
            variant = hyp_meta.get("variant")
            arch_guidance = COPY_PROMPTS.get("concept_framework", {})
            if hyp_type == "awareness_stage" and variant:
                concept["audience_stage"] = _framework_item("audience_stage", variant)
                guidance = arch_guidance.get("awareness_stage", {}).get(variant)
                if guidance:
                    concept["awareness_stage_guidance"] = guidance
            elif hyp_type == "concept_angle" and variant:
                concept["lead_angle"] = _framework_item("lead_angle", variant)
                guidance = arch_guidance.get("concept_angle", {}).get(variant)
                if guidance:
                    concept["concept_angle_guidance"] = guidance
            elif hyp_type == "concept_structure" and variant:
                concept["message_structure"] = _framework_item("message_structure", variant)
                guidance = arch_guidance.get("concept_structure", {}).get(variant)
                if guidance:
                    concept["concept_structure_guidance"] = guidance
            elif hyp_type == "hook_structure" and variant:
                concept["hook_structure_override"] = variant
                guidance = arch_guidance.get("hook_structure", {}).get(variant)
                if guidance:
                    concept["hook_structure_guidance"] = guidance
            elif hyp_type == "proof_style" and variant:
                concept["proof_style_override"] = variant
                guidance = arch_guidance.get("proof_style", {}).get(variant)
                if guidance:
                    concept["proof_style_guidance"] = guidance
            elif hyp_type == "cta_voice" and variant:
                concept["cta_voice_override"] = variant
                guidance = arch_guidance.get("cta_voice", {}).get(variant)
                if guidance:
                    concept["cta_voice_guidance"] = guidance
            copy_req["concept_variation"] = concept

            # Update headline_architecture if hypothesis changed the structural driver
            arch_src = None
            arch_variant = None
            if hyp_type == "concept_structure" and variant:
                arch_src = "concept_structure"
                arch_variant = variant
            elif hyp_type == "hook_structure" and variant:
                arch_src = "hook_structure"
                arch_variant = variant
            if arch_src and arch_variant:
                ha_group = COPY_ARCH.get("headline_architectures", {}).get(arch_src, {})
                ha_entry = ha_group.get(arch_variant)
                if ha_entry:
                    copy_req["headline_architecture"] = {
                        "template": ha_entry.get("template", ""),
                        "examples": ha_entry.get("examples", []),
                        "source": arch_src,
                        "variant": arch_variant,
                    }

        ads_context.append(
            {
                "persona": persona_payload,
                "format_rules": format_payload,
                "format": fmt,
                "copy_requirements": copy_req,
                "hypothesis": hyp_meta,
                "visual_archetype": item.get("visual_archetype"),
                "visual_pattern_reused_from_run_id": item.get("visual_pattern_reused_from_run_id"),
                "visual_pattern_reuse_key": item.get("visual_pattern_reuse_key"),
                "creative_index": item.get("creative_index", 1),
                "creative_total": item.get("creative_total", 1),
                "background_group_key": item.get("background_group_key"),
                "share_background_across_personas": item.get("share_background_across_personas", False),
            }
        )

    banlist_result = run_cmd(["python3", "scripts/registry_banlist.py", "--last", "150"], cwd=ROOT)
    banlist_payload = parse_json_stdout(banlist_result, "registry_banlist")

    full_context = {
        "generated_at": now_iso(),
        "run_id": run_id,
        "language_mode": resolve_language_mode(cfg),
        "context_source": product_ctx_source,
        "context_extractor_model": extractor_model,
        "product_file_path": str(product_file),
        "ads": ads_context,
        "banlist": banlist_payload,
    }
    (run_dir / "context" / "run_context.json").write_text(
        json.dumps(
            {
                "generated_at": full_context["generated_at"],
                "run_id": full_context["run_id"],
                "language_mode": full_context["language_mode"],
                "context_source": full_context["context_source"],
                "context_extractor_model": full_context["context_extractor_model"],
                "product_file_path": full_context["product_file_path"],
            },
            ensure_ascii=False,
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )

    llm_mode = "opencode"
    copy_json = call_opencode_compatible(cfg, full_context, run_dir)
    used_template_fallback = False
    if not copy_json:
        llm_mode = "fallback_template"
        used_template_fallback = True
        (run_dir / "logs" / "opencode_fallback.txt").write_text(
            "OpenCode copy generation unavailable; using deterministic schema-compatible fallback copy.\n",
            encoding="utf-8",
        )
        copy_json = build_template_copy(full_context, run_id)
    opencode_failures = copy_json.pop("_opencode_failures", []) if isinstance(copy_json, dict) else []
    opencode_warnings = copy_json.pop("_opencode_warnings", []) if isinstance(copy_json, dict) else []
    opencode_session_fallback = bool(copy_json.pop("_opencode_session_fallback", False)) if isinstance(copy_json, dict) else False
    if opencode_failures and llm_mode == "opencode":
        llm_mode = "opencode_partial_fallback"
        (run_dir / "logs" / "opencode_fallback.txt").write_text(
            "Some OpenCode ad generations failed; normalize_generated_copy filled those outputs with deterministic template copy.\n\n"
            + "\n\n---\n\n".join(opencode_failures),
            encoding="utf-8",
        )
    copy_json = normalize_generated_copy(copy_json, full_context, run_id)
    copy_json = strip_internal_markers_from_payload(copy_json)
    copy_json = enforce_unique_ctas(copy_json, full_context)
    copy_json = scrub_on_image_copy(copy_json)
    reuse_backgrounds_from_run_id = str(cfg.get("reuse_backgrounds_from_run_id") or "").strip()
    if reuse_backgrounds_from_run_id:
        locks = collect_background_reuse_locks(reuse_backgrounds_from_run_id)
        copy_json, applied_locks = apply_background_reuse_locks(
            copy_json,
            locks,
            share_across_personas=bool(cfg.get("share_background_across_personas")),
        )
        (run_dir / "context" / "background_reuse.json").write_text(
            json.dumps(
                {
                    "source_run_id": reuse_backgrounds_from_run_id,
                    "available_locks": len(locks),
                    "applied_ads": applied_locks,
                    "share_background_across_personas": bool(cfg.get("share_background_across_personas")),
                },
                ensure_ascii=False,
                indent=2,
            )
            + "\n",
            encoding="utf-8",
        )
    generated_copy_error = validate_generated_copy_payload(copy_json, ads_context)
    if generated_copy_error:
        (run_dir / "logs" / "opencode_error.txt").write_text(
            generated_copy_error + "\n\nGenerated payload:\n" + json.dumps(copy_json, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        raise HTTPException(status_code=502, detail="OpenCode copy generation returned incomplete copy. Prompt production stopped; check run logs.")

    copy_file = run_dir / "context" / "copy_batch.json"
    copy_file.write_text(json.dumps(copy_json, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    assembler_result = run_cmd(
        [
            "python3",
            "scripts/generate_ads.py",
            "--copy-file",
            str(copy_file),
            "--language-mode",
            assembler_language_mode(cfg),
            "--skip-uniqueness-check",
        ],
        cwd=ROOT,
    )
    if assembler_result.returncode != 0:
        assembler_error = assembler_result.stderr or assembler_result.stdout
        (run_dir / "logs" / "assembler_error.txt").write_text(assembler_error, encoding="utf-8")
        raise HTTPException(status_code=500, detail="Prompt assembly failed. Check run logs.")

    batch_match = re.search(r"Batch:\s*(v\d+)", assembler_result.stdout)
    if not batch_match:
        raise HTTPException(status_code=500, detail="Could not parse batch from assembler output")
    batch = batch_match.group(1)

    manifest = collect_run_result(run_dir, batch, image_generated=False)
    manifest["llm_mode"] = llm_mode
    if llm_mode == "fallback_template":
        manifest["copy_source"] = "deterministic fallback template"
    elif llm_mode == "opencode_partial_fallback":
        manifest["copy_source"] = f"opencode generated copy with template fallback for {len(opencode_failures)} failed ad(s)"
    else:
        manifest["copy_source"] = "opencode generated copy"
    if opencode_failures:
        manifest["copy_generation_failures"] = len(opencode_failures)
        manifest["copy_fallback_log"] = str((run_dir / "logs" / "opencode_fallback.txt").relative_to(ROOT))
    if used_template_fallback:
        manifest["copy_generation_failures"] = max(int(manifest.get("copy_generation_failures") or 0), 1)
        manifest["copy_fallback_log"] = str((run_dir / "logs" / "opencode_fallback.txt").relative_to(ROOT))
        manifest["copy_generation_notes"] = ["OpenCode copy generation unavailable; deterministic fallback copy was used."]
    if opencode_warnings:
        manifest["copy_generation_warnings"] = len(opencode_warnings)
        manifest["copy_warning_log"] = str((run_dir / "logs" / "opencode_error.txt").relative_to(ROOT))
        manifest["copy_generation_notes"] = [str(item).splitlines()[0] for item in opencode_warnings[:3]]
    if opencode_session_fallback:
        manifest["copy_session_fallback"] = True
        manifest["copy_session_log"] = str((run_dir / "logs" / "opencode_session.log").relative_to(ROOT))
    manifest["context_source"] = product_ctx_source
    manifest["context_extractor_model"] = extractor_model
    manifest["image_sources_file"] = str(image_sources_file_path)
    manifest["input_images_dir"] = str(INPUT_IMAGES_DIR.relative_to(ROOT)).replace("\\", "/")
    manifest["input_images_uploaded"] = saved_input_images
    if reuse_backgrounds_from_run_id:
        manifest["background_reuse_from_run_id"] = reuse_backgrounds_from_run_id
    if reuse_visual_patterns_from_run_id:
        manifest["visual_pattern_reuse_from_run_id"] = reuse_visual_patterns_from_run_id
    (run_dir / "manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return manifest


# Chrome process tracking
_chrome_process: subprocess.Popen | None = None


def api_launch_visible_browser() -> dict[str, Any]:
    """Launch a visible Chrome instance with CDP enabled so the user can log in
    before automation begins."""
    global _chrome_process

    chrome_bin = None
    for candidate in [
        "/usr/bin/google-chrome",
        "/usr/bin/google-chrome-stable",
    ]:
        if Path(candidate).exists():
            chrome_bin = candidate
            break

    if not chrome_bin:
        raise HTTPException(status_code=500, detail="Chrome binary not found on system")

    cmd = [
        chrome_bin,
        "--remote-debugging-port=9222",
        f"--user-data-dir={os.path.expandvars('$HOME')}/.config/google-chrome-gemini-cdp",
        "--no-first-run",
        "--no-default-browser-check",
    ]
    _chrome_process = subprocess.Popen(cmd, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    time.sleep(2)

    for attempt in range(10):
        try:
            resp = urllib.request.urlopen("http://127.0.0.1:9222/json/version", timeout=2)
            if resp.status == 200:
                return {
                    "status": "launched",
                    "cdp_url": "http://127.0.0.1:9222",
                    "message": "Chrome launched. Log in, then trigger generation.",
                }
        except Exception:
            time.sleep(1)

    raise HTTPException(status_code=500, detail="Chrome launched but CDP not responding on port 9222")


def api_kill_chrome() -> dict[str, Any]:
    """Kill the Chrome process started by launch-visible-browser and stop any running automation."""
    global _chrome_process
    killed = False
    if _chrome_process and _chrome_process.poll() is None:
        try:
            _chrome_process.terminate()
            try:
                _chrome_process.wait(timeout=5)
            except subprocess.TimeoutExpired:
                _chrome_process.kill()
                _chrome_process.wait(timeout=3)
            killed = True
        except Exception:
            pass
        _chrome_process = None

    # Also kill any running gemini automation processes
    gemini_killed = 0
    for proc in psutil.process_iter(["pid", "name", "cmdline"]):
        try:
            cmdline = proc.info.get("cmdline") or []
            if any("gemini_web_automation" in c for c in cmdline):
                proc.kill()
                gemini_killed += 1
        except (psutil.NoSuchProcess, psutil.AccessDenied):
            continue

    # Also kill any running chatgpt automation processes
    chatgpt_killed = 0
    for proc in psutil.process_iter(["pid", "name", "cmdline"]):
        try:
            cmdline = proc.info.get("cmdline") or []
            if any("chatgpt_web_sutomation" in c for c in cmdline):
                proc.kill()
                chatgpt_killed += 1
        except (psutil.NoSuchProcess, psutil.AccessDenied):
            continue

    return {"status": "killed", "chrome": killed, "gemini_processes": gemini_killed, "chatgpt_processes": chatgpt_killed}


def api_edit_prompt(run_id: str, payload: dict[str, Any] = Body(...)) -> dict[str, Any]:
    """Edit a prompt file in-place, replacing only the EXACT ON-IMAGE COPY block."""
    run_dir = RUNS_ROOT / run_id
    prompt_path = payload.get("prompt_file", "")
    new_text = payload.get("text", "")
    if not prompt_path or not new_text:
        raise HTTPException(status_code=400, detail="prompt_file and text are required")

    full_path = ROOT / prompt_path
    if not full_path.exists():
        raise HTTPException(status_code=404, detail="Prompt file not found")

    old_text = full_path.read_text(encoding="utf-8")
    updated_text = _replace_exact_copy_block(old_text, new_text)
    if updated_text is None:
        raise HTTPException(status_code=400, detail="No EXACT ON-IMAGE COPY block found in prompt file")
    full_path.write_text(updated_text, encoding="utf-8")
    return {"status": "saved", "prompt_file": prompt_path}


def api_delete_prompt(run_id: str, payload: dict[str, Any] = Body(...)) -> dict[str, Any]:
    """Delete a prompt file and remove it from the run manifest."""
    run_dir = RUNS_ROOT / run_id
    manifest_path = run_dir / "manifest.json"
    prompt_path = payload.get("prompt_file", "")
    if not prompt_path:
        raise HTTPException(status_code=400, detail="prompt_file is required")

    full_path = ROOT / prompt_path
    if full_path.exists():
        full_path.unlink()

    if manifest_path.exists():
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        manifest["prompt_files"] = [p for p in manifest.get("prompt_files", []) if p != prompt_path]
        manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    return {"status": "deleted", "prompt_file": prompt_path}


def api_delete_image(run_id: str, payload: dict[str, Any] = Body(...)) -> dict[str, Any]:
    """Delete a generated image and its metadata JSON."""
    run_dir = RUNS_ROOT / run_id
    manifest_path = run_dir / "manifest.json"
    image_path = payload.get("image_file", "")
    if not image_path:
        raise HTTPException(status_code=400, detail="image_file is required")

    full_path = ROOT / image_path
    if full_path.exists():
        full_path.unlink()

    # Also delete companion JSON metadata if it exists
    for json_path in (full_path.with_suffix(".json"), full_path.with_suffix(full_path.suffix + ".json")):
        if json_path.exists():
            json_path.unlink()

    if manifest_path.exists():
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        manifest["image_files"] = [p for p in manifest.get("image_files", []) if p != image_path]
        manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    return {"status": "deleted", "image_file": image_path}


async def api_replace_image(run_id: str, image_file: str = Form(...), replacement_file: UploadFile = File(...)) -> dict[str, Any]:
    run_dir = RUNS_ROOT / run_id
    full_path = resolve_safe_path(image_file)
    generated_root = GENERATED_IMAGES_ROOT.resolve()
    if generated_root not in full_path.resolve().parents:
        raise HTTPException(status_code=400, detail="image_file must be under generated_images")
    if not full_path.exists() or not full_path.is_file():
        raise HTTPException(status_code=404, detail="Generated image not found")

    allowed = {".png", ".jpg", ".jpeg", ".webp"}
    upload_name = Path(replacement_file.filename or "").name
    upload_ext = Path(upload_name).suffix.lower()
    if upload_ext not in allowed:
        raise HTTPException(status_code=400, detail="Replacement must be png, jpg, jpeg, or webp")

    data = await replacement_file.read()
    if not data:
        raise HTTPException(status_code=400, detail="Replacement file is empty")
    full_path.write_bytes(data)

    meta_path = full_path.with_suffix(".json")
    if meta_path.exists():
        try:
            meta = json.loads(meta_path.read_text(encoding="utf-8"))
        except Exception:
            meta = {}
    else:
        meta = {"type": "ad_image", "status": "success", "saved_file": str(full_path)}
    replacements = meta.setdefault("replacements", [])
    if isinstance(replacements, list):
        replacements.append(
            {
                "timestamp": int(time.time()),
                "source_filename": upload_name,
                "size_bytes": len(data),
            }
        )
    meta["status"] = "success"
    meta["saved_file"] = str(full_path)
    meta["replaced"] = True
    meta["replacement_timestamp"] = int(time.time())
    meta["replacement_source_filename"] = upload_name
    meta_path.write_text(json.dumps(meta, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    manifest_path = run_dir / "manifest.json"
    if manifest_path.exists():
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        manifest.setdefault("image_files", [])
        if image_file not in manifest["image_files"]:
            manifest["image_files"].append(image_file)
        manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    return {"status": "replaced", "image_file": image_file, "size_bytes": len(data)}


def _parse_image_naming(image_path_str: str, run_dir: Path | None) -> dict[str, str]:
    """Extract format, persona, language from an image's companion JSON metadata
    and build a human-readable stem for download naming."""
    full_path = ROOT / image_path_str
    meta_path = full_path.with_suffix(".json")
    legacy_meta_path = full_path.with_suffix(full_path.suffix + ".json")
    base = {"format": "UNKNOWN", "persona": "00", "lang": "EN", "stem": "image"}
    hyp_label = ""

    if meta_path.exists() or legacy_meta_path.exists():
        try:
            if not meta_path.exists() and legacy_meta_path.exists():
                meta_path = legacy_meta_path
            meta = json.loads(meta_path.read_text(encoding="utf-8"))
        except Exception:
            meta = {}
        fmt_value = str(meta.get("format") or meta.get("format_id") or "").strip().upper()
        persona_value = str(meta.get("persona") or meta.get("persona_id") or "").strip().upper()
        lang_value = str(meta.get("language") or meta.get("lang") or meta.get("lang_id") or "").strip().upper()
        if fmt_value:
            base["format"] = fmt_value
        persona_match = re.search(r"P?(\d+)", persona_value)
        if persona_match:
            base["persona"] = f"P{int(persona_match.group(1)):02d}"
        if lang_value:
            base["lang"] = lang_value
        prompt_file = str(meta.get("prompt_file_relative") or meta.get("prompt_file") or "").strip().replace("\\", "/")
        if not prompt_file:
            prompt_file = str(meta.get("prompt_file_relative") or meta.get("prompt_file") or "").strip().replace("\\", "/")
        parsed = parse_prompt_filename(prompt_file)
        if parsed:
            fmt, lang, persona_num = parsed
            base["format"] = fmt
            base["persona"] = f"P{persona_num:02d}" if persona_num else "P00"
            base["lang"] = lang
        creative_total = int(meta.get("creative_total") or 1) if str(meta.get("creative_total") or "1").isdigit() else 1
        creative_index = int(meta.get("creative_index") or 1) if str(meta.get("creative_index") or "1").isdigit() else 1
        if creative_total > 1:
            base["creative_suffix"] = f"_A{creative_index:02d}"

        if not hyp_label:
            htype = str(meta.get("hypothesis_type") or "")
            hvar = str(meta.get("hypothesis_variant") or "")
            if htype and htype != "none":
                parts = [htype]
                if hvar:
                    parts.append(hvar)
                hyp_label = "_" + "_".join(parts)

    if base["format"] == "UNKNOWN" or base["persona"] in {"00", "P00"}:
        name = Path(image_path_str).stem.lower()
        match = re.search(r"(?:gemini|chatgpt)-(?P<fmt>[a-z0-9]+)-p(?P<num>\d+)-(?P<lang>[a-z0-9]+)(?:-a(?P<creative>\d+))?", name)
        if match:
            base["format"] = match.group("fmt").upper()
            base["persona"] = f"P{int(match.group('num')):02d}"
            base["lang"] = match.group("lang").upper()
            if match.group("creative"):
                base["creative_suffix"] = f"_A{int(match.group('creative')):02d}"

    # Try hypothesis
    if run_dir is not None:
        hyp_path = run_dir / "context" / "hypothesis_config.json"
        if hyp_path.exists():
            try:
                hyp_cfg = json.loads(hyp_path.read_text(encoding="utf-8"))
                htype = hyp_cfg.get("type", "")
                hvar = hyp_cfg.get("variant", "")
                if htype and htype != "none":
                    parts = [htype]
                    if hvar:
                        parts.append(hvar)
                    hyp_label = "_" + "_".join(parts)
            except Exception:
                pass

    ext = Path(image_path_str).suffix
    stem = f"{base['format']}_{base['persona']}_{base['lang']}{base.get('creative_suffix', '')}{hyp_label}"
    base["stem"] = stem
    base["ext"] = ext
    return base


def _build_persona_name_map(run_dir: Path) -> dict[str, str]:
    """Map persona number (P01) to persona name from run's copy_batch.json."""
    copy_path = run_dir / "context" / "copy_batch.json"
    if not copy_path.exists():
        return {}
    try:
        data = json.loads(copy_path.read_text(encoding="utf-8"))
        ads = data.get("ads") if isinstance(data, dict) else []
        if not isinstance(ads, list):
            return {}
        mapping: dict[str, str] = {}
        for ad in ads:
            p = ad.get("persona") if isinstance(ad, dict) else {}
            if not isinstance(p, dict):
                continue
            num = p.get("number")
            name = p.get("name") or p.get("persona_name") or ""
            if isinstance(num, int) and name:
                mapping[f"P{num:02d}"] = str(name)
        return mapping
    except Exception:
        return {}


def _clean_metadata_for_download(meta: dict[str, Any], img_path: str, run_dir: Path | None) -> dict[str, Any]:
    """Strip excessive internal keys from image metadata and enrich with
    hypothesis info, persona name, and clean format labels for download ZIP."""
    clean = dict(meta)

    # Strip internal plumbing
    for key in ("generated_image_src", "saved_ext", "output_dir", "metadata_file", "type"):
        clean.pop(key, None)

    # Normalise key names
    if "format" not in clean and "format_id" in clean:
        clean["format"] = clean.pop("format_id")
    if clean.get("format_id"):
        clean.pop("format_id", None)
    if "persona" not in clean and "persona_id" in clean:
        clean["persona"] = clean.pop("persona_id")
    if clean.get("persona_id"):
        clean.pop("persona_id", None)
    if "language" not in clean and "lang_id" in clean:
        clean["language"] = clean.pop("lang_id")
    if clean.get("lang_id"):
        clean.pop("lang_id", None)

    # Ensure hypothesis keys are always present
    hyp_type = clean.get("hypothesis_type") or ""
    hyp_var = clean.get("hypothesis_variant") or ""
    clean["hypothesis_type"] = hyp_type
    clean["hypothesis_variant"] = hyp_var

    # Enrich with persona name if we have a run_dir
    if run_dir is not None:
        persona_val = clean.get("persona", "")
        if persona_val:
            mapping = _build_persona_name_map(run_dir)
            if persona_val in mapping:
                clean["persona_name"] = mapping[persona_val]

    return clean


def api_download_single_image(run_id: str, image_file: str):
    """Return a zip containing the image file and its metadata JSON."""
    run_dir = RUNS_ROOT / run_id
    full_path = ROOT / image_file
    if not full_path.exists():
        raise HTTPException(status_code=404, detail="Image file not found")

    naming = _parse_image_naming(image_file, run_dir)
    meta_path = full_path.with_suffix(".json")
    legacy_meta_path = full_path.with_suffix(full_path.suffix + ".json")

    import io, zipfile
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.write(full_path, f"{naming['stem']}{naming['ext']}")
        meta_content = {"source": image_file}
        if meta_path.exists() or legacy_meta_path.exists():
            try:
                if not meta_path.exists() and legacy_meta_path.exists():
                    meta_path = legacy_meta_path
                meta_content = json.loads(meta_path.read_text(encoding="utf-8"))
            except Exception:
                pass
        meta_content = _clean_metadata_for_download(meta_content, image_file, run_dir)
        meta_content["_download_name"] = naming["stem"]
        zf.writestr(f"{naming['stem']}_metadata.json", json.dumps(meta_content, ensure_ascii=False, indent=2))

    buf.seek(0)
    return StreamingResponse(buf, media_type="application/zip",
                             headers={"Content-Disposition": f'attachment; filename="{naming["stem"]}.zip"'})


def api_download_batch_images(run_id: str):
    """Return a zip of all images grouped by VN subfolders with metadata.
    Always scans the filesystem directly so newly generated images
    (e.g. 9:16 added after the manifest was saved) are included."""
    run_dir = RUNS_ROOT / run_id

    # Refresh cached thumbnail summary before scanning
    batch_label = run_id
    manifest_path = run_dir / "manifest.json"
    if manifest_path.exists():
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        batch_label = manifest.get("batch", run_id)

    # Always scan the filesystem — manifest may be stale
    image_files = scan_image_files_for_batch(batch_label) if batch_label != run_id else []

    import io, zipfile
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        vns: set[str] = set()
        for img_path in image_files:
            full_path = ROOT / img_path
            if not full_path.exists():
                continue

            vn = _extract_vn_from_image_path(img_path) or batch_label or "images"
            vns.add(vn)
            aspect = _extract_aspect_from_image_path(img_path)
            naming = _parse_image_naming(img_path, run_dir)
            meta_path = full_path.with_suffix(".json")
            legacy_meta_path = full_path.with_suffix(full_path.suffix + ".json")

            folder = f"{vn}/{aspect}" if aspect else vn
            zf.write(full_path, f"{folder}/{naming['stem']}{naming['ext']}")

            meta_content = {"source": img_path}
            if meta_path.exists() or legacy_meta_path.exists():
                try:
                    if not meta_path.exists() and legacy_meta_path.exists():
                        meta_path = legacy_meta_path
                    meta_content = json.loads(meta_path.read_text(encoding="utf-8"))
                except Exception:
                    pass
            meta_content = _clean_metadata_for_download(meta_content, img_path, run_dir)
            meta_content["_download_name"] = naming["stem"]
            zf.writestr(f"{folder}/{naming['stem']}_metadata.json",
                        json.dumps(meta_content, ensure_ascii=False, indent=2))

        if not image_files:
            zf.writestr("README.txt",
                        "No generated images found for this run.\n"
                        "Run image generation first, then try again.")

    buf.seek(0)
    label = "_".join(sorted(vns)) if vns else (batch_label if batch_label != run_id else run_id)
    return StreamingResponse(buf, media_type="application/zip",
                             headers={"Content-Disposition": f'attachment; filename="batch_{label}.zip"'})


def api_download_batches(batch_names: list[str]):
    """Return a zip of all images for given batch names, grouped by VN folder."""
    image_files_by_vn: dict[str, list[str]] = {}
    for batch_name in batch_names:
        files = scan_image_files_for_batch(batch_name)
        if files:
            image_files_by_vn[batch_name] = files

    import io, zipfile
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for vn, files in image_files_by_vn.items():
            for img_path in files:
                full_path = ROOT / img_path
                if not full_path.exists():
                    continue
                naming = _parse_image_naming(img_path, None)
                aspect = _extract_aspect_from_image_path(img_path)
                meta_path = full_path.with_suffix(".json")
                legacy_meta_path = full_path.with_suffix(full_path.suffix + ".json")
                folder = f"{vn}/{aspect}" if aspect else vn
                zf.write(full_path, f"{folder}/{naming['stem']}{naming['ext']}")
                meta_content = {"source": img_path}
                if meta_path.exists() or legacy_meta_path.exists():
                    try:
                        if not meta_path.exists() and legacy_meta_path.exists():
                            meta_path = legacy_meta_path
                        meta_content = json.loads(meta_path.read_text(encoding="utf-8"))
                    except Exception:
                        pass
                meta_content = _clean_metadata_for_download(meta_content, img_path, None)
                meta_content["_download_name"] = naming["stem"]
                zf.writestr(f"{folder}/{naming['stem']}_metadata.json",
                            json.dumps(meta_content, ensure_ascii=False, indent=2))

        if not image_files_by_vn:
            zf.writestr("README.txt",
                        "No generated images found for selected batch(es).\n"
                        "Run image generation first, then try again.")

    buf.seek(0)
    label = "_".join(batch_names) if batch_names else "batches"
    return StreamingResponse(buf, media_type="application/zip",
                             headers={"Content-Disposition": f'attachment; filename="{label}.zip"'})


# ── Modular routes ───────────────────────────────────────────────────────────
from dashboard.backend.routes import defaults, progress, runs, generate, batch, export_import, execute, chrome

app.include_router(defaults.router)
app.include_router(progress.router)
app.include_router(runs.router)
app.include_router(generate.router)
app.include_router(batch.router)
app.include_router(export_import.router)
app.include_router(execute.router)
app.include_router(chrome.router)

app.mount("/storage", StaticFiles(directory=str(STORAGE_ROOT)), name="storage")
app.mount("/output", StaticFiles(directory=str(ROOT / "output")), name="output")
app.mount("/input", StaticFiles(directory=str(ROOT / "input")), name="input")
GENERATED_IMAGES_ROOT.mkdir(parents=True, exist_ok=True)
app.mount("/generated_images", StaticFiles(directory=str(GENERATED_IMAGES_ROOT)), name="generated_images")
app.mount("/", StaticFiles(directory=str(ROOT / "dashboard" / "frontend"), html=True), name="frontend")
